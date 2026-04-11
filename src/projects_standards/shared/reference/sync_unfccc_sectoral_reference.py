# Objetivo do script:
# Adicionar ou atualizar no workbook de referencias as tabelas de escopos setoriais e areas tecnicas da UNFCCC.
# Processo:
# 1. Definir URL do documento fonte UNFCCC e nomes de tabelas de referencia.
# 2. Embutir dados fixos de SECTORAL_SCOPES (18 escopos por A6.4-STAN-ACCR-001).
# 3. Criar/atualizar tabela sectoral_scopes no workbook de referencia.
# 4. Criar/atualizar tabela technical_areas no workbook de referencia.
# 5. Manter taxonomia UNFCCC autoritativa para classificacao de projetos.


from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parents[3]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


from src.projects_standards.shared.reference.build_reference_dataset import (  # noqa: E402
    DEFAULT_OUTPUT_PATH,
    strip_worksheet_autofilters,
    validate_reference_dataset,
)


SOURCE_URL = "https://unfccc.int/sites/default/files/resource/A6.4-STAN-ACCR-001.pdf"
SOURCE_DOCUMENT = "UNFCCC A6.4-STAN-ACCR-001"
SOURCE_TABLE_NAME = "Table. Sectoral scopes and required sector technical knowledge"


SECTORAL_SCOPES = [
    ["SS 1", "Energy industries (renewable/non-renewable sources)", "Energy industries", "Industrias de energia (fontes renovaveis/nao renovaveis)", "Industrias de energia", "Industrias de energia (fuentes renovables/no renovables)", "Industrias de energia", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 42, 43, SOURCE_URL],
    ["SS 2", "Energy distribution", "Energy distribution", "Distribuicao de energia", "Distribuicao de energia", "Distribucion de energia", "Distribucion de energia", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 43, 43, SOURCE_URL],
    ["SS 3", "Energy demand", "Energy demand", "Demanda de energia", "Demanda de energia", "Demanda de energia", "Demanda de energia", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 44, 44, SOURCE_URL],
    ["SS 4", "Manufacturing industries", "Manufacturing", "Industrias de manufatura", "Manufatura", "Industrias manufactureras", "Manufactura", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 44, 45, SOURCE_URL],
    ["SS 5", "Chemical industry", "Chemical industry", "Industria quimica", "Industria quimica", "Industria quimica", "Industria quimica", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 45, 46, SOURCE_URL],
    ["SS 6", "Construction", "Construction", "Construcao", "Construcao", "Construccion", "Construccion", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 46, 46, SOURCE_URL],
    ["SS 7", "Transport", "Transport", "Transporte", "Transporte", "Transporte", "Transporte", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 46, 47, SOURCE_URL],
    ["SS 8", "Mining/mineral production", "Mining & minerals", "Mineracao/producao mineral", "Mineracao e minerais", "Mineria/produccion mineral", "Mineria y minerales", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 47, 47, SOURCE_URL],
    ["SS 9", "Metal production", "Metal production", "Producao de metais", "Producao de metais", "Produccion de metales", "Produccion de metales", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 47, 48, SOURCE_URL],
    ["SS 10", "Fugitive emissions from fuels (solid, oil and gas)", "Fuel fugitive emissions", "Emissoes fugitivas de combustiveis (solidos, petroleo e gas)", "Emissoes fugitivas de combustiveis", "Emisiones fugitivas de combustibles (solidos, petroleo y gas)", "Emisiones fugitivas de combustibles", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 48, 49, SOURCE_URL],
    ["SS 11", "Fugitive emissions from production and consumption of halocarbons and sulphur hexafluoride", "Halocarbon fugitive emissions", "Emissoes fugitivas da producao e do consumo de halocarbonos e hexafluoreto de enxofre", "Emissoes fugitivas de halocarbonos", "Emisiones fugitivas de la produccion y el consumo de halocarbonos y hexafluoruro de azufre", "Emisiones fugitivas de halocarbonos", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 49, 50, SOURCE_URL],
    ["SS 12", "Solvents use", "Solvents use", "Uso de solventes", "Uso de solventes", "Uso de solventes", "Uso de solventes", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 50, 50, SOURCE_URL],
    ["SS 13", "Waste handling and disposal", "Waste handling", "Manejo e disposicao de residuos", "Manejo de residuos", "Manejo y disposicion de residuos", "Manejo de residuos", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 50, 51, SOURCE_URL],
    ["SS 14", "Afforestation and reforestation", "Afforestation & reforestation", "Florestamento e reflorestamento", "Florestamento e reflorestamento", "Forestacion y reforestacion", "Forestacion y reforestacion", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 51, 51, SOURCE_URL],
    ["SS 15", "Agriculture", "Agriculture", "Agricultura", "Agricultura", "Agricultura", "Agricultura", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 51, 52, SOURCE_URL],
    ["SS 16", "Carbon capture and storage of CO2 in geological formation", "Geological CO2 storage", "Captura e armazenamento de CO2 em formacao geologica", "Armazenamento geologico de CO2", "Captura y almacenamiento de CO2 en formacion geologica", "Almacenamiento geologico de CO2", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 52, 53, SOURCE_URL],
    ["SS 17", "Other activities involving removals", "Other removals", "Outras atividades envolvendo remocoes", "Outras remocoes", "Otras actividades relacionadas con remociones", "Otras remociones", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 53, 54, SOURCE_URL],
]


TECHNICAL_AREAS = [
    ["TA 1.1", "SS 1", "Thermal energy generation", "Thermal generation", "Geracao termica de energia", "Geracao termica", "Generacion termica de energia", "Generacion termica", "Power and heat generation from non-renewable sources and biomass, including new plants, retrofits, efficiency gains, fuel switching, district heating systems and power grids. Typical emissions include CO2 from combustion and upstream fuel-chain emissions.", "Geracao de energia e calor a partir de fontes nao renovaveis e biomassa, incluindo novas plantas, retrofit, eficiencia energetica, troca de combustivel, aquecimento distrital e redes eletricas. As emissoes tipicas incluem CO2 da combustao e emissoes a montante da cadeia de combustiveis.", "Generacion de energia y calor a partir de fuentes no renovables y biomasa, incluyendo nuevas plantas, retrofit, eficiencia energetica, cambio de combustible, calefaccion distrital y redes electricas. Las emisiones tipicas incluyen CO2 por combustion y emisiones aguas arriba de la cadena de combustibles.", "Methods to evaluate mass and energy flows; characteristics of combustion devices, heat plants and power plants; grid dispatch and dispatch-based emission evaluation; and methods to estimate upstream fuel-related emissions with standard factors.", "Metodos para avaliar fluxos de massa e energia; caracteristicas de equipamentos de combustao, usinas termicas e eletricas; despacho de redes eletricas e avaliacao de emissoes com base no despacho; e metodos para estimar emissoes a montante relacionadas ao uso de combustiveis com fatores padrao.", "Metodos para evaluar flujos de masa y energia; caracteristicas de equipos de combustion, plantas termicas y electricas; despacho de redes electricas y evaluacion de emisiones basada en el despacho; y metodos para estimar emisiones aguas arriba relacionadas con el uso de combustibles mediante factores estandar.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 42, 42, SOURCE_URL],
    ["TA 1.2", "SS 1", "Renewables", "Energias renovaveis", "Energias renovables", "Power and heat generation from renewable sources, including new plants, retrofits, efficiency improvements and fuel switching. Typical emissions include combustion-related CO2 and technology-specific CO2 and CH4, such as reservoir or geothermal emissions.", "Geracao de energia e calor a partir de fontes renovaveis, incluindo novas plantas, retrofit, ganhos de eficiencia e troca de combustivel. As emissoes tipicas incluem CO2 relacionado a combustao e CO2/CH4 especificos de tecnologias, como reservatorios e geotermia.", "Generacion de energia y calor a partir de fuentes renovables, incluyendo nuevas plantas, retrofit, mejoras de eficiencia y cambio de combustible. Las emisiones tipicas incluyen CO2 relacionado con combustion y CO2/CH4 especificos de tecnologias, como embalses y geotermia.", "Methods to evaluate mass and energy flows; characteristics of renewable power plants, such as capacity, load factor, intermittency, auxiliary fuel use and GHG emissions; and grid dispatch analysis for GHG evaluation.", "Metodos para avaliar fluxos de massa e energia; caracteristicas de plantas renovaveis, como capacidade, fator de carga, intermitencia, uso de combustivel auxiliar e emissoes; e analise de despacho de rede para avaliacao de GEE.", "Metodos para evaluar flujos de masa y energia; caracteristicas de plantas renovables, como capacidad, factor de carga, intermitencia, uso de combustible auxiliar y emisiones; y analisis de despacho de red para evaluacion de GEI.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 43, 43, SOURCE_URL],
    ["TA 2.1", "SS 2", "Energy distribution", "Distribuicao de energia", "Distribucion de energia", "Energy efficiency measures in electricity transmission and distribution. Typical emissions include CO2 from fuel combustion for power and heat generation.", "Medidas de eficiencia energetica em transmissao e distribuicao de eletricidade. As emissoes tipicas incluem CO2 da combustao de combustiveis para geracao de energia e calor.", "Medidas de eficiencia energetica en transmision y distribucion de electricidad. Las emisiones tipicas incluyen CO2 por combustion de combustibles para generacion de energia y calor.", "Knowledge of efficiency measures in transmission and distribution systems, transformer upgrades, AC/DC transmission losses, energy-saving evaluation and transmission-voltage upgrading.", "Conhecimento sobre medidas de eficiencia em sistemas de transmissao e distribuicao, transformadores, perdas em transmissao CA/CC, avaliacao de economia de energia e elevacao de tensao de transmissao.", "Conocimiento sobre medidas de eficiencia en sistemas de transmision y distribucion, transformadores, perdidas en transmision CA/CC, evaluacion de ahorro de energia y aumento del voltaje de transmision.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 43, 44, SOURCE_URL],
    ["TA 3.1", "SS 3", "Energy demand", "Demanda de energia", "Demanda de energia", "Demand-side energy efficiency in sectors such as pumping, lighting, household appliances and buildings. Typical emissions include CO2 from commercial and non-commercial fuel combustion for power and heat generation.", "Eficiencia energetica do lado da demanda em setores como bombeamento, iluminacao, eletrodomesticos e edificios. As emissoes tipicas incluem CO2 da combustao comercial e nao comercial de combustiveis para geracao de energia e calor.", "Eficiencia energetica del lado de la demanda en sectores como bombeo, iluminacion, electrodomesticos y edificios. Las emisiones tipicas incluyen CO2 por combustion comercial y no comercial de combustibles para generacion de energia y calor.", "Methods to evaluate mass and energy flows in end-use energy demand, including direct monitoring, balances, energy-use factors and energy-efficiency factors.", "Metodos para avaliar fluxos de massa e energia no uso final de energia, incluindo monitoramento direto, balancos, fatores de uso de energia e fatores de eficiencia energetica.", "Metodos para evaluar flujos de masa y energia en el uso final de energia, incluyendo monitoreo directo, balances, factores de uso de energia y factores de eficiencia energetica.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 44, 44, SOURCE_URL],
    ["TA 4.1", "SS 4", "Cement and lime production", "Producao de cimento e cal", "Produccion de cemento y cal", "Cement production activities, especially fuel switching and alternative raw materials. Typical emissions include process emissions from calcination of carbonated raw materials.", "Atividades de producao de cimento, especialmente troca de combustivel e uso de materias-primas alternativas. As emissoes tipicas incluem emissoes de processo da calcinacao de materias-primas carbonatadas.", "Actividades de produccion de cemento, especialmente cambio de combustible y uso de materias primas alternativas. Las emisiones tipicas incluyen emisiones de proceso por calcinacion de materias primas carbonatadas.", "Unit operations in cement and lime production; raw materials and fuels such as limestone, dolomite, magnesite and kiln fuels; mass and energy balances; and methods to determine carbonate content of raw materials.", "Operacoes unitarias na producao de cimento e cal; materias-primas e combustiveis como calcario, dolomita, magnesita e combustiveis de forno; balancos de massa e energia; e metodos para determinar o teor de carbonato das materias-primas.", "Operaciones unitarias en la produccion de cemento y cal; materias primas y combustibles como caliza, dolomita, magnesita y combustibles de horno; balances de masa y energia; y metodos para determinar el contenido de carbonato de las materias primas.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 44, 45, SOURCE_URL],
    ["TA 5.1", "SS 5", "Chemical industry", "Industria quimica", "Industria quimica", "Production of chemicals and manufactured materials such as biodiesel, charcoal, upgraded biogas, ammonia, urea, CO2-based chemicals and hydrogen. Typical emissions arise from processes such as transesterification, pyrolysis, carbonization, reforming and gas upgrading.", "Producao de produtos quimicos e materiais manufaturados, como biodiesel, carvao vegetal, biogas purificado, amonia, ureia, produtos quimicos a base de CO2 e hidrogenio. As emissoes tipicas decorrem de processos como transesterificacao, pirolise, carbonizacao, reforma e purificacao de gas.", "Produccion de productos quimicos y materiales manufacturados, como biodiesel, carbon vegetal, biogas mejorado, amoniaco, urea, productos quimicos basados en CO2 e hidrogeno. Las emisiones tipicas provienen de procesos como transesterificacion, pirolisis, carbonizacion, reformado y acondicionamiento de gas.", "Chemical processes, reactions and stoichiometry; unit operations in the chemical process industry; and mass and energy balances for chemical and manufacturing processes.", "Processos quimicos, reacoes e estequiometria; operacoes unitarias da industria de processos quimicos; e balancos de massa e energia para processos quimicos e manufatureiros.", "Procesos quimicos, reacciones y estequiometria; operaciones unitarias de la industria de procesos quimicos; y balances de masa y energia para procesos quimicos y manufactureros.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 45, 45, SOURCE_URL],
    ["TA 5.2", "SS 5", "Caprolactam, nitric and adipic acid", "Caprolactama, acido nitrico e acido adipico", "Caprolactama, acido nitrico y acido adipico", "Management and abatement of N2O emissions from caprolactam, nitric acid and adipic acid plants. Typical emissions are N2O from those production processes.", "Gestao e abatimento de emissoes de N2O em plantas de caprolactama, acido nitrico e acido adipico. As emissoes tipicas sao as de N2O nesses processos produtivos.", "Gestion y abatimiento de emisiones de N2O en plantas de caprolactama, acido nitrico y acido adipico. Las emisiones tipicas son las de N2O en esos procesos productivos.", "Knowledge of reactions, stoichiometry, mass and energy balances in those processes; methods to evaluate N2O sources; and primary, secondary and tertiary N2O abatement options.", "Conhecimento de reacoes, estequiometria, balancos de massa e energia nesses processos; metodos para avaliar fontes de N2O; e opcoes de abatimento primario, secundario e terciario de N2O.", "Conocimiento de reacciones, estequiometria, balances de masa y energia en esos procesos; metodos para evaluar fuentes de N2O; y opciones de abatimiento primario, secundario y terciario de N2O.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 45, 46, SOURCE_URL],
    ["TA 6.1", "SS 6", "Construction", "Construcao", "Construccion", "Construction activities focused on lower-GHG materials and techniques for buildings. The document notes that energy efficiency in buildings belongs under energy demand and that technical knowledge here is only indicative.", "Atividades de construcao focadas em materiais e tecnicas de menor intensidade de GEE para edificios. O documento observa que eficiencia energetica em edificios pertence a demanda de energia e que o conhecimento tecnico aqui e apenas indicativo.", "Actividades de construccion centradas en materiales y tecnicas de menor intensidad de GEI para edificios. El documento indica que la eficiencia energetica en edificios pertenece a demanda de energia y que el conocimiento tecnico aqui es solo indicativo.", "Knowledge of building construction, foundations, structural systems and material requirements; GHG sources from material production and transport; and regional codes and best practices to define baselines and baseline emissions.", "Conhecimento de construcao de edificios, fundacoes, sistemas estruturais e requisitos de materiais; fontes de GEE na producao e no transporte de materiais; e codigos e boas praticas regionais para definir linhas de base e emissoes de linha de base.", "Conocimiento de construccion de edificios, cimentaciones, sistemas estructurales y requisitos de materiales; fuentes de GEI en la produccion y el transporte de materiales; y codigos y buenas practicas regionales para definir lineas base y emisiones de linea base.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 46, 46, SOURCE_URL],
    ["TA 7.1", "SS 7", "Transport", "Transporte", "Transporte", "Modal shifts, fuel switches and lower-GHG transport modes for freight and passengers. Typical emissions are CO2 from fossil fuel combustion in transport activities.", "Mudancas modais, troca de combustivel e modos de transporte de menor intensidade de GEE para cargas e passageiros. As emissoes tipicas sao CO2 da combustao de combustiveis fosseis nas atividades de transporte.", "Cambios modales, cambio de combustible y modos de transporte de menor intensidad de GEI para carga y pasajeros. Las emisiones tipicas son CO2 por combustion de combustibles fosiles en actividades de transporte.", "Transport system modelling; service levels, travel distances and baseline modes; surveys and sampling for alternative scenarios; rebound and induced-traffic effects; and methods to quantify primary energy use and GHG emissions for transport modes.", "Modelagem de sistemas de transporte; niveis de servico, distancias percorridas e modos de linha de base; pesquisas e amostragem para cenarios alternativos; efeitos rebote e trafego induzido; e metodos para quantificar uso de energia primaria e emissoes de GEE dos modos de transporte.", "Modelizacion de sistemas de transporte; niveles de servicio, distancias recorridas y modos de linea base; encuestas y muestreo para escenarios alternativos; efectos rebote y trafico inducido; y metodos para cuantificar el uso de energia primaria y las emisiones de GEI de los modos de transporte.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 46, 47, SOURCE_URL],
    ["TA 8.1", "SS 8", "Mining/mineral production", "Mineracao/producao mineral", "Mineria/produccion mineral", "Mine methane management and capture/use of waste gas. Typical emissions include CH4 from metal ore and coal mining.", "Gestao de metano de mina e captura/uso de gases residuais. As emissoes tipicas incluem CH4 da mineracao de minerio metalico e carvao.", "Gestion de metano de mina y captura/uso de gases residuales. Las emisiones tipicas incluyen CH4 de la mineria de mineral metalico y carbon.", "Unit operations in mining and coal industries, such as drilling, blasting, hauling, ventilation and drainage; mass and energy balances; and potential uses, flaring and venting of waste streams and mine methane.", "Operacoes unitarias das industrias de mineracao e carvao, como perfuracao, desmonte, transporte, ventilacao e drenagem; balancos de massa e energia; e usos potenciais, queima e ventilacao de correntes residuais e metano de mina.", "Operaciones unitarias de las industrias minera y del carbon, como perforacion, voladura, acarreo, ventilacion y drenaje; balances de masa y energia; y usos potenciales, quema y venteo de corrientes residuales y metano de mina.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 47, 47, SOURCE_URL],
    ["TA 9.1", "SS 9", "Aluminium and magnesium production", "Producao de aluminio e magnesio", "Produccion de aluminio y magnesio", "Management of PFC emissions in aluminium production. Typical emissions include PFCs, SF6 and other cover gases, plus CO2 from combustion for power and heat generation.", "Gestao de emissoes de PFC na producao de aluminio. As emissoes tipicas incluem PFCs, SF6 e outros gases de cobertura, alem de CO2 da combustao para geracao de energia e calor.", "Gestion de emisiones de PFC en la produccion de aluminio. Las emisiones tipicas incluyen PFC, SF6 y otros gases de cobertura, ademas de CO2 por combustion para generacion de energia y calor.", "Unit operations in metallurgy; mass and energy balances; evaluation of specific energy consumption of furnaces and kilns; and understanding of anode effects, PFC emissions and mitigation in aluminium and magnesium processes.", "Operacoes unitarias em metalurgia; balancos de massa e energia; avaliacao do consumo especifico de energia de fornos; e compreensao de efeitos anodicos, emissoes de PFC e mitigacao em processos de aluminio e magnesio.", "Operaciones unitarias en metalurgia; balances de masa y energia; evaluacion del consumo especifico de energia de hornos; y comprension de efectos anodicos, emisiones de PFC y mitigacion en procesos de aluminio y magnesio.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 47, 48, SOURCE_URL],
    ["TA 9.2", "SS 9", "Iron, steel and ferro-alloy production", "Producao de ferro, aco e ferroligas", "Produccion de hierro, acero y ferroaleaciones", "Management of CO2 emissions in iron production and recovery/use of waste gases in iron and steel production. Typical emissions include CO2 from iron reduction and fuel combustion for power and heat generation.", "Gestao de emissoes de CO2 na producao de ferro e recuperacao/uso de gases residuais na producao de ferro e aco. As emissoes tipicas incluem CO2 da reducao de ferro e da combustao para geracao de energia e calor.", "Gestion de emisiones de CO2 en la produccion de hierro y recuperacion/uso de gases residuales en la produccion de hierro y acero. Las emisiones tipicas incluyen CO2 por reduccion del hierro y por combustion para generacion de energia y calor.", "Unit operations in metallurgy; mass and energy balances; evaluation of furnace and kiln energy performance; and energy recovery and utilization from blast furnace, coke oven and converter gases.", "Operacoes unitarias em metalurgia; balancos de massa e energia; avaliacao do desempenho energetico de fornos; e recuperacao e uso energetico de gases de alto-forno, coqueria e conversor.", "Operaciones unitarias en metalurgia; balances de masa y energia; evaluacion del desempeno energetico de hornos; y recuperacion y uso energetico de gases de alto horno, coqueria y convertidor.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 47, 48, SOURCE_URL],
    ["TA 10.1", "SS 10", "Fugitive emissions from oil and gas", "Emissoes fugitivas de oleo e gas", "Emisiones fugitivas de petroleo y gas", "Management of leakage, venting and flaring of natural gas and associated petroleum gas in oil and gas facilities. Typical emissions include CH4 from gas systems and CO2 from gas flaring.", "Gestao de vazamentos, ventilacao e queima de gas natural e gas associado de petroleo em instalacoes de oleo e gas. As emissoes tipicas incluem CH4 de sistemas de gas e CO2 da queima de gas.", "Gestion de fugas, venteo y quema de gas natural y gas asociado del petroleo en instalaciones de petroleo y gas. Las emisiones tipicas incluyen CH4 de sistemas de gas y CO2 de la quema de gas.", "Unit operations in oil and gas; reservoir dynamics, enhanced oil recovery, gas lift and associated gas production; mass and energy balances; uses, flaring and venting of waste streams; methane monitoring technologies; and the OGMP 2.0 reporting and mitigation framework.", "Operacoes unitarias de oleo e gas; dinamica de reservatorios, recuperacao avancada, gas lift e producao de gas associado; balancos de massa e energia; usos, queima e ventilacao de correntes residuais; tecnologias de monitoramento de metano; e o framework OGMP 2.0 de reporte e mitigacao.", "Operaciones unitarias de petroleo y gas; dinamica de reservorios, recuperacion mejorada, gas lift y produccion de gas asociado; balances de masa y energia; usos, quema y venteo de corrientes residuales; tecnologias de monitoreo de metano; y el marco OGMP 2.0 de reporte y mitigacion.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 48, 49, SOURCE_URL],
    ["TA 11.1", "SS 11", "Emissions of fluorinated gases", "Emissoes de gases fluorados", "Emisiones de gases fluorados", "Mitigation of HFC emissions used as refrigerants and blowing agents, mitigation of SF6 in electrical equipment and mitigation of fluorinated gases in semiconductor manufacturing. Typical emissions include HFC, SF6 and other fluorinated GHGs.", "Mitigacao de emissoes de HFC usados como refrigerantes e agentes expansores, mitigacao de SF6 em equipamentos eletricos e mitigacao de gases fluorados na fabricacao de semicondutores. As emissoes tipicas incluem HFC, SF6 e outros GEE fluorados.", "Mitigacion de emisiones de HFC usados como refrigerantes y agentes espumantes, mitigacion de SF6 en equipos electricos y mitigacion de gases fluorados en la fabricacion de semiconductores. Las emisiones tipicas incluyen HFC, SF6 y otros GEI fluorados.", "Applications of HFC, SF6 and other fluorinated gases in manufacturing; mitigation and abatement of fluorinated GHG emissions; and monitoring methods such as FTIR, quadrupole mass spectrometry, mass balances and gas chromatography.", "Aplicacoes de HFC, SF6 e outros gases fluorados em processos produtivos; mitigacao e abatimento de emissoes de GEE fluorados; e metodos de monitoramento como FTIR, espectrometria de massa quadrupolar, balancos de massa e cromatografia gasosa.", "Aplicaciones de HFC, SF6 y otros gases fluorados en procesos productivos; mitigacion y abatimiento de emisiones de GEI fluorados; y metodos de monitoreo como FTIR, espectrometria de masas cuadrupolar, balances de masa y cromatografia de gases.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 49, 49, SOURCE_URL],
    ["TA 11.2", "SS 11", "Refrigerant gas production", "Producao de gas refrigerante", "Produccion de gas refrigerante", "Production of HCFC-22 refrigerant gas. Typical emissions include HFC-23.", "Producao de gas refrigerante HCFC-22. As emissoes tipicas incluem HFC-23.", "Produccion de gas refrigerante HCFC-22. Las emisiones tipicas incluyen HFC-23.", "Unit operations in swing and non-swing HCFC-22 plants; formation of HFC-23 streams and mitigation measures; use of mass balances to evaluate HFC-23 generation and emissions; and monitoring of HFC streams with flow meters and gas chromatography.", "Operacoes unitarias em plantas HCFC-22 swing e non-swing; formacao de correntes de HFC-23 e medidas de mitigacao; uso de balancos de massa para avaliar geracao e emissoes de HFC-23; e monitoramento de correntes de HFC com medidores de vazao e cromatografia gasosa.", "Operaciones unitarias en plantas HCFC-22 swing y non-swing; formacion de corrientes de HFC-23 y medidas de mitigacion; uso de balances de masa para evaluar generacion y emisiones de HFC-23; y monitoreo de corrientes de HFC con medidores de flujo y cromatografia de gases.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 49, 50, SOURCE_URL],
    ["TA 12.1", "SS 12", "Chemical industry", "Industria quimica", "Industria quimica", "Projects involving the use of solvents. Typical emissions are GHG emissions related to solvent use.", "Projetos envolvendo o uso de solventes. As emissoes tipicas sao emissoes de GEE relacionadas ao uso de solventes.", "Proyectos que implican el uso de solventes. Las emisiones tipicas son emisiones de GEI relacionadas con el uso de solventes.", "Chemical processes, reactions and stoichiometry; unit operations in the chemical process industry; and mass and energy balances in chemical and manufacturing processes.", "Processos quimicos, reacoes e estequiometria; operacoes unitarias da industria de processos quimicos; e balancos de massa e energia em processos quimicos e manufatureiros.", "Procesos quimicos, reacciones y estequiometria; operaciones unitarias de la industria de procesos quimicos; y balances de masa y energia en procesos quimicos y manufactureros.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 50, 50, SOURCE_URL],
    ["TA 13.1", "SS 13", "Solid waste and wastewater", "Residuos solidos e aguas residuais", "Residuos solidos y aguas residuales", "Landfilling, alternative solid waste management such as gasification, incineration, recycling and refuse-derived fuel, wastewater treatment systems and biogas management. Typical emissions include CH4 from anaerobic decay in waste and wastewater.", "Aterros, gestao alternativa de residuos solidos como gaseificacao, incineracao, reciclagem e combustivel derivado de residuos, sistemas de tratamento de aguas residuais e gestao de biogas. As emissoes tipicas incluem CH4 da decomposicao anaerobia em residuos e efluentes.", "Rellenos sanitarios, gestion alternativa de residuos solidos como gasificacion, incineracion, reciclaje y combustible derivado de residuos, sistemas de tratamiento de aguas residuales y gestion de biogas. Las emisiones tipicas incluyen CH4 por descomposicion anaerobia en residuos y efluentes.", "Biomass decay under aerobic and anaerobic conditions; types and composition of solid waste and wastewater; decay models and standard emission factors; and alternative methods for waste disposal, management and treatment.", "Decomposicao da biomassa em condicoes aerobias e anaerobias; tipos e composicao de residuos solidos e aguas residuais; modelos de decaimento e fatores padrao de emissao; e metodos alternativos de disposicao, manejo e tratamento de residuos.", "Descomposicion de biomasa en condiciones aerobias y anaerobias; tipos y composicion de residuos solidos y aguas residuales; modelos de decaimiento y factores estandar de emision; y metodos alternativos de disposicion, manejo y tratamiento de residuos.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 50, 50, SOURCE_URL],
    ["TA 13.2", "SS 13", "Manure", "Esterco", "Estiercol", "Manure management systems and biogas management. Typical emissions include CH4 from anaerobic decay of organic matter in manure.", "Sistemas de manejo de esterco e gestao de biogas. As emissoes tipicas incluem CH4 da decomposicao anaerobia da materia organica no esterco.", "Sistemas de manejo de estiercol y gestion de biogas. Las emisiones tipicas incluyen CH4 por descomposicion anaerobia de la materia organica en el estiercol.", "Biomass decay and biogas generation; types of manure, their composition and characterization; livestock types and dietary factors affecting manure generation; and decay models and standard factors for manure management emissions.", "Decomposicao da biomassa e geracao de biogas; tipos de esterco, sua composicao e caracterizacao; tipos de rebanho e fatores alimentares que afetam a geracao de esterco; e modelos de decaimento e fatores padrao para emissoes de manejo de esterco.", "Descomposicion de biomasa y generacion de biogas; tipos de estiercol, su composicion y caracterizacion; tipos de ganado y factores dieteticos que afectan la generacion de estiercol; y modelos de decaimiento y factores estandar para emisiones del manejo de estiercol.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 50, 51, SOURCE_URL],
    ["TA 14.1", "SS 14", "Afforestation and reforestation", "Florestamento e reflorestamento", "Forestacion y reforestacion", "Afforestation, reforestation, agroforestry and mangrove projects. Typical reservoirs are carbon stocks in biomass, dead wood, litter and soil carbon.", "Projetos de florestamento, reflorestamento, agrofloresta e manguezais. Os reservatorios tipicos sao estoques de carbono em biomassa, madeira morta, serapilheira e carbono do solo.", "Proyectos de forestacion, reforestacion, agroforesteria y manglares. Los reservorios tipicos son existencias de carbono en biomasa, madera muerta, hojarasca y carbono del suelo.", "Quantification of carbon stocks and changes in biomass and soil; GHG emissions from displacement of pre-project agricultural activities; and definition and identification of degraded and degrading lands in the Article 6.4 context.", "Quantificacao de estoques de carbono e de suas mudancas em biomassa e solo; emissoes de GEE decorrentes do deslocamento de atividades agricolas pre-projeto; e definicao e identificacao de terras degradadas e em degradacao no contexto do Artigo 6.4.", "Cuantificacion de existencias de carbono y de sus cambios en biomasa y suelo; emisiones de GEI derivadas del desplazamiento de actividades agricolas previas al proyecto; y definicion e identificacion de tierras degradadas y en degradacion en el contexto del Articulo 6.4.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 51, 51, SOURCE_URL],
    ["TA 15.1", "SS 15", "Agriculture", "Agricultura", "Agricultura", "Management of agricultural operations to reduce emissions and management of fertilizer application. Typical emissions include N2O from fertilizer application, changes in carbon stocks and CO2 from fuel combustion.", "Gestao de operacoes agricolas para reduzir emissoes e gestao da aplicacao de fertilizantes. As emissoes tipicas incluem N2O da aplicacao de fertilizantes, mudancas nos estoques de carbono e CO2 da combustao de combustiveis.", "Gestion de operaciones agricolas para reducir emisiones y gestion de la aplicacion de fertilizantes. Las emisiones tipicas incluyen N2O por aplicacion de fertilizantes, cambios en las existencias de carbono y CO2 por combustion de combustibles.", "Knowledge of agricultural operations and their emission sources; fossil fuel and electricity use; emissions from synthetic and organic fertilizers, urea, dolomite and limestone; field burning; soil carbon stocks and land management; displacement effects; and identification of degraded lands.", "Conhecimento de operacoes agricolas e suas fontes de emissao; uso de combustiveis fosseis e eletricidade; emissoes de fertilizantes sinteticos e organicos, ureia, dolomita e calcario; queima em campo; estoques de carbono do solo e manejo da terra; efeitos de deslocamento; e identificacao de terras degradadas.", "Conocimiento de operaciones agricolas y sus fuentes de emision; uso de combustibles fosiles y electricidad; emisiones de fertilizantes sinteticos y organicos, urea, dolomita y caliza; quema en campo; existencias de carbono del suelo y manejo de la tierra; efectos de desplazamiento; e identificacion de tierras degradadas.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 51, 52, SOURCE_URL],
    ["TA 16.1", "SS 16", "Carbon Capture and Storage", "Captura e Armazenamento de Carbono", "Captura y Almacenamiento de Carbono", "Activities related to CCS, DACCS and BECCS with geological storage. The typical reservoir is the geological formation.", "Atividades relacionadas a CCS, DACCS e BECCS com armazenamento geologico. O reservatorio tipico e a formacao geologica.", "Actividades relacionadas con CCS, DACCS y BECCS con almacenamiento geologico. El reservorio tipico es la formacion geologica.", "Unit operations in CCS/DACCS/BECCS facilities; boundaries of geological storage sites and complexes; migration of CO2 plumes; estimation of emissions through mass balance and direct monitoring; and procedures to determine emissions from leakage and seepage events.", "Operacoes unitarias em instalacoes de CCS/DACCS/BECCS; limites de sitios e complexos de armazenamento geologico; migracao de plumas de CO2; estimativa de emissoes por balanco de massa e monitoramento direto; e procedimentos para determinar emissoes de vazamentos e seepage.", "Operaciones unitarias en instalaciones de CCS/DACCS/BECCS; limites de sitios y complejos de almacenamiento geologico; migracion de plumas de CO2; estimacion de emisiones por balance de masa y monitoreo directo; y procedimientos para determinar emisiones de fugas y filtraciones.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 52, 53, SOURCE_URL],
    ["TA 17.1", "SS 17", "Land-based biological removal activities", "Atividades biologicas terrestres de remocao", "Actividades biologicas terrestres de remocion", "Anthropogenic CO2 removal activities using biological methods, such as biochar, improved forest management, peatland and coastal wetland restoration, and soil carbon sequestration in croplands and grasslands. Typical reservoirs include vegetation, buried biomass and biochar.", "Atividades antropogenicas de remocao de CO2 por meios biologicos, como biochar, manejo florestal melhorado, restauracao de turfeiras e zonas umidas costeiras, e sequestro de carbono no solo em areas agricolas e pastagens. Os reservatorios tipicos incluem vegetacao, biomassa enterrada e biochar.", "Actividades antropogenicas de remocion de CO2 mediante metodos biologicos, como biochar, manejo forestal mejorado, restauracion de turberas y humedales costeros, y secuestro de carbono en suelos de cultivos y pastizales. Los reservorios tipicos incluyen vegetacion, biomasa enterrada y biochar.", "Biological carbon removal processes; non-permanence and reversal risks; approaches to remedy reversals such as buffer pools, insurance and replacement; potential leakage sources; and monitoring, quantification, accounting, estimation and reporting of removals.", "Processos biologicos de remocao de carbono; riscos de nao permanencia e reversao; abordagens para remediar reversoes, como buffer pools, seguros e substituicao; fontes potenciais de leakage; e monitoramento, quantificacao, contabilizacao, estimativa e reporte das remocoes.", "Procesos biologicos de remocion de carbono; riesgos de no permanencia y reversion; enfoques para remediar reversiones, como buffer pools, seguros y reemplazo; fuentes potenciales de fugas; y monitoreo, cuantificacion, contabilizacion, estimacion y reporte de las remociones.", SOURCE_DOCUMENT, SOURCE_TABLE_NAME, 53, 54, SOURCE_URL],
]


TECHNICAL_AREA_SHORT_NAMES = {
    "TA 1.1": ("Thermal generation", "Geracao termica", "Generacion termica"),
    "TA 1.2": ("Renewables", "Renovaveis", "Renovables"),
    "TA 2.1": ("Power distribution", "Distribuicao eletrica", "Distribucion electrica"),
    "TA 3.1": ("End-use efficiency", "Eficiencia no uso final", "Eficiencia de uso final"),
    "TA 4.1": ("Cement & lime", "Cimento e cal", "Cemento y cal"),
    "TA 5.1": ("Chemical production", "Producao quimica", "Produccion quimica"),
    "TA 5.2": ("Caprolactam & acids", "Caprolactama e acidos", "Caprolactama y acidos"),
    "TA 6.1": ("Low-carbon construction", "Construcao baixo carbono", "Construccion baja en carbono"),
    "TA 7.1": ("Low-carbon transport", "Transporte baixo carbono", "Transporte bajo carbono"),
    "TA 8.1": ("Mine methane", "Metano de mina", "Metano de mina"),
    "TA 9.1": ("Aluminium & magnesium", "Aluminio e magnesio", "Aluminio y magnesio"),
    "TA 9.2": ("Iron & steel", "Ferro e aco", "Hierro y acero"),
    "TA 10.1": ("Oil & gas fugitives", "Fugitivas de oleo e gas", "Fugitivas de petroleo y gas"),
    "TA 11.1": ("Fluorinated gases", "Gases fluorados", "Gases fluorados"),
    "TA 11.2": ("HCFC-22 production", "Producao de HCFC-22", "Produccion de HCFC-22"),
    "TA 12.1": ("Solvent use", "Uso de solventes", "Uso de solventes"),
    "TA 13.1": ("Waste & wastewater", "Residuos e efluentes", "Residuos y efluentes"),
    "TA 13.2": ("Manure management", "Manejo de esterco", "Manejo de estiercol"),
    "TA 14.1": ("Afforestation & reforestation", "Florestamento e reflorestamento", "Forestacion y reforestacion"),
    "TA 15.1": ("Agricultural operations", "Operacoes agricolas", "Operaciones agricolas"),
    "TA 16.1": ("CCS/DACCS/BECCS", "CCS/DACCS/BECCS", "CCS/DACCS/BECCS"),
    "TA 17.1": ("Land removals", "Remocoes terrestres", "Remociones terrestres"),
}


SECTORAL_SCOPE_HEADERS = [
    "sectoral_scope_id",
    "sectoral_scope_name_en",
    "sectoral_scope_short_name_en",
    "sectoral_scope_name_pt",
    "sectoral_scope_short_name_pt",
    "sectoral_scope_name_es",
    "sectoral_scope_short_name_es",
    "source_document",
    "source_table_name",
    "source_page_start",
    "source_page_end",
    "source_url",
]


TECHNICAL_AREA_HEADERS = [
    "technical_area_id",
    "sectoral_scope_id",
    "technical_area_name_en",
    "technical_area_short_name_en",
    "technical_area_name_pt",
    "technical_area_short_name_pt",
    "technical_area_name_es",
    "technical_area_short_name_es",
    "typical_activities_and_ghg_en",
    "typical_activities_and_ghg_pt",
    "typical_activities_and_ghg_es",
    "technical_knowledge_en",
    "technical_knowledge_pt",
    "technical_knowledge_es",
    "source_document",
    "source_table_name",
    "source_page_start",
    "source_page_end",
    "source_url",
]


# Expande as linhas de areas tecnicas antigas para o formato que inclui short names.
def normalize_technical_area_rows() -> None:
    normalized_rows = []
    for row in TECHNICAL_AREAS:
        if len(row) == 19:
            normalized_rows.append(row)
            continue

        if len(row) != 16:
            raise ValueError(f"Estrutura inesperada em TECHNICAL_AREAS para {row[0]}: {len(row)} colunas.")

        short_name_en, short_name_pt, short_name_es = TECHNICAL_AREA_SHORT_NAMES[row[0]]
        normalized_rows.append(
            [
                row[0],
                row[1],
                row[2],
                short_name_en,
                row[3],
                short_name_pt,
                row[4],
                short_name_es,
                row[5],
                row[6],
                row[7],
                row[8],
                row[9],
                row[10],
                row[11],
                row[12],
                row[13],
                row[14],
                row[15],
            ]
        )

    TECHNICAL_AREAS[:] = normalized_rows


normalize_technical_area_rows()


# Remove a aba de destino quando ela ja existe para permitir regeneracao idempotente.
def remove_sheet_if_exists(workbook, sheet_name: str) -> None:
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])


# Escreve uma aba tabular, aplica estilos basicos e cria a tabela estruturada.
def write_table_sheet(workbook, sheet_name: str, table_name: str, headers: list[str], rows: list[list[object]]) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    for cell in worksheet[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_index in range(2, worksheet.max_row + 1):
        url_cell = worksheet.cell(row=row_index, column=worksheet.max_column)
        if isinstance(url_cell.value, str) and url_cell.value.startswith("http"):
            url_cell.hyperlink = url_cell.value
            url_cell.style = "Hyperlink"

    width_map = {
        "A": 16,
        "B": 18,
        "C": 36,
        "D": 36,
        "E": 36,
        "F": 55,
        "G": 55,
        "H": 55,
        "I": 55,
        "J": 55,
        "K": 55,
        "L": 24,
        "M": 38,
        "N": 14,
        "O": 14,
        "P": 62,
        "Q": 24,
        "R": 14,
        "S": 62,
    }
    for column_letter, width in width_map.items():
        worksheet.column_dimensions[column_letter].width = width

    worksheet.freeze_panes = "A2"
    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium15",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


# Atualiza o workbook consolidado com as referencias setoriais da UNFCCC.
def sync_unfccc_sectoral_reference(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    try:
        remove_sheet_if_exists(workbook, "sectoral_scopes")
        remove_sheet_if_exists(workbook, "technical_areas")

        write_table_sheet(
            workbook=workbook,
            sheet_name="sectoral_scopes",
            table_name="tb_ref_sectoral_scopes",
            headers=SECTORAL_SCOPE_HEADERS,
            rows=SECTORAL_SCOPES,
        )
        write_table_sheet(
            workbook=workbook,
            sheet_name="technical_areas",
            table_name="tb_ref_technical_areas",
            headers=TECHNICAL_AREA_HEADERS,
            rows=TECHNICAL_AREAS,
        )

        workbook.save(workbook_path)
    finally:
        workbook.close()

    strip_worksheet_autofilters(workbook_path)
    validate_reference_dataset(workbook_path)


# Executa a sincronizacao das tabelas setoriais da UNFCCC no workbook consolidado.
def main() -> int:
    sync_unfccc_sectoral_reference(DEFAULT_OUTPUT_PATH)
    print(f"UNFCCC sectoral reference sincronizada com sucesso: {DEFAULT_OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
