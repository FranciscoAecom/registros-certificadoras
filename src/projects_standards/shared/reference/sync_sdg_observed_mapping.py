# Objetivo do script:
# Gerar no workbook consolidado a tabela com todas as formas observadas de SDGs na silver e seu mapeamento para o goal_id padrao.
# Processo:
# 1. Definir nomes de aba/tabela e cabecalhos para sdg_observed_mapping.
# 2. Escanear todos os datasets silver em 02_silver/.
# 3. Coletar formas observadas de SDGs dos campos sdg_targets.
# 4. Agregar contagens de ocorrencia com Counter.
# 5. Normalizar texto (trim, collapse whitespace, unicode).
# 6. Criar/atualizar tabela sdg_observed_mapping com [sdg_raw, observed_count, sdg_goal_id].
# 7. Preservar mapeamentos goal_id existentes e identificar novas formas.


from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parents[3]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


from src.projects_standards.shared.reference.build_reference_dataset import (  # noqa: E402
    DEFAULT_OUTPUT_PATH,
    strip_worksheet_autofilters,
    validate_reference_dataset,
)


SILVER_DIR = ROOT_DIR / "data" / "project_standards" / "02_silver"
SHEET_NAME = "sdg_observed_mapping"
TABLE_NAME = "tb_ref_sdg_observed_mapping"
HEADERS = ["sdg_raw", "observed_count", "sdg_goal_id"]


SDG_KEYWORDS = {
    1: ["goal 1", "1:", "1.", "1 -", "no poverty", "poverty"],
    2: ["goal 2", "2:", "2.", "2 -", "zero hunger", "hunger"],
    3: ["goal 3", "3:", "3.", "3 -", "good health", "well-being", "well being"],
    4: ["goal 4", "4:", "4.", "4 -", "quality education", "education"],
    5: ["goal 5", "5:", "5.", "5 -", "gender equality"],
    6: ["goal 6", "6:", "6.", "6 -", "clean water", "sanitation"],
    7: ["goal 7", "7:", "7.", "7 -", "clean energy", "affordable and clean energy"],
    8: ["goal 8", "8:", "8.", "8 -", "decent work", "economic growth"],
    9: ["goal 9", "9:", "9.", "9 -", "industry, innovation and infrastructure", "industry and infrastructure"],
    10: ["goal 10", "10:", "10.", "10 -", "reduced inequalities"],
    11: ["goal 11", "11:", "11.", "11 -", "sustainable cities and communities"],
    12: ["goal 12", "12:", "12.", "12 -", "responsible consumption and production", "responsible production and consumption", "responsible consumption"],
    13: ["goal 13", "13:", "13.", "13 -", "climate action"],
    14: ["goal 14", "14:", "14.", "14 -", "life below water", "life underwater"],
    15: ["goal 15", "15:", "15.", "15 -", "life on land"],
    16: ["goal 16", "16:", "16.", "16 -", "peace, justice and strong institutions", "peace justice and strong institutions"],
    17: ["goal 17", "17:", "17.", "17 -", "partnerships for the goals", "partnership for the goals", "partnership"],
}


# Normaliza texto para facilitar comparacao de variacoes observadas nos datasets silver.
def normalize_text(text: object) -> str:
    value = "" if text is None else str(text)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", value).strip().lower()


# Mapeia a forma observada de SDG para o goal_id padrao quando houver correspondencia confiavel.
def map_sdg_goal_id(raw_value: str) -> int | None:
    normalized = normalize_text(raw_value)
    if normalized.startswith("sdg "):
        normalized = normalized.replace("sdg ", "", 1)
    match = re.match(r"goal\s+(\d{1,2})\b", normalized)
    if match:
        return int(match.group(1))
    match = re.match(r"(\d{1,2})\s*[:.\-]\s*", normalized)
    if match:
        return int(match.group(1))
    for goal_id, keywords in SDG_KEYWORDS.items():
        if any(keyword in normalized for keyword in keywords):
            return goal_id
    return None


# Coleta todas as formas observadas de SDGs na camada silver.
def collect_observed_sdgs() -> Counter:
    counter = Counter()
    for file_path in SILVER_DIR.rglob("allprojects.json"):
        try:
            dataset = json.loads(file_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        for project in dataset.get("projects") or []:
            values = project.get("sdg_targets") or []
            if not isinstance(values, list):
                continue
            for value in values:
                raw = "" if value is None else str(value).strip()
                if raw:
                    counter[raw] += 1
    return counter


# Carrega mapeamentos ja existentes para preservar overrides manuais do workbook.
def load_existing_sdg_mapping(workbook_path: Path) -> dict[str, int | None]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            return {}
        worksheet = workbook[SHEET_NAME]
        headers = [cell.value for cell in worksheet[1]]
        header_index = {str(header): index for index, header in enumerate(headers) if header is not None}
        if "sdg_raw" not in header_index or "sdg_goal_id" not in header_index:
            return {}

        existing: dict[str, int | None] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            raw_value = row[header_index["sdg_raw"]]
            goal_id = row[header_index["sdg_goal_id"]]
            if raw_value in (None, ""):
                continue
            existing[str(raw_value).strip()] = None if goal_id in (None, "") else int(goal_id)
        return existing
    finally:
        workbook.close()


# Escreve a tabela de mapeamento de SDGs observadas no workbook consolidado.
def sync_sdg_observed_mapping(workbook_path: Path) -> Counter:
    observed_sdgs = collect_observed_sdgs()
    existing_mapping = load_existing_sdg_mapping(workbook_path)
    workbook = load_workbook(workbook_path)
    try:
        if SHEET_NAME in workbook.sheetnames:
            workbook.remove(workbook[SHEET_NAME])
        worksheet = workbook.create_sheet(SHEET_NAME)
        worksheet.append(HEADERS)

        for raw_value, observed_count in sorted(observed_sdgs.items(), key=lambda item: (-item[1], item[0])):
            preserved_goal_id = existing_mapping.get(raw_value)
            worksheet.append(
                [
                    raw_value,
                    observed_count,
                    preserved_goal_id if preserved_goal_id is not None else map_sdg_goal_id(raw_value),
                ]
            )

        for cell in worksheet[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        worksheet.column_dimensions["A"].width = 64
        worksheet.column_dimensions["B"].width = 16
        worksheet.column_dimensions["C"].width = 16
        worksheet.freeze_panes = "A2"

        table = Table(displayName=TABLE_NAME, ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium15",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

        workbook.save(workbook_path)
        return observed_sdgs
    finally:
        workbook.close()


# Executa a sincronizacao da tabela de SDGs observadas.
def main() -> int:
    observed_sdgs = sync_sdg_observed_mapping(DEFAULT_OUTPUT_PATH)
    strip_worksheet_autofilters(DEFAULT_OUTPUT_PATH)
    validate_reference_dataset(DEFAULT_OUTPUT_PATH)
    print(f"sdg observed mapping synchronized successfully: {DEFAULT_OUTPUT_PATH}")
    print(f"unique_observed_sdgs={len(observed_sdgs)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
