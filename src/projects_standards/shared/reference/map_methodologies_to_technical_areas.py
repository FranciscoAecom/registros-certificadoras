# Objetivo do script:
# Vincular metodologias do reference_dataset.xlsx a uma area tecnica da UNFCCC de forma deterministica e auditavel.
# Processo:
# 1. Definir NAME_OVERRIDES para mapeamentos manuais explicitos.
# 2. Carregar aba methodologies do reference_dataset.xlsx.
# 3. Iterar todas as linhas de metodologia.
# 4. Tentar mapeamento deterministico: nome exato primeiro, depois pattern matching.
# 5. Preencher coluna technical_area_id com base em overrides + pattern matching.
# 6. Gerar trilha de auditoria das decisoes de mapeamento.


from __future__ import annotations

import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parents[3]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


from src.projects_standards.shared.reference.build_reference_dataset import (  # noqa: E402
    DEFAULT_OUTPUT_PATH,
    strip_worksheet_autofilters,
    validate_reference_dataset,
)


METHODOLOGY_SHEET = "methodologies"
TARGET_HEADERS = [
    "technical_area_id",
]


NAME_OVERRIDES = {
    "bio-oil geological storage": ("TA 16.1", "high", "exact_name_ccs"),
    "biogenic carbon capture and storage": ("TA 16.1", "high", "exact_name_ccs"),
    "biomass geological storage": ("TA 16.1", "high", "exact_name_ccs"),
    "direct air capture": ("TA 16.1", "high", "exact_name_daccs"),
    "subsurface biomass carbon removal and storage": ("TA 16.1", "high", "exact_name_ccs"),
    "geologically stored carbon": ("TA 16.1", "high", "exact_name_ccs"),
    "geologically stored carbon, 2024": ("TA 16.1", "high", "exact_name_ccs"),
    "enhanced weathering in agriculture": ("TA 17.1", "medium", "exact_name_other_removals_approx"),
    "ocean alkalinity enhancement from coastal outfalls": ("TA 17.1", "low", "exact_name_other_removals_approx"),
    "river alkalinity enhancement": ("TA 17.1", "low", "exact_name_other_removals_approx"),
    "wastewater alkalinity enhancement": ("TA 17.1", "low", "exact_name_other_removals_approx"),
    "carbonated materials": ("TA 4.1", "medium", "exact_name_mineral_carbonation"),
    "wooden building elements": ("TA 6.1", "medium", "exact_name_building_material"),
    "biochar production and storage": ("TA 17.1", "high", "exact_name_biochar"),
    "soil amendment": ("TA 17.1", "medium", "exact_name_soil_carbon"),
    "terrestrial storage of biomass": ("TA 17.1", "high", "exact_name_biological_storage"),
    "reforestation": ("TA 14.1", "high", "exact_name_arr"),
    "climate action reserve forest project protocol": ("TA 17.1", "high", "exact_name_forest_protocol"),
    "restoration of california deltaic and coastal wetlands": ("TA 17.1", "high", "exact_name_wetland_restoration"),
    "ecy compliance offset protocol: u.s. forest projects": ("TA 17.1", "medium", "exact_name_forest_projects"),
    "emission reductions in rice management systems": ("TA 15.1", "high", "exact_name_rice"),
    "emissions reductions through anti-idling technologies": ("TA 7.1", "high", "exact_name_transport"),
    "nitrous oxide abatement protocol": ("TA 5.2", "medium", "exact_name_n2o"),
    "truck stop electrification": ("TA 7.1", "high", "exact_name_transport"),
    "m/lu/f-w01: modular methodology for climate change mitigation activities on forest lands and wetlands": ("TA 17.1", "high", "exact_name_forest_wetlands"),
    "ntc 6208 - mitigation actions in the land use, change in land use and forestry sector at the rural level, incorporating social and biodiversity considerations": ("TA 17.1", "high", "exact_name_land_use_forestry"),
    "gs advanced hull coatings v2.": ("TA 7.1", "medium", "exact_name_shipping"),
    "gs alternative ignition coal fires v1.": ("TA 8.1", "medium", "exact_name_coal"),
    "gs baseline and monitoring methodology biodigester v1.": ("TA 13.2", "medium", "exact_name_biodigester"),
    "gs ms microscale electrification and energization v1.": ("TA 1.2", "medium", "exact_name_micro_electrification"),
    "gs methodology for improved cook stoves and kitchen regimes v1.": ("TA 3.1", "high", "exact_name_cooking"),
    "gs methodology for improved cook stoves and kitchen regimes v2.": ("TA 3.1", "high", "exact_name_cooking"),
    "gs tpddtec v 1.": ("TA 3.1", "medium", "exact_name_decentralized_thermal"),
    "gs tpddtec v 2.": ("TA 3.1", "medium", "exact_name_decentralized_thermal"),
    "gs tpddtec v 3.": ("TA 3.1", "medium", "exact_name_decentralized_thermal"),
    "gs tpddtec v3.1": ("TA 3.1", "medium", "exact_name_decentralized_thermal"),
    "gs water access and wash methodology v1.": ("TA 3.1", "medium", "exact_name_water_access"),
    "methane emission reduction by adjusted water management practice in rice cultivation": ("TA 15.1", "high", "exact_name_rice"),
    "methodology for metered & measured energy cooking devices": ("TA 3.1", "high", "exact_name_cooking"),
    "soil organic carbon framework methodology": ("TA 17.1", "high", "exact_name_soil_carbon"),
    "scm0004": ("TA 17.1", "medium", "exact_name_ecosystem_regeneration"),
    "scm0007": ("TA 17.1", "low", "exact_name_freshwater_removal_approx"),
    "vcs v1 project specific": (None, "low", "non_methodology_tool_placeholder"),
    "not provided": (None, "low", "placeholder"),
    "other": (None, "low", "placeholder"),
}


CODE_OVERRIDES = {
    "ACM0008": ("TA 8.1", "high", "official_code_scope"),
    "ACM0009": ("TA 1.1", "high", "official_code_scope"),
    "ACM0011": ("TA 1.1", "high", "official_code_scope"),
    "AM0009": ("TA 10.1", "high", "official_code_scope"),
    "AM0010": ("TA 13.1", "high", "official_code_scope"),
    "AM0014": ("TA 1.1", "high", "official_code_scope"),
    "AM0016": ("TA 13.2", "high", "official_code_scope"),
    "AM0024": ("TA 4.1", "high", "official_code_scope"),
    "AM0025": ("TA 13.1", "high", "official_code_scope"),
    "AM0029": ("TA 1.1", "high", "official_code_scope"),
    "AM0034": ("TA 5.2", "high", "official_code_scope"),
    "AM0046": ("TA 3.1", "high", "official_code_scope"),
    "AM0047": ("TA 5.1", "high", "official_code_scope"),
    "AM0059": ("TA 9.1", "high", "official_code_scope"),
    "AMS-III.AU": ("TA 15.1", "high", "official_code_scope"),
    "AMS-III.B": ("TA 1.1", "medium", "code_inference"),
    "AMS-III.BB": ("TA 2.1", "high", "official_code_scope"),
    "AMS-III.F": ("TA 13.1", "high", "official_code_scope"),
    "AMS-III.J": ("TA 5.1", "medium", "code_inference"),
    "AMS-III.M": ("TA 5.1", "medium", "code_inference"),
    "AR-ACM0002": ("TA 14.1", "high", "arr_code_family"),
    "AR-AM0002": ("TA 14.1", "high", "arr_code_family"),
    "AR-AM0003": ("TA 14.1", "high", "arr_code_family"),
    "AR-AM0005": ("TA 14.1", "high", "arr_code_family"),
    "AR-AM0007": ("TA 14.1", "high", "arr_code_family"),
    "AR-AM0010": ("TA 14.1", "high", "arr_code_family"),
    "AR-AM0014": ("TA 14.1", "high", "arr_code_family"),
    "AR-AMS0001": ("TA 14.1", "high", "arr_code_family"),
    "AR-AMS0002": ("TA 14.1", "high", "arr_code_family"),
    "AR-AMS0003": ("TA 14.1", "high", "arr_code_family"),
    "AR-AMS0004": ("TA 14.1", "high", "arr_code_family"),
    "AR-AMS0005": ("TA 14.1", "high", "arr_code_family"),
    "AR-AMS0006": ("TA 14.1", "high", "arr_code_family"),
    "AR-AMS0007": ("TA 14.1", "high", "arr_code_family"),
    "BCR0003": ("TA 17.1", "high", "exact_code_land_conservation"),
    "BCR0005": ("TA 17.1", "high", "exact_code_land_conservation"),
    "M000": ("TA 14.1", "high", "exact_code_arr"),
    "M001": ("TA 14.1", "high", "exact_code_arr"),
    "VM0003": ("TA 17.1", "high", "exact_code_forest_management"),
    "VM0009": ("TA 17.1", "high", "exact_code_redd"),
    "VM0010": ("TA 17.1", "high", "exact_code_ifm"),
    "VM0012": ("TA 17.1", "high", "exact_code_ifm"),
    "VM0013": ("TA 7.1", "high", "official_code_scope"),
    "VM0014": ("TA 8.1", "high", "official_code_scope"),
    "VM0015": ("TA 17.1", "high", "official_code_scope"),
    "VM0017": ("TA 15.1", "high", "official_code_scope"),
    "VM0019": ("TA 7.1", "high", "official_code_scope"),
    "VM0021": ("TA 17.1", "high", "official_code_scope"),
    "VM0023": ("TA 5.1", "high", "official_code_scope"),
    "VM0024": ("TA 17.1", "high", "official_code_scope"),
    "VM0026": ("TA 17.1", "high", "official_code_scope"),
    "VM0032": ("TA 17.1", "high", "official_code_scope"),
    "VM0033": ("TA 17.1", "high", "official_code_scope"),
    "VM0034": ("TA 17.1", "high", "exact_code_forest_management"),
    "VM0039": ("TA 6.1", "high", "official_code_scope"),
    "VM0040": ("TA 5.1", "high", "official_code_scope"),
    "VM0043": ("TA 4.1", "high", "official_code_scope"),
    "VM0048": ("TA 17.1", "high", "official_code_scope"),
    "VM0049": ("TA 16.1", "high", "official_code_scope"),
    "VMR0002": ("TA 8.1", "high", "official_code_scope"),
    "VMR0003": ("TA 13.2", "high", "official_code_scope"),
    "VMR0008": ("TA 13.1", "high", "code_inference"),
}


KEYWORD_RULES = [
    ("TA 11.2", "high", "keyword_hcfc22", [r"hcfc-22", r"hfc-23", r"refrigerant gas production"]),
    ("TA 11.1", "high", "keyword_fluorinated_gases", [r"\bsf6\b", r"\bhfc\b", r"refriger", r"ozone deplet", r"semiconductor", r"cover gas", r"foam", r"fluor"]),
    ("TA 9.1", "high", "keyword_aluminium_magnesium", [r"magnesium", r"aluminium", r"aluminum", r"\bpfc\b"]),
    ("TA 5.2", "high", "keyword_nitric_adipic_caprolactam", [r"caprolactam", r"adipic", r"nitric acid", r"acido nitrico", r"acido adipico", r"nitrous oxide"]),
    ("TA 10.1", "high", "keyword_oil_gas", [r"oil and gas", r"oleo e gas", r"petroleo e gas", r"natural gas systems", r"pneumatic controller", r"field gas", r"gas flar", r"orphaned oil and gas wells", r"oog", r"ldar", r"petroleum gas"]),
    ("TA 8.1", "high", "keyword_mining", [r"mine methane", r"metano de mina", r"coal mine", r"mina de carvao", r"trona", r"mining", r"coal fires"]),
    ("TA 13.2", "high", "keyword_manure", [r"manure", r"esterco", r"estiercol", r"dejetos animais", r"animal manure", r"livestock", r"biodigest"]),
    ("TA 13.1", "high", "keyword_waste_water", [r"wastewater", r"aguas residu", r"landfill", r"aterro", r"compost", r"solid waste", r"residu", r"biogas management", r"sewage", r"waste treatment", r"recycling", r"anaerobic digestion", r"water purification"]),
    ("TA 7.1", "high", "keyword_transport", [r"transport", r"transit", r"vehicle", r"cable car", r"tachograph", r"freight", r"truck", r"hull coating"]),
    ("TA 14.1", "high", "keyword_arr", [r"afforestation", r"reforestation", r"\barr\b", r"reveget", r"forest restoration"]),
    ("TA 17.1", "high", "keyword_land_removals", [r"redd", r"improved forest management", r"\bifm\b", r"conserv", r"peatland", r"turfeira", r"grassland", r"shrubland", r"avoided conversion", r"biochar", r"high mountain", r"savana", r"soil organic carbon", r"wetland restoration", r"humedal", r"thicket"]),
    ("TA 15.1", "high", "keyword_agriculture", [r"\bagric", r"cropland", r"pasture", r"pastagem", r"fertiliz", r"farm", r"rice", r"arroz", r"land management", r"regenerative"]),
    ("TA 4.1", "high", "keyword_cement_lime", [r"cement", r"cimento", r"clinker", r"cal\b", r"cal virgem", r"lime", r"carbonated material", r"quicklime"]),
    ("TA 5.1", "high", "keyword_chemical_manufacturing", [r"biodiesel", r"charcoal", r"carvao vegetal", r"bio-oil", r"biocombust", r"chemical", r"quimic", r"ammonia", r"urea", r"paper", r"cardboard", r"fiberboard", r"fibreboard", r"brick", r"industrial process", r"industrial facility"]),
    ("TA 16.1", "high", "keyword_ccs", [r"carbon capture and storage", r"captura e armazenamento", r"geological", r"geologic", r"daccs", r"beccs", r"direct air capture", r"armazenamento geologico"]),
    ("TA 2.1", "high", "keyword_distribution", [r"transmission and distribution", r"transmissao e distribuicao", r"power distribution"]),
    ("TA 3.1", "high", "keyword_demand_efficiency", [r"efficiency", r"eficiencia", r"lighting", r"cook", r"household", r"appliance", r"building", r"electrification of communities", r"water access"]),
    ("TA 1.2", "high", "keyword_renewables", [r"renewable", r"renovavel", r"renovable", r"solar", r"wind", r"hydro", r"geothermal", r"electricity generation by the user", r"mini-grid", r"rural communities"]),
    ("TA 1.1", "high", "keyword_thermal_generation", [r"thermal energy", r"cogeneration", r"biomass", r"fossil fuel", r"district heating", r"waste energy recovery", r"heat generation", r"power-only plants", r"fuel switch", r"combined cycle", r"electricity and heat", r"gas based energy generation"]),
]


def normalize_text(text: object) -> str:
    value = "" if text is None else str(text)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return value.lower().strip()


def extract_code(name: str) -> str | None:
    text = normalize_text(name).replace("_", " ").replace(":", " ")
    patterns = [
        r"\b(ar-acm\d{4})\b",
        r"\b(ar-am\d{4})\b",
        r"\b(ar-ams\d{4})\b",
        r"\b(acm\d{4})\b",
        r"\b(am\d{4})\b",
        r"\b(ams-[ivx]+\.[a-z]{1,3})\b",
        r"\b(vmr\d{4})\b",
        r"\b(vm\d{4})\b",
        r"\b(scm\d{4})\b",
        r"\b(bcr\d{4})\b",
        r"\b(m\d{3})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1).upper()
    return None


def classify_methodology(name: str, description: str, inherited_codes: dict[str, str]) -> tuple[str | None, str, str]:
    normalized_name = normalize_text(name)
    combined_text = normalize_text(f"{name} || {description}")
    code = extract_code(name)

    if normalized_name in NAME_OVERRIDES:
        return NAME_OVERRIDES[normalized_name]

    if code in CODE_OVERRIDES:
        return CODE_OVERRIDES[code]

    for technical_area_id, confidence, note, patterns in KEYWORD_RULES:
        if any(re.search(pattern, combined_text) for pattern in patterns):
            return technical_area_id, confidence, note

    if code in inherited_codes:
        return inherited_codes[code], "medium", f"inherited_code_{code}"

    return None, "low", "unmapped_after_rules"


def build_code_inheritance(rows: list[dict[str, object]]) -> dict[str, str]:
    code_to_matches: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        code = extract_code(str(row["project_methodology"]))
        if not code:
            continue
        technical_area_id, _, note = classify_methodology(
            str(row["project_methodology"]),
            str(row["methodology_description"]),
            {},
        )
        if technical_area_id and note != "unmapped_after_rules":
            code_to_matches[code].append(technical_area_id)
    return {code: values[0] for code, values in code_to_matches.items() if len(set(values)) == 1}


def map_methodologies_to_technical_areas(workbook_path: Path) -> Counter:
    workbook = load_workbook(workbook_path)
    try:
        worksheet = workbook[METHODOLOGY_SHEET]
        headers = [cell.value for cell in worksheet[1]]
        header_index = {header: index + 1 for index, header in enumerate(headers)}

        rows = []
        for row_index in range(2, worksheet.max_row + 1):
            rows.append(
                {
                    "row_index": row_index,
                    "standard_acronym": worksheet.cell(row=row_index, column=header_index["standard_acronym"]).value,
                    "project_methodology": worksheet.cell(row=row_index, column=header_index["project_methodology"]).value,
                    "methodology_description": worksheet.cell(
                        row=row_index,
                        column=header_index.get("methodology_description_pt", header_index["methodology_description"]),
                    ).value,
                }
            )

        inherited_codes = build_code_inheritance(rows)

        next_column = worksheet.max_column + 1
        for header in TARGET_HEADERS:
            if header not in header_index:
                worksheet.cell(row=1, column=next_column, value=header)
                header_index[header] = next_column
                next_column += 1

        removable_headers = [
            "sectoral_scope_id",
            "technical_area_match_confidence",
            "technical_area_match_note",
        ]
        for removable_header in removable_headers:
            if removable_header in header_index:
                column_index = header_index[removable_header]
                worksheet.delete_cols(column_index)
                headers = [cell.value for cell in worksheet[1]]
                header_index = {header: index + 1 for index, header in enumerate(headers)}

        stats = Counter()
        for row in rows:
            technical_area_id, _, _ = classify_methodology(
                str(row["project_methodology"] or ""),
                str(row["methodology_description"] or ""),
                inherited_codes,
            )
            worksheet.cell(row=row["row_index"], column=header_index["technical_area_id"], value=technical_area_id)
            stats["total_rows"] += 1
            if technical_area_id:
                stats["mapped_rows"] += 1
                stats[f"mapped_{technical_area_id}"] += 1
            else:
                stats["unmapped_rows"] += 1

        for table_name in worksheet.tables.keys():
            worksheet.tables[table_name].ref = worksheet.dimensions

        workbook.save(workbook_path)
        return stats
    finally:
        workbook.close()


def main() -> int:
    stats = map_methodologies_to_technical_areas(DEFAULT_OUTPUT_PATH)
    strip_worksheet_autofilters(DEFAULT_OUTPUT_PATH)
    validate_reference_dataset(DEFAULT_OUTPUT_PATH)
    print(f"methodologies mapped successfully: {DEFAULT_OUTPUT_PATH}")
    print(f"total_rows={stats['total_rows']} mapped_rows={stats['mapped_rows']} unmapped_rows={stats['unmapped_rows']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
