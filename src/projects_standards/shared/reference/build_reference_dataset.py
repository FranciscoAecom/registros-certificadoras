# Objetivo do script:
# Consolidar as tabelas estruturadas dos workbooks de referencia em um unico arquivo XLSX compativel com o Excel.
# Processo:
# 1. Ler argumentos CLI opcionais.
# 2. Descobrir fontes de referencia legadas em _legacy_sources/.
# 3. Mapear cada fonte para aba de destino no workbook consolidado.
# 4. Consolidar todas as tabelas em reference_dataset.xlsx.
# 5. Aplicar formatacao Excel (estilos de tabela, alinhamento, autofiltros).
# 6. Validar compatibilidade OOXML.


from __future__ import annotations

import argparse
import sys
import tempfile
import zipfile
from copy import copy
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parents[3]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


REFERENCE_DIR = ROOT_DIR / "data" / "project_standards" / "00_reference"
DEFAULT_OUTPUT_PATH = REFERENCE_DIR / "reference_dataset.xlsx"
LEGACY_REFERENCE_DIR = REFERENCE_DIR / "_legacy_sources"
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


# Retorna a configuracao fixa de origem e destino para cada tabela do consolidado.
def build_sheet_plan() -> list[dict[str, str | Path]]:
    return [
        {
            "path": LEGACY_REFERENCE_DIR / "certificadoras" / "certificadoras.xlsx",
            "source_sheet": "certificadora",
            "source_table": "tb_certificadora",
            "target_sheet": "standards_catalog",
            "target_table": "tb_ref_standards_catalog",
            "header_map": {
                "sigla": "standard_acronym",
                "nome": "standard_name",
            },
        },
        {
            "path": LEGACY_REFERENCE_DIR / "certificadoras" / "certificadoras.xlsx",
            "source_sheet": "certificadora_status",
            "source_table": "Tabela2",
            "target_sheet": "standards_status",
            "target_table": "tb_ref_standards_status",
            "header_map": {
                "sigla": "standard_acronym",
                "status_certificadora": "status_standard",
            },
        },
        {
            "path": LEGACY_REFERENCE_DIR / "certificadoras" / "certificadoras.xlsx",
            "source_sheet": "pipeline_status",
            "source_table": "tb_pipelineStatus",
            "target_sheet": "common_pipeline_status",
            "target_table": "tb_ref_common_pipeline_status",
            "header_map": {},
        },
        {
            "path": LEGACY_REFERENCE_DIR / "metodologias" / "metodologias.xlsx",
            "source_sheet": "standard_methodologies",
            "source_table": "Tabela1",
            "target_sheet": "methodologies",
            "target_table": "tb_ref_methodologies",
            "header_map": {
                "sigla": "standard_acronym",
            },
        },
        {
            "path": LEGACY_REFERENCE_DIR / "paises" / "paises.xlsx",
            "source_sheet": "pais_padrao",
            "source_table": "tb_paisPadrao",
            "target_sheet": "countries_standard",
            "target_table": "tb_ref_countries_standard",
            "header_map": {},
        },
        {
            "path": LEGACY_REFERENCE_DIR / "paises" / "paises.xlsx",
            "source_sheet": "pais_certificadoras",
            "source_table": "tb_mapPaisCertificadora",
            "target_sheet": "countries_observed_mapping",
            "target_table": "tb_ref_countries_observed_mapping",
            "header_map": {
                "Pais_certificadora": "country_raw",
            },
        },
        {
            "path": LEGACY_REFERENCE_DIR / "sdgs" / "sdgs.xlsx",
            "source_sheet": "goals",
            "source_table": "Tabela2",
            "target_sheet": "sdg_goals",
            "target_table": "tb_ref_sdg_goals",
            "header_map": {},
        },
        {
            "path": LEGACY_REFERENCE_DIR / "sdgs" / "sdgs.xlsx",
            "source_sheet": "targets",
            "source_table": "Tabela1",
            "target_sheet": "sdg_targets",
            "target_table": "tb_ref_sdg_targets",
            "header_map": {},
        },
    ]


# Copia o intervalo da tabela de origem para a aba de destino preservando valores, estilos e hyperlinks.
def copy_table_range(
    source_workbook_path: Path,
    source_sheet_name: str,
    source_table_name: str,
    target_sheet,
    header_map: dict[str, str] | None = None,
) -> None:
    source_wb = load_workbook(source_workbook_path)
    try:
        source_ws = source_wb[source_sheet_name]
        source_table = source_ws.tables[source_table_name]
        start_cell, end_cell = source_table.ref.split(":")
        start_col = source_ws[start_cell].column
        start_row = source_ws[start_cell].row
        end_col = source_ws[end_cell].column
        end_row = source_ws[end_cell].row

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                source_cell = source_ws.cell(row=row, column=col)
                target_cell = target_sheet.cell(
                    row=row - start_row + 1,
                    column=col - start_col + 1,
                    value=source_cell.value,
                )
                if row == start_row and isinstance(target_cell.value, str) and header_map:
                    target_cell.value = header_map.get(target_cell.value, target_cell.value)
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.font:
                    target_cell.font = copy(source_cell.font)
                if source_cell.fill:
                    target_cell.fill = copy(source_cell.fill)
                if source_cell.border:
                    target_cell.border = copy(source_cell.border)
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
                if source_cell.protection:
                    target_cell.protection = copy(source_cell.protection)
                if source_cell.hyperlink:
                    target_cell._hyperlink = copy(source_cell.hyperlink)
                if source_cell.comment:
                    target_cell.comment = copy(source_cell.comment)

        for col_idx in range(start_col, end_col + 1):
            source_letter = source_ws.cell(row=start_row, column=col_idx).column_letter
            target_letter = target_sheet.cell(row=1, column=col_idx - start_col + 1).column_letter
            width = source_ws.column_dimensions[source_letter].width
            if width:
                target_sheet.column_dimensions[target_letter].width = width
    finally:
        source_wb.close()


# Configura a tabela estruturada de destino sem adicionar AutoFilter no worksheet fora da tabela.
def add_target_table(target_sheet, target_table_name: str) -> None:
    table = Table(displayName=target_table_name, ref=target_sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium15",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    target_sheet.add_table(table)


# Itera pelas tabelas de uma aba retornando sempre o nome e o objeto Table.
def iter_worksheet_tables(worksheet):
    for table_name in worksheet.tables.keys():
        yield table_name, worksheet.tables[table_name]


# Remove AutoFilter do worksheet e os defined names auxiliares para manter compatibilidade com tabelas no Excel.
def strip_worksheet_autofilters(workbook_path: Path) -> None:
    ET.register_namespace("", MAIN_NS)
    source_path = workbook_path
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        temp_path = Path(tmp_file.name)

    with zipfile.ZipFile(source_path, "r") as source_zip, zipfile.ZipFile(
        temp_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as target_zip:
        for item in source_zip.infolist():
            data = source_zip.read(item.filename)

            if item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml"):
                root = ET.fromstring(data)
                for auto_filter in root.findall(f"{{{MAIN_NS}}}autoFilter"):
                    root.remove(auto_filter)
                data = ET.tostring(root, encoding="utf-8", xml_declaration=False)

            elif item.filename == "xl/workbook.xml":
                root = ET.fromstring(data)
                defined_names = root.find(f"{{{MAIN_NS}}}definedNames")
                if defined_names is not None:
                    for defined_name in list(defined_names):
                        if defined_name.attrib.get("name") == "_xlnm._FilterDatabase":
                            defined_names.remove(defined_name)
                    if len(defined_names) == 0:
                        root.remove(defined_names)
                data = ET.tostring(root, encoding="utf-8", xml_declaration=False)

            target_zip.writestr(item, data)

    temp_path.replace(workbook_path)


# Sincroniza os nomes das colunas das tabelas com os cabecalhos reais da primeira linha de cada aba.
def sync_table_headers(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    try:
        for worksheet in workbook.worksheets:
            headers = [cell.value for cell in worksheet[1]]
            for _, table in iter_worksheet_tables(worksheet):
                table.ref = worksheet.dimensions
                if len(table.tableColumns) != len(headers):
                    table.tableColumns = []
                    table.autoFilter = None
                    table._initialise_columns()
                header_slice = headers[: len(table.tableColumns)]
                for column, header in zip(table.tableColumns, header_slice):
                    column.name = "" if header is None else str(header)
        workbook.save(workbook_path)
    finally:
        workbook.close()


# Valida se os metadados das tabelas continuam coerentes com os cabecalhos e com o worksheet.
def validate_reference_dataset(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    try:
        for worksheet in workbook.worksheets:
            if not worksheet.tables:
                continue
            headers = [cell.value for cell in worksheet[1]]
            for table_name, table in iter_worksheet_tables(worksheet):
                if table.headerRowCount != 1:
                    raise ValueError(
                        f"A tabela {table_name} da aba {worksheet.title} nao possui headerRowCount=1."
                    )
                column_names = [column.name for column in table.tableColumns]
                expected_headers = headers[: len(column_names)]
                if column_names != expected_headers:
                    raise ValueError(
                        f"A tabela {table_name} da aba {worksheet.title} esta dessincronizada dos cabecalhos."
                    )
    finally:
        workbook.close()


# Gera o workbook consolidado com uma aba por tabela estruturada de referencia.
def build_reference_dataset(output_path: Path) -> None:
    source_specs = build_sheet_plan()
    missing_sources = sorted({str(spec["path"]) for spec in source_specs if not Path(spec["path"]).exists()})

    if missing_sources:
        if not output_path.exists():
            raise FileNotFoundError(
                "Nao foi possivel reconstruir o reference_dataset.xlsx porque os workbooks-fonte nao estao disponiveis. "
                f"Fontes ausentes: {', '.join(missing_sources)}"
            )
        sync_table_headers(output_path)
        strip_worksheet_autofilters(output_path)
        validate_reference_dataset(output_path)
        return

    workbook = Workbook()
    workbook.remove(workbook.active)

    for spec in source_specs:
        worksheet = workbook.create_sheet(spec["target_sheet"])
        copy_table_range(
            source_workbook_path=spec["path"],
            source_sheet_name=str(spec["source_sheet"]),
            source_table_name=str(spec["source_table"]),
            target_sheet=worksheet,
            header_map=spec.get("header_map", {}),
        )
        for cell in worksheet[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.freeze_panes = "A2"
        add_target_table(worksheet, str(spec["target_table"]))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    sync_table_headers(output_path)
    strip_worksheet_autofilters(output_path)
    validate_reference_dataset(output_path)


# Monta o parser de argumentos do builder do workbook consolidado.
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Consolida as tabelas estruturadas dos workbooks de referencia em um unico XLSX."
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT_PATH),
        help=f"Arquivo XLSX de saida. Padrao: {DEFAULT_OUTPUT_PATH}",
    )
    return parser


# Executa o script de consolidacao do workbook de referencia.
def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    output_path = Path(args.output)
    build_reference_dataset(output_path)
    print(f"reference dataset gerado com sucesso: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
