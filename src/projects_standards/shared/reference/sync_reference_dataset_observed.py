# Objetivo do script:
# Sincronizar no workbook consolidado as formas observadas em silver para paises, metodologias, status e SDGs.
# Processo:
# 1. Ler argumentos CLI opcionais.
# 2. Escanear todos os datasets silver em 02_silver/.
# 3. Coletar formas observadas de paises -> aba countries_observed_mapping.
# 4. Coletar formas observadas de metodologias -> aba methodologies.
# 5. Coletar formas observadas de status -> aba standards_status.
# 6. Delegar sync de SDGs ao modulo sync_sdg_observed_mapping.
# 7. Preservar colunas curadas manualmente e inserir apenas novas formas.
# 8. Preencher correspondencias exatas seguras automaticamente.


from __future__ import annotations

import argparse
import json
import sys
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parents[3]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


from src.projects_standards.shared.reference.build_reference_dataset import (  # noqa: E402
    DEFAULT_OUTPUT_PATH,
    strip_worksheet_autofilters,
    sync_table_headers,
    validate_reference_dataset,
)
from src.projects_standards.shared.reference.sync_sdg_observed_mapping import sync_sdg_observed_mapping  # noqa: E402


SILVER_DIR = ROOT_DIR / "data" / "project_standards" / "02_silver"
COUNTRIES_SHEET = "countries_standard"
COUNTRIES_OBSERVED_SHEET = "countries_observed_mapping"
METHODOLOGIES_SHEET = "methodologies"
STATUS_SHEET = "standards_status"


# Normaliza texto para joins tolerantes sem alterar o valor salvo no workbook.
def normalize_lookup_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if text == "":
        return None
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.casefold().split())


# Garante que um valor seja tratado como lista simples.
def ensure_list(value: Any) -> list[Any]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return value
    return [value]


# Percorre todos os datasets silver e coleta as formas observadas que precisam entrar na referencia.
def collect_observed_values() -> dict[str, Any]:
    countries: set[str] = set()
    methodologies: set[tuple[str, str]] = set()
    statuses: set[tuple[str, str, str]] = set()

    for dataset_path in sorted(SILVER_DIR.rglob("allprojects.json")):
        payload = json.loads(dataset_path.read_text(encoding="utf-8"))
        for project in payload.get("projects") or []:
            standard_acronym = project.get("standard_acronym")
            if standard_acronym not in (None, ""):
                standard_acronym = str(standard_acronym).strip()

            country = project.get("country")
            if country not in (None, ""):
                countries.add(str(country).strip())

            for methodology in ensure_list(project.get("project_methodology")):
                if standard_acronym and methodology not in (None, ""):
                    methodologies.add((standard_acronym, str(methodology).strip()))

            voluntary_status = project.get("project_voluntary_status")
            if standard_acronym and voluntary_status not in (None, ""):
                statuses.add((standard_acronym, "voluntary", str(voluntary_status).strip()))

            regulatory_status = project.get("project_regulatory_status")
            if standard_acronym and regulatory_status not in (None, ""):
                statuses.add((standard_acronym, "regulatory", str(regulatory_status).strip()))

    return {
        "countries": sorted(countries),
        "methodologies": sorted(methodologies),
        "statuses": sorted(statuses),
    }


# Monta um lookup tolerante de pais observado -> nome padrao em ingles.
def build_country_standard_lookup(workbook) -> dict[str, str]:
    worksheet = workbook[COUNTRIES_SHEET]
    headers = [cell.value for cell in worksheet[1]]
    header_index = {str(header): index for index, header in enumerate(headers) if header is not None}

    lookup: dict[str, str] = {}
    english_idx = header_index["name_en"]
    candidate_headers = ["name_pt", "name_en", "name_es", "alpha_2", "alpha_3", "numeric"]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        english_name = row[english_idx]
        if english_name in (None, ""):
            continue
        english_name_str = str(english_name).strip()
        for header in candidate_headers:
            value = row[header_index[header]]
            key = normalize_lookup_text(value)
            if key is not None:
                lookup[key] = english_name_str
    return lookup


# Atualiza a aba de paises observados com novas formas e preenche correspondencias exatas quando possivel.
def sync_countries_sheet(workbook, observed_countries: list[str]) -> int:
    worksheet = workbook[COUNTRIES_OBSERVED_SHEET]
    headers = [cell.value for cell in worksheet[1]]
    header_index = {str(header): index + 1 for index, header in enumerate(headers) if header is not None}
    country_lookup = build_country_standard_lookup(workbook)

    existing_rows: dict[str, int] = {}
    changed = 0
    for row_index in range(2, worksheet.max_row + 1):
        raw_value = worksheet.cell(row=row_index, column=header_index["country_raw"]).value
        raw_key = normalize_lookup_text(raw_value)
        if raw_key is None:
            continue
        existing_rows[raw_key] = row_index
        standard_cell = worksheet.cell(row=row_index, column=header_index["country_standard"])
        if standard_cell.value in (None, ""):
            inferred_standard = country_lookup.get(raw_key)
            if inferred_standard is not None:
                standard_cell.value = inferred_standard
                changed += 1

    inserted = 0
    for raw_country in observed_countries:
        raw_key = normalize_lookup_text(raw_country)
        if raw_key is None or raw_key in existing_rows:
            continue
        worksheet.append([raw_country, country_lookup.get(raw_key)])
        existing_rows[raw_key] = worksheet.max_row
        inserted += 1

    if worksheet.tables:
        for table_name in worksheet.tables.keys():
            worksheet.tables[table_name].ref = worksheet.dimensions
    return inserted + changed


# Atualiza a aba de metodologias preservando as colunas curadas e adicionando apenas novas combinacoes observadas.
def sync_methodologies_sheet(workbook, observed_methodologies: list[tuple[str, str]]) -> int:
    worksheet = workbook[METHODOLOGIES_SHEET]
    headers = [cell.value for cell in worksheet[1]]
    header_index = {str(header): index + 1 for index, header in enumerate(headers) if header is not None}

    existing: set[tuple[str, str]] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        standard = normalize_lookup_text(row[header_index["standard_acronym"] - 1] if len(row) >= header_index["standard_acronym"] else None)
        methodology = normalize_lookup_text(row[header_index["project_methodology"] - 1] if len(row) >= header_index["project_methodology"] else None)
        if standard and methodology:
            existing.add((standard, methodology))

    inserted = 0
    for standard_acronym, project_methodology in observed_methodologies:
        key = (normalize_lookup_text(standard_acronym), normalize_lookup_text(project_methodology))
        if None in key or key in existing:
            continue
        new_row = [None] * len(headers)
        new_row[header_index["standard_acronym"] - 1] = standard_acronym
        new_row[header_index["project_methodology"] - 1] = project_methodology
        worksheet.append(new_row)
        existing.add((key[0], key[1]))
        inserted += 1

    if worksheet.tables:
        for table_name in worksheet.tables.keys():
            worksheet.tables[table_name].ref = worksheet.dimensions
    return inserted


# Atualiza a aba de status preservando as colunas curadas e adicionando novos status observados.
def sync_status_sheet(workbook, observed_statuses: list[tuple[str, str, str]]) -> int:
    worksheet = workbook[STATUS_SHEET]
    headers = [cell.value for cell in worksheet[1]]
    header_index = {str(header): index + 1 for index, header in enumerate(headers) if header is not None}

    existing: set[tuple[str, str, str]] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        standard = normalize_lookup_text(row[header_index["standard_acronym"] - 1] if len(row) >= header_index["standard_acronym"] else None)
        market = normalize_lookup_text(row[header_index["market"] - 1] if len(row) >= header_index["market"] else None)
        status = normalize_lookup_text(row[header_index["status_standard"] - 1] if len(row) >= header_index["status_standard"] else None)
        if standard and market and status:
            existing.add((standard, market, status))

    inserted = 0
    for standard_acronym, market, status_standard in observed_statuses:
        key = (
            normalize_lookup_text(standard_acronym),
            normalize_lookup_text(market),
            normalize_lookup_text(status_standard),
        )
        if None in key or key in existing:
            continue
        new_row = [None] * len(headers)
        new_row[header_index["standard_acronym"] - 1] = standard_acronym
        new_row[header_index["market"] - 1] = market
        new_row[header_index["status_standard"] - 1] = status_standard
        worksheet.append(new_row)
        existing.add((key[0], key[1], key[2]))
        inserted += 1

    if worksheet.tables:
        for table_name in worksheet.tables.keys():
            worksheet.tables[table_name].ref = worksheet.dimensions
    return inserted


# Sincroniza o workbook consolidado com todas as formas observadas nos datasets silver.
def sync_reference_dataset_observed(workbook_path: Path = DEFAULT_OUTPUT_PATH) -> dict[str, int]:
    observed = collect_observed_values()
    try:
        workbook = load_workbook(workbook_path)
        try:
            changes = {
                "countries_observed_mapping_changed": sync_countries_sheet(workbook, observed["countries"]),
                "methodologies_inserted": sync_methodologies_sheet(workbook, observed["methodologies"]),
                "standards_status_inserted": sync_status_sheet(workbook, observed["statuses"]),
            }
            workbook.save(workbook_path)
        finally:
            workbook.close()

        observed_sdgs = sync_sdg_observed_mapping(workbook_path)
    except PermissionError as exc:
        raise PermissionError(
            f"Nao foi possivel salvar o workbook de referencia em {workbook_path}. "
            "Feche o arquivo no Excel e execute novamente a construcao da silver ou a sincronizacao consolidada."
        ) from exc

    sync_table_headers(workbook_path)
    strip_worksheet_autofilters(workbook_path)
    validate_reference_dataset(workbook_path)
    changes["sdg_observed_rows"] = len(observed_sdgs)
    return changes


# Monta o parser do sincronizador consolidado de observados.
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Sincroniza no reference_dataset.xlsx as formas observadas em todos os datasets silver."
    )
    parser.add_argument(
        "--workbook",
        default=str(DEFAULT_OUTPUT_PATH),
        help=f"Workbook consolidado de referencia. Padrao: {DEFAULT_OUTPUT_PATH}",
    )
    return parser


# Executa a sincronizacao completa das formas observadas para o workbook consolidado.
def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    workbook_path = Path(args.workbook)
    changes = sync_reference_dataset_observed(workbook_path)
    print(f"reference observed mappings synchronized successfully: {workbook_path}")
    print(json.dumps(changes, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
