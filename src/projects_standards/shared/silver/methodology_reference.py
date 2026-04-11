# Objetivo do modulo:
# Normalizar e sincronizar a referencia de metodologias observadas nos datasets silver.
# Processo:
# 1. Importar caminhos do workbook de referencia e utilitarios regex.
# 2. Implementar normalize_project_methodology() para converter escalar ou lista em lista limpa.
# 3. Tratar metodologias delimitadas (virgula, ponto-e-virgula, pipe).
# 4. Usado pelo sync de metodologias para catalogar formas observadas na silver.


import argparse
import json
import re
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from .missing import normalize_missing
from .normalize import scalar_or_list, unique_non_empty


ROOT_DIR = Path(__file__).resolve().parents[4]
REFERENCE_WORKBOOK_PATH = ROOT_DIR / "data" / "project_standards" / "00_reference" / "reference_dataset.xlsx"
METHODOLOGY_SHEET_NAME = "methodologies"
METHODOLOGY_TABLE_NAME = "tb_ref_methodologies"


# Converte um valor escalar ou lista em uma lista limpa de metodologias observadas.
def normalize_project_methodology(value: Any, split_pattern: str | None = None) -> Any:
    items: list[str] = []

    def visit(current: Any) -> None:
        if isinstance(current, list):
            for item in current:
                visit(item)
            return
        normalized = normalize_missing(current)
        if normalized is None:
            return
        text = str(normalized).strip()
        if split_pattern:
            parts = [part.strip() for part in re.split(split_pattern, text) if part.strip()]
            if parts:
                items.extend(parts)
                return
        items.append(text)

    visit(value)
    unique_items = unique_non_empty(items)
    return scalar_or_list(unique_items)


# Normaliza um valor escalar ou lista para uma lista simples de metodologias nao vazias.
def normalize_methodology_values(value: Any) -> list[str]:
    normalized = normalize_project_methodology(value)
    if normalized is None:
        return []
    return normalized if isinstance(normalized, list) else [normalized]


# Extrai linhas de metodologia do projeto usando a sigla e a metodologia canonica da silver.
def collect_default_methodology_rows(project: dict[str, Any]) -> list[tuple[str, str]]:
    sigla = normalize_missing(project.get("standard_acronym"))
    if sigla is None:
        return []

    methodologies = normalize_methodology_values(project.get("project_methodology"))
    return [(str(sigla).strip(), methodology) for methodology in methodologies]


# Insere novas linhas de metodologia na tabela de referencia sem duplicar combinacoes existentes.
def sync_methodology_reference_rows(
    rows: list[tuple[str, str]],
    workbook_path: Path = REFERENCE_WORKBOOK_PATH,
) -> int:
    if not rows:
        return 0

    workbook = load_workbook(workbook_path)
    try:
        worksheet = workbook[METHODOLOGY_SHEET_NAME]
        table = worksheet.tables[METHODOLOGY_TABLE_NAME]

        existing: set[tuple[str, str]] = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            sigla = normalize_missing(row[0] if len(row) > 0 else None)
            methodology = normalize_missing(row[1] if len(row) > 1 else None)
            if sigla is None or methodology is None:
                continue
            existing.add((str(sigla).strip(), str(methodology).strip()))

        pending = sorted(
            {
                (str(sigla).strip(), str(methodology).strip())
                for sigla, methodology in rows
                if normalize_missing(sigla) is not None and normalize_missing(methodology) is not None
            },
            key=lambda row: (row[0], row[1]),
        )

        inserted = 0
        for row in pending:
            if row in existing:
                continue
            worksheet.append(list(row))
            existing.add(row)
            inserted += 1

        if inserted > 0:
            table.ref = worksheet.dimensions
            try:
                workbook.save(workbook_path)
            except PermissionError as exc:
                raise PermissionError(
                    f"Nao foi possivel salvar a planilha de metodologias em {workbook_path}. "
                    "Feche o arquivo no Excel e execute novamente."
                ) from exc
        return inserted
    finally:
        workbook.close()


# Sincroniza as metodologias de uma lista de projetos ja carregados em memoria.
def sync_methodology_reference_projects(
    reference_name: str,
    projects: list[dict[str, Any]],
    collector: Callable[[dict[str, Any]], list[tuple[str, str]]] = collect_default_methodology_rows,
    workbook_path: Path = REFERENCE_WORKBOOK_PATH,
) -> int:
    rows: list[tuple[str, str]] = []
    for project in projects:
        rows.extend(collector(project))
    inserted = sync_methodology_reference_rows(rows, workbook_path=workbook_path)
    print(f"metodologias de referencia inseridas para {reference_name}: {inserted}")
    return inserted


# Monta o parser usado pelos scripts de sincronizacao de metodologias.
def build_methodology_sync_parser(display_name: str, bronze_slug: str) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=f"Sincroniza a tabela de referencia de metodologias da {display_name} a partir do dataset silver."
    )
    parser.add_argument("--date", required=True, help="Data de referencia no formato YYYYMMDD.")
    parser.add_argument(
        "--input",
        default=None,
        help=f"Arquivo de entrada. Padrao: data/project_standards/02_silver/{bronze_slug}/<date>/allprojects.json",
    )
    return parser


# Executa o fluxo completo do script de sincronizacao de metodologias.
def run_methodology_sync(config: dict[str, Any]) -> int:
    parser = build_methodology_sync_parser(config["display_name"], config["bronze_slug"])
    args = parser.parse_args()
    input_path = (
        Path(args.input)
        if args.input
        else ROOT_DIR / "data" / "project_standards" / "02_silver" / config["bronze_slug"] / args.date / "allprojects.json"
    )
    if not input_path.exists():
        raise SystemExit(f"Dataset silver nao encontrado: {input_path}")

    payload = json.loads(input_path.read_text(encoding="utf-8"))
    projects = payload.get("projects")
    if not isinstance(projects, list):
        raise SystemExit(f"Campo 'projects' invalido em {input_path}")

    sync_methodology_reference_projects(
        reference_name=config["reference_name"],
        projects=projects,
        collector=config.get("methodology_collector", collect_default_methodology_rows),
        workbook_path=config.get("workbook_path", REFERENCE_WORKBOOK_PATH),
    )
    print(f"sincronizacao de metodologias concluida: {input_path}")
    return 0
