# Objetivo do modulo:
# Sincronizar a tabela de referencia de status das certificadoras a partir dos datasets silver.
# Processo:
# 1. Importar workbook de referencia e utilitarios openpyxl.
# 2. Implementar load_standard_sigla() para resolver sigla da certificadora no standards_catalog.
# 3. Iterar linhas para casar nome de referencia e retornar sigla correspondente.
# 4. Usado pelo sync de status para identificar a certificadora no reference_dataset.xlsx.


import argparse
import json
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from .missing import normalize_missing


ROOT_DIR = Path(__file__).resolve().parents[4]
REFERENCE_WORKBOOK_PATH = ROOT_DIR / "data" / "project_standards" / "00_reference" / "reference_dataset.xlsx"
CERT_SHEET_NAME = "standards_catalog"
STATUS_SHEET_NAME = "standards_status"
STATUS_TABLE_NAME = "tb_ref_standards_status"
STATUS_KEY_HEADERS = ("standard_acronym", "market", "status_standard")


# Resolve a sigla da certificadora a partir da aba principal da planilha.
def load_standard_sigla(reference_name: str, workbook_path: Path = REFERENCE_WORKBOOK_PATH) -> str:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[CERT_SHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        header_index = {header: index for index, header in enumerate(headers)}
        nome_idx = header_index.get("standard_name")
        sigla_idx = header_index.get("standard_acronym")
        if nome_idx is None or sigla_idx is None:
            raise SystemExit("Colunas 'standard_name' e 'standard_acronym' nao encontradas na aba 'standards_catalog'.")

        for row in rows:
            row_name = row[nome_idx]
            if row_name is None:
                continue
            if str(row_name).strip().lower() == reference_name.strip().lower():
                sigla = row[sigla_idx]
                if sigla in (None, ""):
                    raise SystemExit(f"Sigla vazia para a certificadora {reference_name}.")
                return str(sigla).strip()
    finally:
        workbook.close()

    raise SystemExit(f"Standard nao encontrada na aba 'standards_catalog': {reference_name}")


# Extrai linhas de status usando a regra padrao de mercado por campo.
def collect_default_status_rows(project: dict[str, Any]) -> list[tuple[str, str, str | None]]:
    rows: list[tuple[str, str, str | None]] = []
    voluntary_status = normalize_missing(project.get("project_voluntary_status"))
    regulatory_status = normalize_missing(project.get("project_regulatory_status"))

    if voluntary_status is not None:
        rows.append(("voluntary", str(voluntary_status).strip(), None))
    if regulatory_status is not None:
        rows.append(("regulatory", str(regulatory_status).strip(), None))
    return rows


# Insere novas linhas de status na tabela de referencia sem duplicar combinacoes existentes.
def sync_status_reference_rows(
    reference_name: str,
    rows: list[tuple[str, str, str | None]],
    workbook_path: Path = REFERENCE_WORKBOOK_PATH,
) -> int:
    if not rows:
        return 0

    sigla = load_standard_sigla(reference_name, workbook_path=workbook_path)
    workbook = load_workbook(workbook_path)
    try:
        worksheet = workbook[STATUS_SHEET_NAME]
        table = worksheet.tables[STATUS_TABLE_NAME]
        headers = [cell.value for cell in worksheet[1]]
        header_index = {str(header).strip(): index for index, header in enumerate(headers) if header is not None}
        missing_headers = [header for header in STATUS_KEY_HEADERS if header not in header_index]
        if missing_headers:
            raise SystemExit(
                "Colunas obrigatorias ausentes na aba 'standards_status': "
                + ", ".join(missing_headers)
            )

        existing: set[tuple[str, str, str]] = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            current_sigla = normalize_missing(
                row[header_index["standard_acronym"]] if len(row) > header_index["standard_acronym"] else None
            )
            current_market = normalize_missing(
                row[header_index["market"]] if len(row) > header_index["market"] else None
            )
            current_status = normalize_missing(
                row[header_index["status_standard"]]
                if len(row) > header_index["status_standard"]
                else None
            )
            if current_sigla is None or current_market is None or current_status is None:
                continue
            existing.add((str(current_sigla).strip(), str(current_market).strip(), str(current_status).strip()))

        inserted = 0
        pending = sorted(
            {
                (
                    sigla,
                    str(market).strip(),
                    str(status).strip(),
                    description if description not in ("", None) else None,
                )
                for market, status, description in rows
                if normalize_missing(market) is not None and normalize_missing(status) is not None
            }
        )

        for row_sigla, market, status, description in pending:
            key = (row_sigla, market, status)
            if key in existing:
                continue
            new_row = [None] * len(headers)
            new_row[header_index["standard_acronym"]] = row_sigla
            new_row[header_index["market"]] = market
            new_row[header_index["status_standard"]] = status
            if "status_description" in header_index:
                new_row[header_index["status_description"]] = description
            worksheet.append(new_row)
            existing.add(key)
            inserted += 1

        if inserted > 0:
            table.ref = f"A1:{worksheet.cell(row=1, column=worksheet.max_column).column_letter}{worksheet.max_row}"
            workbook.save(workbook_path)
        return inserted
    finally:
        workbook.close()


# Sincroniza os status de uma lista de projetos ja carregados em memoria.
def sync_status_reference_projects(
    reference_name: str,
    projects: list[dict[str, Any]],
    collector: Callable[[dict[str, Any]], list[tuple[str, str, str | None]]] = collect_default_status_rows,
    workbook_path: Path = REFERENCE_WORKBOOK_PATH,
) -> int:
    rows: list[tuple[str, str, str | None]] = []
    for project in projects:
        rows.extend(collector(project))
    inserted = sync_status_reference_rows(reference_name, rows, workbook_path=workbook_path)
    print(f"status de referencia inseridos para {reference_name}: {inserted}")
    return inserted


# Monta o parser usado pelos scripts de sincronizacao de status.
def build_status_sync_parser(display_name: str, bronze_slug: str) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=f"Sincroniza a tabela de referencia de status da {display_name} a partir do dataset silver."
    )
    parser.add_argument("--date", required=True, help="Data de referencia no formato YYYYMMDD.")
    parser.add_argument(
        "--input",
        default=None,
        help=f"Arquivo de entrada. Padrao: data/project_standards/02_silver/{bronze_slug}/<date>/allprojects.json",
    )
    return parser


# Executa o fluxo completo do script de sincronizacao de status.
def run_status_sync(config: dict[str, Any]) -> int:
    parser = build_status_sync_parser(config["display_name"], config["bronze_slug"])
    args = parser.parse_args()
    input_path = (
        Path(args.input)
        if args.input
        else ROOT_DIR / "data" / "project_standards" / "02_silver" / config["bronze_slug"] / args.date / "allprojects.json"
    )
    if not input_path.exists():
        raise SystemExit(f"Dataset silver nao encontrado: {input_path}")

    payload = json.loads(input_path.read_text(encoding="utf-8"))
    projects = payload.get("projects")
    if not isinstance(projects, list):
        raise SystemExit(f"Campo 'projects' invalido em {input_path}")

    sync_status_reference_projects(
        reference_name=config["reference_name"],
        projects=projects,
        collector=config.get("status_collector", collect_default_status_rows),
        workbook_path=config.get("workbook_path", REFERENCE_WORKBOOK_PATH),
    )
    print(f"sincronizacao de status concluida: {input_path}")
    return 0
