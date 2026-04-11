# Objetivo do modulo:
# Sincronizar o mapeamento de paises observados nos datasets silver.
# Processo:
# 1. Importar caminhos do workbook de referencia e nomes de abas.
# 2. Implementar collect_default_country_rows() para extrair valores brutos de pais por projeto.
# 3. Implementar normalize_country_key() para comparacao tolerante entre linguas e codigos.
# 4. Usado pelo sync de paises para catalogar formas observadas na silver.


import argparse
import json
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from .missing import normalize_missing


ROOT_DIR = Path(__file__).resolve().parents[4]
REFERENCE_WORKBOOK_PATH = ROOT_DIR / "data" / "project_standards" / "00_reference" / "reference_dataset.xlsx"
STANDARD_COUNTRY_SHEET_NAME = "countries_standard"
STANDARD_COUNTRY_TABLE_NAME = "tb_ref_countries_standard"
COUNTRY_MAPPING_SHEET_NAMES = ("countries_observed_mapping",)
COUNTRY_MAPPING_TABLE_NAMES = ("tb_ref_countries_observed_mapping",)


# Extrai os paises observados no dataset silver preservando a forma bruta vinda da certificadora.
def collect_default_country_rows(project: dict[str, Any]) -> list[str]:
    country = normalize_missing(project.get("country"))
    if country is None:
        return []
    return [str(country).strip()]


# Normaliza chaves textuais para comparacao tolerante entre idiomas e codigos.
def normalize_country_key(value: Any) -> str | None:
    normalized = normalize_missing(value)
    if normalized is None:
        return None
    return str(normalized).strip().casefold()


# Localiza a planilha configurada para o mapeamento dos paises da certificadora.
def get_country_mapping_sheet(workbook: Any) -> Any:
    for sheet_name in COUNTRY_MAPPING_SHEET_NAMES:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    available = ", ".join(workbook.sheetnames)
    raise KeyError(
        "Aba de mapeamento de paises nao encontrada. Esperadas: "
        + ", ".join(COUNTRY_MAPPING_SHEET_NAMES)
        + f". Disponiveis: {available}"
    )


# Localiza a tabela configurada para o mapeamento dos paises da certificadora.
def get_country_mapping_table(worksheet: Any) -> Any:
    for table_name in COUNTRY_MAPPING_TABLE_NAMES:
        if table_name in worksheet.tables:
            return worksheet.tables[table_name]
    available = ", ".join(sorted(worksheet.tables.keys()))
    raise KeyError(
        "Tabela de mapeamento de paises nao encontrada na aba "
        f"{worksheet.title}. Disponiveis: {available}"
    )


# Monta o indice de paises padrao a partir da tabela de referencia.
def build_standard_country_lookup(workbook: Any) -> dict[str, str]:
    worksheet = workbook[STANDARD_COUNTRY_SHEET_NAME]
    table = worksheet.tables[STANDARD_COUNTRY_TABLE_NAME]
    headers = [cell.value for cell in worksheet[1]]
    header_index = {str(header): index for index, header in enumerate(headers) if header}

    required_headers = ["name_pt", "name_en", "name_es", "alpha_2", "alpha_3", "numeric"]
    missing_headers = [header for header in required_headers if header not in header_index]
    if missing_headers:
        raise KeyError(
            "Colunas obrigatorias ausentes em tb_ref_countries_standard: "
            + ", ".join(missing_headers)
        )

    lookup: dict[str, str] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        english_name = normalize_missing(row[header_index["name_en"]])
        if english_name is None:
            continue
        english_name_str = str(english_name).strip()
        for header in required_headers:
            key = normalize_country_key(row[header_index[header]])
            if key is not None:
                lookup[key] = english_name_str
    return lookup


# Insere novos mapeamentos de pais na referencia e preenche o pais padrao quando possivel.
def sync_country_reference_rows(rows: list[str], workbook_path: Path = REFERENCE_WORKBOOK_PATH) -> int:
    if not rows:
        return 0

    workbook = load_workbook(workbook_path)
    try:
        mapping_worksheet = get_country_mapping_sheet(workbook)
        mapping_table = get_country_mapping_table(mapping_worksheet)
        standard_lookup = build_standard_country_lookup(workbook)

        existing: set[str] = set()
        changed = 0
        for row_index, row in enumerate(mapping_worksheet.iter_rows(min_row=2, values_only=True), start=2):
            current_country = normalize_country_key(row[0] if len(row) > 0 else None)
            if current_country is None:
                continue
            existing.add(current_country)
            current_standard = normalize_missing(row[1] if len(row) > 1 else None)
            expected_standard = standard_lookup.get(current_country)
            if current_standard is None and expected_standard is not None:
                mapping_worksheet.cell(row=row_index, column=2, value=expected_standard)
                changed += 1

        inserted = 0
        pending = sorted(
            {
                str(country).strip()
                for country in rows
                if normalize_missing(country) is not None
            }
        )

        for country in pending:
            country_key = normalize_country_key(country)
            if country_key is None or country_key in existing:
                continue
            standard_country = standard_lookup.get(country_key)
            mapping_worksheet.append([country, standard_country])
            existing.add(country_key)
            inserted += 1

        if inserted > 0 or changed > 0:
            mapping_table.ref = f"A1:B{mapping_worksheet.max_row}"
            try:
                workbook.save(workbook_path)
            except PermissionError as exc:
                raise PermissionError(
                    f"Nao foi possivel salvar a planilha de paises em {workbook_path}. "
                    "Feche o arquivo no Excel e execute novamente."
                ) from exc
        return inserted + changed
    finally:
        workbook.close()


# Sincroniza os paises de uma lista de projetos ja carregados em memoria.
def sync_country_reference_projects(
    reference_name: str,
    projects: list[dict[str, Any]],
    collector: Callable[[dict[str, Any]], list[str]] = collect_default_country_rows,
    workbook_path: Path = REFERENCE_WORKBOOK_PATH,
) -> int:
    rows: list[str] = []
    for project in projects:
        rows.extend(collector(project))
    changed = sync_country_reference_rows(rows, workbook_path=workbook_path)
    print(f"paises de referencia inseridos/atualizados a partir de {reference_name}: {changed}")
    return changed


# Monta o parser usado pelos scripts de sincronizacao de paises.
def build_country_sync_parser(display_name: str, bronze_slug: str) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=f"Sincroniza a tabela de referencia de paises da {display_name} a partir do dataset silver."
    )
    parser.add_argument("--date", required=True, help="Data de referencia no formato YYYYMMDD.")
    parser.add_argument(
        "--input",
        default=None,
        help=f"Arquivo de entrada. Padrao: data/project_standards/02_silver/{bronze_slug}/<date>/allprojects.json",
    )
    return parser


# Executa o fluxo completo do script de sincronizacao de paises.
def run_country_sync(config: dict[str, Any]) -> int:
    parser = build_country_sync_parser(config["display_name"], config["bronze_slug"])
    args = parser.parse_args()
    input_path = (
        Path(args.input)
        if args.input
        else ROOT_DIR / "data" / "project_standards" / "02_silver" / config["bronze_slug"] / args.date / "allprojects.json"
    )
    if not input_path.exists():
        raise SystemExit(f"Dataset silver nao encontrado: {input_path}")

    payload = json.loads(input_path.read_text(encoding="utf-8"))
    projects = payload.get("projects")
    if not isinstance(projects, list):
        raise SystemExit(f"Campo 'projects' invalido em {input_path}")

    sync_country_reference_projects(
        reference_name=config["reference_name"],
        projects=projects,
        collector=config.get("country_collector", collect_default_country_rows),
        workbook_path=config.get("workbook_path", REFERENCE_WORKBOOK_PATH),
    )
    print(f"sincronizacao de paises concluida: {input_path}")
    return 0
