# Objetivo do modulo:
# Reunir utilitarios compartilhados para gerar mapeamentos e datasets da camada silver a partir dos snapshots bronze das certificadoras.
# Processo:
# 1. Importar modulos especializados (dates, missing, normalize, numbers, quality_checks).
# 2. Definir dataclass FieldSpec e constantes de configuracao (fracao de amostra, minimos).
# 3. Carregar workbook de referencia (reference_dataset.xlsx).
# 4. Prover funcoes de orquestracao para gerar datasets e mapeamentos silver.
# 5. Descompactar automaticamente o bronze zipado antes do processamento e recompactar ao final.
# 6. Servir como hub central importado pelos build_silver_dataset.py de cada certificadora.


import argparse
import base64
import io
import json
import math
import random
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from ..archive_data import pack_directory, unpack_archive
from .dates import format_datetime_iso, parse_date
from .missing import normalize_missing
from .normalize import ensure_list, normalize_record, scalar_or_list, unique_non_empty
from .numbers import parse_number
from .quality_checks import collect_quality_issues
from ..reference.sync_reference_dataset_observed import sync_reference_dataset_observed


ROOT_DIR = Path(__file__).resolve().parents[4]
GUIDE_PATH = ROOT_DIR / "docs" / "agentes" / "guia_silver.md"
REFERENCE_PATH = ROOT_DIR / "data" / "project_standards" / "00_reference" / "reference_dataset.xlsx"
REFERENCE_SHEET_NAME = "standards_catalog"
DEFAULT_SAMPLE_FRACTION = 0.10
DEFAULT_SAMPLE_MIN_FILES = 10
DEFAULT_SAMPLE_LARGEST_RATIO = 0.50
GEOJSON_GEOMETRY_TYPES = {
    "Point",
    "MultiPoint",
    "LineString",
    "MultiLineString",
    "Polygon",
    "MultiPolygon",
    "GeometryCollection",
}
GEOMETRY_CANDIDATE_PATHS = (
    "detail_data.geometry",
    "detail_data.geom",
    "detail_data.geojson",
    "detail_data.location.geometry",
    "detail_data.project.geometry",
    "detail_data.project.geojson",
    "detail_data.project.boundary",
    "detail_data.project.boundaries",
    "detail_data.project.polygon",
    "detail_data.boundary",
    "detail_data.boundaries",
    "detail_data.polygon",
    "detail_data.footprint",
    "list_data.geometry",
    "list_data.geom",
    "list_data.geojson",
    "list_data.location.geometry",
    "list_data.project.geometry",
)
LATITUDE_CANDIDATE_PATHS = (
    "detail_data.location.latitude",
    "detail_data.location.lat",
    "detail_data.latitude",
    "detail_data.lat",
    "detail_data.project.latitude",
    "detail_data.project.lat",
    "list_data.latitude",
    "list_data.lat",
)
LONGITUDE_CANDIDATE_PATHS = (
    "detail_data.location.longitude",
    "detail_data.location.lng",
    "detail_data.location.lon",
    "detail_data.longitude",
    "detail_data.lng",
    "detail_data.lon",
    "detail_data.project.longitude",
    "detail_data.project.lng",
    "detail_data.project.lon",
    "list_data.longitude",
    "list_data.lng",
    "list_data.lon",
)


@dataclass(frozen=True)
class FieldSpec:
    section: str
    name: str


@dataclass(frozen=True)
class CandidateSource:
    source_section: str
    source_path: str
    rule_type: str
    notes: str
    extractor: Callable[[dict[str, Any], Path], Any]


# Carrega um arquivo JSON e valida o tipo do payload.
def load_json(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise SystemExit(f"JSON invalido em {path}")
    return payload


# Percorre um caminho com notacao por ponto dentro do payload.
def get_path(payload: dict[str, Any], dotted_path: str) -> Any:
    current: Any = payload
    for part in dotted_path.split("."):
        if isinstance(current, list):
            if not part.isdigit():
                return None
            index = int(part)
            if index < 0 or index >= len(current):
                return None
            current = current[index]
            continue
        if not isinstance(current, dict) or part not in current:
            return None
        current = current[part]
    return current


# Deriva o caminho relativo do arquivo de detalhe no filesystem.
def derive_bronze_file_path(file_path: Path) -> str:
    return str(file_path.relative_to(ROOT_DIR)).replace("\\", "/")


# Retorna o primeiro valor util entre varios caminhos candidatos.
def first_non_empty(payload: dict[str, Any], *paths: str) -> Any:
    for path in paths:
        value = normalize_missing(get_path(payload, path))
        if value is not None:
            return value
    return None


# Converte um valor em float usando o parser numerico compartilhado para coordenadas.
def _to_coordinate(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    parsed = parse_number(value, number_kind="coordinate", number_style="auto")
    if isinstance(parsed, (int, float)):
        return float(parsed)
    return None


# Converte listas de pontos em coordenadas [longitude, latitude].
def _parse_vertex_list(value: Any) -> list[list[float]] | None:
    if not isinstance(value, list):
        return None
    coords: list[list[float]] = []
    for item in value:
        lon: float | None = None
        lat: float | None = None
        if isinstance(item, dict):
            lon = _to_coordinate(
                first_non_empty(
                    {"point": item},
                    "point.longitude",
                    "point.lng",
                    "point.lon",
                    "point.x",
                )
            )
            lat = _to_coordinate(
                first_non_empty(
                    {"point": item},
                    "point.latitude",
                    "point.lat",
                    "point.y",
                )
            )
        elif isinstance(item, list) and len(item) >= 2:
            lon = _to_coordinate(item[0])
            lat = _to_coordinate(item[1])
        if lon is None or lat is None:
            continue
        coords.append([lon, lat])
    if len(coords) < 3:
        return None
    if coords[0] != coords[-1]:
        coords.append(coords[0])
    return coords


# Identifica quando uma estrutura representa uma posicao [longitude, latitude].
def _is_position(value: Any) -> bool:
    if not isinstance(value, list) or len(value) < 2:
        return False
    return _to_coordinate(value[0]) is not None and _to_coordinate(value[1]) is not None


# Converte uma posicao em coordenadas [longitude, latitude].
def _position_to_lon_lat(value: list[Any]) -> list[float] | None:
    if len(value) < 2:
        return None
    lon = _to_coordinate(value[0])
    lat = _to_coordinate(value[1])
    if lon is None or lat is None:
        return None
    return [lon, lat]


# Fecha um anel de poligono, repetindo o primeiro ponto no final quando necessario.
def _close_ring(ring: list[list[float]]) -> list[list[float]]:
    if ring and ring[0] != ring[-1]:
        ring = [*ring, ring[0]]
    return ring


# Infere geometria GeoJSON-like a partir de um bloco coordinates sem type explicito.
def _infer_geometry_from_coordinates(coordinates: Any) -> dict[str, Any] | None:
    if not isinstance(coordinates, list) or not coordinates:
        return None

    # Point: [lon, lat]
    if _is_position(coordinates):
        point = _position_to_lon_lat(coordinates)
        if point is not None:
            return {"type": "Point", "coordinates": point}

    # Lista de posicoes: pode ser linha ou anel de poligono.
    if all(_is_position(item) for item in coordinates):
        points = [_position_to_lon_lat(item) for item in coordinates]
        normalized_points = [point for point in points if point is not None]
        if len(normalized_points) != len(coordinates):
            return None
        if len(normalized_points) >= 4:
            return {"type": "Polygon", "coordinates": [_close_ring(normalized_points)]}
        return {"type": "LineString", "coordinates": normalized_points}

    # Lista de anéis/linhas.
    if all(isinstance(item, list) and item for item in coordinates):
        if all(all(_is_position(point) for point in item) for item in coordinates):
            rings: list[list[list[float]]] = []
            for item in coordinates:
                points = [_position_to_lon_lat(point) for point in item]
                normalized_points = [point for point in points if point is not None]
                if len(normalized_points) != len(item):
                    return None
                rings.append(_close_ring(normalized_points))
            return {"type": "Polygon", "coordinates": rings}

        # Lista de polígonos (MultiPolygon)
        if all(
            isinstance(poly, list)
            and poly
            and all(isinstance(ring, list) and ring and all(_is_position(point) for point in ring) for ring in poly)
            for poly in coordinates
        ):
            multipolygon: list[list[list[list[float]]]] = []
            for poly in coordinates:
                polygon_rings: list[list[list[float]]] = []
                for ring in poly:
                    points = [_position_to_lon_lat(point) for point in ring]
                    normalized_points = [point for point in points if point is not None]
                    if len(normalized_points) != len(ring):
                        return None
                    polygon_rings.append(_close_ring(normalized_points))
                multipolygon.append(polygon_rings)
            return {"type": "MultiPolygon", "coordinates": multipolygon}

    return None


# Converte texto de coordenadas KML em lista de pontos [lon, lat].
def _parse_kml_coordinates_text(text: str) -> list[list[float]]:
    points: list[list[float]] = []
    for chunk in re.split(r"\s+", text.strip()):
        if not chunk:
            continue
        parts = chunk.split(",")
        if len(parts) < 2:
            continue
        lon = _to_coordinate(parts[0])
        lat = _to_coordinate(parts[1])
        if lon is None or lat is None:
            continue
        points.append([lon, lat])
    return points


# Extrai geometria GeoJSON-like a partir de conteudo KML bruto.
def _extract_geometry_from_kml_text(kml_text: str) -> dict[str, Any] | None:
    if not isinstance(kml_text, str) or not kml_text.strip():
        return None
    try:
        root = ET.fromstring(kml_text)
    except ET.ParseError:
        return None

    coordinate_nodes = root.findall(".//{*}coordinates")
    rings: list[list[list[float]]] = []
    for node in coordinate_nodes:
        if node.text is None:
            continue
        points = _parse_kml_coordinates_text(node.text)
        if len(points) >= 4:
            rings.append(_close_ring(points))

    if not rings:
        return None
    if len(rings) == 1:
        return {"type": "Polygon", "coordinates": [rings[0]]}
    # Cada ring vira um poligono simples dentro do MultiPolygon: [[ring]]
    return {"type": "MultiPolygon", "coordinates": [[ring] for ring in rings]}


# Extrai geometria a partir do bloco bronze detail_data.spatial_documents.
def _extract_geometry_from_spatial_documents(spatial_documents: Any) -> dict[str, Any] | None:
    if not isinstance(spatial_documents, list):
        return None
    for document in spatial_documents:
        if not isinstance(document, dict):
            continue
        content = document.get("content")
        if not isinstance(content, str) or not content.strip():
            continue
        encoding = str(document.get("contentEncoding") or "").strip().lower()
        if encoding == "base64":
            # KMZ: descompacta o primeiro arquivo .kml disponível.
            try:
                kmz_bytes = base64.b64decode(content, validate=False)
                with zipfile.ZipFile(io.BytesIO(kmz_bytes)) as kmz:
                    kml_candidates = [name for name in kmz.namelist() if name.lower().endswith(".kml")]
                    for kml_name in kml_candidates:
                        kml_text = kmz.read(kml_name).decode("utf-8", errors="replace")
                        geometry = _extract_geometry_from_kml_text(kml_text)
                        if geometry is not None:
                            return geometry
            except Exception:  # noqa: BLE001
                continue
            continue

        geometry = _extract_geometry_from_kml_text(content)
        if geometry is not None:
            return geometry
    return None


# Converte diferentes formatos de geometria observados no bronze para um objeto GeoJSON-like.
def _coerce_geometry(value: Any) -> dict[str, Any] | None:
    if value in (None, "", [], {}):
        return None

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            return None
        return _coerce_geometry(parsed)

    if isinstance(value, dict):
        geometry_type = normalize_missing(value.get("type"))
        if isinstance(geometry_type, str):
            clean_type = geometry_type.strip()
            if clean_type in GEOJSON_GEOMETRY_TYPES and "coordinates" in value:
                return {"type": clean_type, "coordinates": value.get("coordinates")}
            if clean_type == "Feature":
                return _coerce_geometry(value.get("geometry"))
            if clean_type == "FeatureCollection":
                features = value.get("features")
                if isinstance(features, list):
                    for feature in features:
                        geometry = _coerce_geometry(feature)
                        if geometry is not None:
                            return geometry

        nested_geometry = value.get("geometry")
        if nested_geometry is not None:
            geometry = _coerce_geometry(nested_geometry)
            if geometry is not None:
                return geometry

        vertices = value.get("vertices")
        if vertices is not None:
            coordinates = _parse_vertex_list(vertices)
            if coordinates is not None:
                return {"type": "Polygon", "coordinates": [coordinates]}

        coordinates = value.get("coordinates")
        if isinstance(coordinates, list):
            inferred = _infer_geometry_from_coordinates(coordinates)
            if inferred is not None:
                return inferred
            if len(coordinates) >= 3 and isinstance(coordinates[0], (dict, list)):
                ring = _parse_vertex_list(coordinates)
                if ring is not None:
                    return {"type": "Polygon", "coordinates": [ring]}

        lat = _to_coordinate(first_non_empty({"point": value}, "point.latitude", "point.lat", "point.y"))
        lon = _to_coordinate(first_non_empty({"point": value}, "point.longitude", "point.lng", "point.lon", "point.x"))
        if lat is not None and lon is not None:
            return {"type": "Point", "coordinates": [lon, lat]}

    if isinstance(value, list):
        ring = _parse_vertex_list(value)
        if ring is not None:
            return {"type": "Polygon", "coordinates": [ring]}

    return None


# Faz uma varredura limitada por profundidade para localizar geometrias dentro do payload.
def _scan_geometry(value: Any, depth: int = 0, max_depth: int = 6) -> dict[str, Any] | None:
    if depth > max_depth:
        return None
    geometry = _coerce_geometry(value)
    if geometry is not None:
        return geometry
    if isinstance(value, dict):
        for nested in value.values():
            found = _scan_geometry(nested, depth + 1, max_depth)
            if found is not None:
                return found
    if isinstance(value, list):
        for nested in value:
            found = _scan_geometry(nested, depth + 1, max_depth)
            if found is not None:
                return found
    return None


# Extrai a melhor geometria disponivel no bronze, com fallback para Point por latitude/longitude.
def extract_project_geometry(payload: dict[str, Any], _file_path: Path | None = None) -> dict[str, Any] | None:
    spatial_geometry = _extract_geometry_from_spatial_documents(get_path(payload, "detail_data.spatial_documents"))
    if spatial_geometry is not None:
        return spatial_geometry

    for path in GEOMETRY_CANDIDATE_PATHS:
        geometry = _coerce_geometry(get_path(payload, path))
        if geometry is not None:
            return geometry

    for section in ("detail_data", "list_data"):
        section_payload = get_path(payload, section)
        geometry = _scan_geometry(section_payload)
        if geometry is not None:
            return geometry

    lat = _to_coordinate(first_non_empty(payload, *LATITUDE_CANDIDATE_PATHS))
    lon = _to_coordinate(first_non_empty(payload, *LONGITUDE_CANDIDATE_PATHS))
    if lat is not None and lon is not None:
        return {"type": "Point", "coordinates": [lon, lat]}

    return None


# Retorna o candidato padrao para mapeamento automatico de geometria.
def geometry_candidate() -> CandidateSource:
    return CandidateSource(
        source_section="detail_data/list_data",
        source_path="geometry|geojson|vertices|latitude/longitude",
        rule_type="derived",
        notes="Extracao automatica de geometria em formato GeoJSON-like com fallback para Point.",
        extractor=extract_project_geometry,
    )


# Le o guia da camada silver e extrai a ordem canonica dos campos.
def parse_guide_fields(guide_path: Path) -> list[FieldSpec]:
    lines = guide_path.read_text(encoding="utf-8").splitlines()
    in_structure = False
    current_section = ""
    fields: list[FieldSpec] = []

    for line in lines:
        if line.startswith("## "):
            if line.strip() == "## Estrutura Canonica Recomendada":
                in_structure = True
                continue
            if in_structure:
                break

        if not in_structure:
            continue

        if line.startswith("### "):
            current_section = re.sub(r"^\d+\.\s*", "", line[4:].strip())
            continue

        match = re.match(r"- `([^`]+)`", line.strip())
        if match and current_section:
            fields.append(FieldSpec(section=current_section, name=match.group(1)))

    if not fields:
        raise SystemExit(f"Nenhum campo encontrado em {guide_path}")
    return fields


# Busca a sigla da certificadora na planilha de referencias.
def load_standard_acronym(reference_path: Path, standard_name: str) -> str | None:
    workbook = load_workbook(reference_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[REFERENCE_SHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        header_index = {header: index for index, header in enumerate(headers)}

        nome_idx = header_index.get("standard_name")
        sigla_idx = header_index.get("standard_acronym")
        if nome_idx is None or sigla_idx is None:
            raise SystemExit(f"Colunas 'standard_name' e 'standard_acronym' nao encontradas em {reference_path}")

        for row in rows:
            row_name = row[nome_idx]
            if row_name is None:
                continue
            if str(row_name).strip().lower() == standard_name.strip().lower():
                sigla = row[sigla_idx]
                return str(sigla).strip() if sigla not in (None, "") else None
    finally:
        workbook.close()
    return None


# Gera uma representacao curta para o exemplo exibido no markdown.
def format_example(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    text = json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else str(value)
    return text if len(text) <= 140 else text[:137] + "..."


# Escapa pipes para manter a tabela markdown valida.
def escape_cell(value: Any) -> str:
    return str(value).replace("|", "\\|")


# Monta uma fonte candidata de caminho direto.
def path_candidate(source_section: str, source_path: str, rule_type: str = "direct", notes: str = "") -> CandidateSource:
    def extractor(payload: dict[str, Any], file_path: Path) -> Any:
        if source_section == "file_system" and source_path == "bronze_file_path":
            return derive_bronze_file_path(file_path)
        if source_section == "file_system" and source_path == "source_file_name":
            return file_path.name
        return get_path(payload, f"{source_section}.{source_path}")

    return CandidateSource(
        source_section=source_section,
        source_path=source_path,
        rule_type=rule_type,
        notes=notes,
        extractor=extractor,
    )


# Monta uma fonte candidata com transformacao aplicada sobre um caminho do payload.
def transformed_candidate(
    source_section: str,
    source_path: str,
    transform: Callable[[Any], Any],
    notes: str,
    rule_type: str,
) -> CandidateSource:
    return CandidateSource(
        source_section=source_section,
        source_path=source_path,
        rule_type=rule_type,
        notes=notes,
        extractor=lambda payload, _file_path: transform(get_path(payload, f"{source_section}.{source_path}")),
    )


# Descompacta o ZIP bronze correspondente se a pasta nao existir.
# Delega ao archive_data.pack_directory / unpack_archive.
def unpack_bronze_if_needed(bronze_dir: Path) -> bool:
    zip_path = bronze_dir.parent / f"{bronze_dir.name}.zip"
    if bronze_dir.exists():
        return False
    if not zip_path.exists():
        return False
    unpack_archive(zip_path, label="bronze", step=1, total=1)
    return True


# Recompacta o bronze em ZIP usando archive_data.pack_directory.
def repack_bronze(bronze_dir: Path) -> None:
    pack_directory(bronze_dir, label="bronze", step=1, total=1)


# Grava um payload JSON UTF-8 no caminho informado.
def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


# Monta e retorna o parser de argumentos para o script de mapeamento.
def build_mapping_parser(display_name: str, default_output_path: Path) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=f"Analisa os arquivos bronze da {display_name} e gera um mapeamento inicial para a camada silver."
    )
    parser.add_argument("--date", required=True, help="Data de referencia no formato YYYYMMDD.")
    parser.add_argument(
        "--output",
        default=str(default_output_path),
        help=f"Arquivo Markdown de saida. Padrao: {default_output_path}",
    )
    parser.add_argument("--limit", type=int, default=None, help="Limita a quantidade de arquivos de detalhe para testes.")
    parser.add_argument(
        "--sample-fraction",
        type=float,
        default=DEFAULT_SAMPLE_FRACTION,
        help=f"Fracao minima de arquivos usada na amostra do mapeamento. Padrao: {DEFAULT_SAMPLE_FRACTION}.",
    )
    parser.add_argument(
        "--sample-min-files",
        type=int,
        default=DEFAULT_SAMPLE_MIN_FILES,
        help=f"Quantidade minima de arquivos na amostra do mapeamento. Padrao: {DEFAULT_SAMPLE_MIN_FILES}.",
    )
    parser.add_argument(
        "--sample-seed",
        type=int,
        default=None,
        help="Seed opcional para a parte aleatoria da amostra. Padrao: usa a data do snapshot.",
    )
    parser.add_argument(
        "--sample-largest-ratio",
        type=float,
        default=DEFAULT_SAMPLE_LARGEST_RATIO,
        help=(
            "Fracao da amostra reservada para os maiores arquivos do snapshot. "
            f"Padrao: {DEFAULT_SAMPLE_LARGEST_RATIO}."
        ),
    )
    return parser


# Monta e retorna o parser de argumentos para o script de consolidacao.
def build_dataset_parser(display_name: str, bronze_slug: str) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=f"Consolida os arquivos bronze da {display_name} em um dataset silver unico em JSON."
    )
    parser.add_argument("--date", required=True, help="Data de referencia no formato YYYYMMDD.")
    parser.add_argument(
        "--output",
        default=None,
        help=f"Arquivo JSON de saida. Padrao: data/project_standards/02_silver/{bronze_slug}/<date>/allprojects.json",
    )
    parser.add_argument("--limit", type=int, default=None, help="Limita a quantidade de projetos processados para testes.")
    parser.add_argument(
        "--failures-output",
        default=None,
        help=f"Arquivo JSON para falhas por projeto. Padrao: src/projects_standards/{bronze_slug}/silver/logs/build_silver_dataset_failures_<date>.json",
    )
    return parser


# Valida a data informada e garante o formato YYYYMMDD.
def validate_date(value: str) -> str:
    if not re.fullmatch(r"\d{8}", value):
        raise SystemExit(f"--date invalida: {value}. Use YYYYMMDD.")
    return value


# Resolve a lista de arquivos de detalhe, com ordenacao e limite opcionais.
def resolve_detail_files(projects_dir: Path, sorter: Callable[[Path], Any], limit: int | None) -> list[Path]:
    detail_files = sorted(projects_dir.glob("*.json"), key=sorter)
    if limit is not None:
        detail_files = detail_files[:limit]
    return detail_files


# Calcula o tamanho minimo da amostra usada no mapeamento.
def calculate_sample_size(total_files: int, sample_fraction: float, sample_min_files: int) -> int:
    if total_files <= 0:
        return 0
    if total_files <= sample_min_files:
        return total_files
    return min(total_files, max(sample_min_files, math.ceil(total_files * sample_fraction)))


# Seleciona uma amostra hibrida com maiores arquivos e parte aleatoria.
def select_sample_files(
    detail_files: list[Path],
    snapshot_date: str,
    sample_fraction: float,
    sample_min_files: int,
    sample_seed: int | None,
    sample_largest_ratio: float,
    sorter: Callable[[Path], Any],
) -> tuple[list[Path], int, int, int]:
    sample_size = calculate_sample_size(
        total_files=len(detail_files),
        sample_fraction=sample_fraction,
        sample_min_files=sample_min_files,
    )
    largest_count = min(sample_size, math.ceil(sample_size * sample_largest_ratio))
    random_count = max(0, sample_size - largest_count)
    effective_seed = sample_seed if sample_seed is not None else int(snapshot_date)

    if sample_size >= len(detail_files):
        return detail_files, effective_seed, len(detail_files), 0

    files_by_size = sorted(detail_files, key=lambda path: (-path.stat().st_size, sorter(path)))
    largest_files = files_by_size[:largest_count]
    largest_set = set(largest_files)
    remaining_files = [path for path in detail_files if path not in largest_set]

    rng = random.Random(effective_seed)
    random_files: list[Path] = []
    if random_count > 0 and remaining_files:
        random_files = rng.sample(remaining_files, min(random_count, len(remaining_files)))

    sampled = sorted({*largest_files, *random_files}, key=sorter)
    return sampled, effective_seed, len(largest_files), len(random_files)


# Mede a cobertura de uma fonte candidata dentro do conjunto de arquivos analisados.
def evaluate_candidate(candidate: CandidateSource, detail_files: list[Path]) -> tuple[int, Any]:
    coverage = 0
    example = None
    for file_path in detail_files:
        payload = load_json(file_path)
        value = normalize_missing(candidate.extractor(payload, file_path))
        if value in (None, "", [], {}):
            continue
        coverage += 1
        if example is None:
            example = value
    return coverage, example


# Resolve a melhor fonte candidata para cada campo canonico.
def build_mapping_rows(fields: list[FieldSpec], detail_files: list[Path], candidate_sources: dict[str, list[CandidateSource]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for field in fields:
        candidates = candidate_sources.get(field.name, [])
        if field.name == "project_geometry" and not candidates:
            candidates = [geometry_candidate()]
        best_row = {
            "target_field": field.name,
            "section": field.section,
            "status": "unmapped",
            "source_section": "",
            "source_path": "",
            "rule_type": "unmapped",
            "coverage": 0,
            "example": "",
            "notes": "Nenhuma regra inicial configurada para este campo.",
        }

        for candidate in candidates:
            coverage, example = evaluate_candidate(candidate, detail_files)
            if coverage > best_row["coverage"]:
                best_row = {
                    "target_field": field.name,
                    "section": field.section,
                    "status": "mapped" if coverage > 0 else "unmapped",
                    "source_section": candidate.source_section,
                    "source_path": candidate.source_path,
                    "rule_type": candidate.rule_type,
                    "coverage": coverage,
                    "example": format_example(example),
                    "notes": candidate.notes if coverage > 0 else (candidate.notes or "Fonte candidata configurada, mas sem valores uteis no snapshot analisado."),
                }

        rows.append(best_row)
    return rows


# Gera o markdown final do mapeamento inicial da certificadora.
def build_mapping_markdown(
    display_name: str,
    bronze_slug: str,
    snapshot_date: str,
    detail_files: list[Path],
    rows: list[dict[str, Any]],
    *,
    total_available_files: int,
    sample_fraction: float,
    sample_min_files: int,
    sample_seed: int,
    sample_largest_ratio: float,
    sample_largest_count: int,
    sample_random_count: int,
) -> str:
    section_summary: dict[str, dict[str, int]] = {}
    for row in rows:
        section = row["section"]
        section_summary.setdefault(section, {"total": 0, "mapped": 0})
        section_summary[section]["total"] += 1
        if row["status"] == "mapped":
            section_summary[section]["mapped"] += 1

    lines = [
        f"# Mapeamento Inicial Silver da {display_name}",
        "",
        f"- Snapshot analisado: `{snapshot_date}`",
        f"- Arquivos de detalhe disponiveis no snapshot: `{total_available_files}`",
        f"- Arquivos de detalhe analisados na amostra: `{len(detail_files)}`",
        f"- Regra de amostragem: `max({sample_min_files}, ceil({sample_fraction:.0%} do snapshot))`, com limite no total disponivel",
        (
            f"- Estrategia da amostra: `{sample_largest_count}` maiores arquivos "
            f"+ `{sample_random_count}` arquivos aleatorios "
            f"(proporcao alvo para maiores arquivos: {sample_largest_ratio:.0%})"
        ),
        f"- Seed da amostra aleatoria: `{sample_seed}`",
        f"- Guia base: `{GUIDE_PATH.relative_to(ROOT_DIR).as_posix()}`",
        "",
        "## Resumo por Secao",
        "",
        "| Secao | Campos | Campos com fonte inicial |",
        "| --- | ---: | ---: |",
    ]

    for section, summary in section_summary.items():
        lines.append(f"| {section} | {summary['total']} | {summary['mapped']} |")

    lines.extend(
        [
            "",
            "## Tabela de Mapeamento Inicial",
            "",
            "| target_field | secao_guia | status | source_section | source_path | rule_type | cobertura | exemplo | notes |",
            "| --- | --- | --- | --- | --- | --- | ---: | --- | --- |",
        ]
    )

    total_files = len(detail_files)
    for row in rows:
        coverage_display = f"{row['coverage']}/{total_files}" if total_files else "0/0"
        lines.append(
            "| "
            + " | ".join(
                [
                    f"`{escape_cell(row['target_field'])}`",
                    escape_cell(row["section"]),
                    escape_cell(row["status"]),
                    escape_cell(row["source_section"]),
                    f"`{escape_cell(row['source_path'])}`" if row["source_path"] else "``",
                    escape_cell(row["rule_type"]),
                    coverage_display,
                    escape_cell(row["example"]),
                    escape_cell(row["notes"]),
                ]
            )
            + " |"
        )

    lines.extend(
        [
            "",
            "## Observacoes",
            "",
            f"- Este arquivo e um ponto de partida para refinarmos o mapeamento `bronze -> silver` da {display_name}.",
            "- Campos com status `unmapped` ainda nao tiveram uma origem confiavel encontrada no bruto analisado.",
            f"- Quando um campo permanecer sem origem confiavel no bruto da {bronze_slug}, ele deve seguir como `null` na `silver`.",
            "- Tratamento de completude, qualidade de registro e preenchimentos derivados devem ficar para a camada `gold`.",
            "- A coluna `cobertura` mostra quantos arquivos da amostra apresentaram valor util na melhor fonte candidata.",
            f"- Este documento deve ser tratado como mapeamento exploratorio ate a estabilizacao do mapeamento canonico da {display_name}.",
        ]
    )
    return "\n".join(lines) + "\n"


# Transforma um projeto bronze em um registro no schema canonico da silver.
def transform_project(
    payload: dict[str, Any],
    file_path: Path,
    field_order: list[FieldSpec],
    transformers: dict[str, Callable[[dict[str, Any], Path], Any]],
) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for field_spec in field_order:
        transformer = transformers.get(field_spec.name)
        if transformer:
            record[field_spec.name] = transformer(payload, file_path)
            continue
        if field_spec.name == "project_geometry":
            record[field_spec.name] = extract_project_geometry(payload, file_path)
            continue
        record[field_spec.name] = None
    return record


# Executa o fluxo completo do script de mapeamento usando a configuracao informada.
# Descompacta o bronze automaticamente se estiver zipado e recompacta ao final.
def run_mapping(config: dict[str, Any]) -> int:
    parser = build_mapping_parser(config["display_name"], config["mapping_output_path"])
    args = parser.parse_args()
    snapshot_date = validate_date(args.date)
    if args.limit is not None and args.limit <= 0:
        raise SystemExit("--limit deve ser maior que zero.")
    if not 0 < args.sample_fraction <= 1:
        raise SystemExit("--sample-fraction deve estar entre 0 e 1.")
    if args.sample_min_files <= 0:
        raise SystemExit("--sample-min-files deve ser maior que zero.")
    if not 0 <= args.sample_largest_ratio <= 1:
        raise SystemExit("--sample-largest-ratio deve estar entre 0 e 1.")

    bronze_dir = ROOT_DIR / "data" / "project_standards" / "01_bronze" / config["bronze_slug"] / snapshot_date
    list_path = bronze_dir / "list" / "projects.json"
    projects_dir = bronze_dir / "projects"

    # Descompacta o bronze automaticamente se estiver zipado
    unpacked = unpack_bronze_if_needed(bronze_dir)

    try:
        if not list_path.exists():
            raise SystemExit(f"Arquivo da lista nao encontrado: {list_path}")
        if not projects_dir.exists():
            raise SystemExit(f"Diretorio de projetos nao encontrado: {projects_dir}")

        sorter = config.get("sort_key", lambda path: path.stem)
        detail_files = resolve_detail_files(projects_dir, sorter, args.limit)
        if not detail_files:
            raise SystemExit(f"Nenhum arquivo de detalhe encontrado em {projects_dir}")

        sampled_files, effective_seed, sample_largest_count, sample_random_count = select_sample_files(
            detail_files=detail_files,
            snapshot_date=snapshot_date,
            sample_fraction=args.sample_fraction,
            sample_min_files=args.sample_min_files,
            sample_seed=args.sample_seed,
            sample_largest_ratio=args.sample_largest_ratio,
            sorter=sorter,
        )

        fields = parse_guide_fields(GUIDE_PATH)
        rows = build_mapping_rows(fields, sampled_files, config["mapping_candidates"]())
        markdown = build_mapping_markdown(
            config["display_name"],
            config["bronze_slug"],
            snapshot_date,
            sampled_files,
            rows,
            total_available_files=len(detail_files),
            sample_fraction=args.sample_fraction,
            sample_min_files=args.sample_min_files,
            sample_seed=effective_seed,
            sample_largest_ratio=args.sample_largest_ratio,
            sample_largest_count=sample_largest_count,
            sample_random_count=sample_random_count,
        )

        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(markdown, encoding="utf-8")

        print(f"Mapeamento inicial silver da {config['display_name']} gerado com sucesso")
        print(f"snapshot analisado: {snapshot_date}")
        print(f"arquivos disponiveis no snapshot: {len(detail_files)}")
        print(f"arquivos analisados na amostra: {len(sampled_files)}")
        print(f"maiores arquivos na amostra: {sample_largest_count}")
        print(f"arquivos aleatorios na amostra: {sample_random_count}")
        print(f"seed da amostra: {effective_seed}")
        print(f"arquivo de saida: {output_path}")
        return 0
    finally:
        if unpacked and bronze_dir.exists():
            repack_bronze(bronze_dir)


# Executa o fluxo completo do script de consolidacao usando a configuracao informada.
# Descompacta o bronze automaticamente se estiver zipado e recompacta ao final.
def run_dataset(config: dict[str, Any]) -> int:
    parser = build_dataset_parser(config["display_name"], config["bronze_slug"])
    args = parser.parse_args()
    snapshot_date = validate_date(args.date)
    if args.limit is not None and args.limit <= 0:
        raise SystemExit("--limit deve ser maior que zero.")

    bronze_dir = ROOT_DIR / "data" / "project_standards" / "01_bronze" / config["bronze_slug"] / snapshot_date
    list_path = bronze_dir / "list" / "projects.json"
    projects_dir = bronze_dir / "projects"
    output_path = Path(args.output) if args.output else Path(str(config["dataset_output_template"]).format(date=snapshot_date))
    failures_path = (
        Path(args.failures_output)
        if args.failures_output
        else Path(str(config["failure_output_template"]).format(date=snapshot_date))
    )

    # Descompacta o silver anterior se existir zipado (para sobrescrita)
    silver_dir = output_path.parent
    silver_zip = silver_dir.parent / f"{silver_dir.name}.zip"
    if not silver_dir.exists() and silver_zip.exists():
        unpack_archive(silver_zip, label="silver", step=1, total=1)

    # Descompacta o bronze automaticamente se estiver zipado
    unpacked = unpack_bronze_if_needed(bronze_dir)

    try:
        if not list_path.exists():
            raise SystemExit(f"Arquivo da lista nao encontrado: {list_path}")
        if not projects_dir.exists():
            raise SystemExit(f"Diretorio de projetos nao encontrado: {projects_dir}")
        if not REFERENCE_PATH.exists():
            raise SystemExit(f"Arquivo de referencia nao encontrado: {REFERENCE_PATH}")

        detail_files = resolve_detail_files(projects_dir, config.get("sort_key", lambda path: path.stem), args.limit)
        if not detail_files:
            raise SystemExit(f"Nenhum arquivo de detalhe encontrado em {projects_dir}")

        field_order = parse_guide_fields(GUIDE_PATH)
        standard_acronym = load_standard_acronym(REFERENCE_PATH, config["reference_name"])
        transformers = config["transformers"](standard_acronym)
        failures: list[dict[str, Any]] = []
        records: list[dict[str, Any]] = []
        quality_issue_counts: dict[str, int] = {}

        print(f"Iniciando consolidacao silver da {config['display_name']}")
        print(f"snapshot analisado: {snapshot_date}")
        print(f"arquivo da lista: {list_path}")
        print(f"diretorio de detalhes: {projects_dir}")
        print(f"total de arquivos de detalhe detectados: {len(detail_files)}")
        print(f"arquivo de saida: {output_path}")

        for index, file_path in enumerate(detail_files, start=1):
            try:
                payload = load_json(file_path)
                bronze_record = transform_project(payload, file_path, field_order, transformers)
                record = normalize_record(bronze_record)
                records.append(record)
                quality_issues = collect_quality_issues(record)
                for issue in quality_issues:
                    quality_issue_counts[issue] = quality_issue_counts.get(issue, 0) + 1
                if index == 1 or index % 250 == 0 or index == len(detail_files):
                    print(f"progresso: {index}/{len(detail_files)} projetos consolidados")
            except Exception as exc:  # noqa: BLE001
                failures.append(
                    {
                        "project_id": file_path.stem,
                        "file_path": derive_bronze_file_path(file_path),
                        "error": str(exc),
                    }
                )
                print(f"falha ao consolidar projeto {file_path.stem}: {exc}")

        if failures:
            write_json(failures_path, failures)
            print(f"log de falhas salvo em: {failures_path}")

        snapshot_payload = {
            "standard_name": config["bronze_slug"],
            "snapshot_date": records[0]["snapshot_date"] if records else None,
            "reference_month": records[0]["reference_month"] if records else None,
            "generated_at": format_datetime_iso(datetime.now(timezone.utc)),
            "total_projects": len(records),
            "failed_projects": len(failures),
            "quality_summary": {
                "issues_detected": sum(quality_issue_counts.values()),
                "issue_counts": dict(sorted(quality_issue_counts.items())),
            },
            "projects": records,
        }
        write_json(output_path, snapshot_payload)
        reference_sync_changes = sync_reference_dataset_observed()
        print(f"referencias observadas sincronizadas: {reference_sync_changes}")

        for hook in config.get("post_build_hooks", []):
            hook(
                records=records,
                snapshot_date=snapshot_date,
                output_path=output_path,
                snapshot_payload=snapshot_payload,
            )

        print(f"projetos consolidados com sucesso: {len(records)}")
        print(f"projetos com falha: {len(failures)}")
        print("consolidacao concluida")

        # Compacta o diretorio silver gerado em ZIP
        silver_dir = output_path.parent
        pack_directory(silver_dir, label="silver", step=1, total=1)

        return 0
    finally:
        # Recompacta o bronze se foi descompactado automaticamente
        if unpacked and bronze_dir.exists():
            repack_bronze(bronze_dir)
