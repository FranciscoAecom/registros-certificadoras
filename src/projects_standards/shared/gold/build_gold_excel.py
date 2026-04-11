# Objetivo do script:
# Gera a base gold em formato Excel (.xlsx) com abas normalizadas para entrega ao usuário.
# Filtra por reference_month e produz um arquivo autossuficiente com dados + referências.
# Processo:
# 1. Ler argumentos CLI (--reference-month para filtro).
# 2. Localizar e descompactar o arquivo allprojects_YYYYMM.json compactado na pasta do mês (03_gold/projects/YYYY-MM/allprojects_YYYYMM.json.zip).
# 3. Carregar o dataset gold JSON do mês informado.
# 4. Extrair 5 tabelas normalizadas: projects, project_history, methodologies, SDGs, sectors.
# 5. Fazer backup do Excel anterior.
# 6. Criar workbook Excel com abas de dados + 11 abas de referência.
# 7. Aplicar formatação (tabelas, autofiltros, alinhamento).
# 8. Recompactar novamente o JSON após gerar o Excel.


from __future__ import annotations

import argparse
import json
import re
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Reutiliza toda a logica de negocio do builder gold existente.
from src.projects_standards.shared.gold.build_gold_dataset import (
    RecordContext,
    build_gold_keys,
    build_gold_project,
    build_quality_report,
    deduplicate_gold_projects,
    load_reference_data,
    normalize_text,
    parse_date,
    parse_datetime,
)

# Utilitários de compactação/descompactação
from src.projects_standards.shared import archive_data

CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parents[3]
REFERENCE_WORKBOOK_PATH = ROOT_DIR / 'data' / 'project_standards' / '00_reference' / 'reference_dataset.xlsx'
SILVER_BASE_DIR = ROOT_DIR / 'data' / 'project_standards' / '02_silver'
DEFAULT_OUTPUT_DIR = ROOT_DIR / 'data' / 'project_standards' / '03_gold' / 'projects'
DEFAULT_BACKUP_DIR = DEFAULT_OUTPUT_DIR / 'backup'

# Abas de referencia do reference_dataset.xlsx que serao copiadas para o Excel gold.
REFERENCE_SHEETS = [
    'standards_catalog',
    'standards_status',
    'common_pipeline_status',
    'countries_standard',
    'countries_observed_mapping',
    'methodologies',
    'technical_areas',
    'sectoral_scopes',
    'sdg_goals',
    'sdg_targets',
    'sdg_observed_mapping',
]

# Colunas da aba projects (identidade fixa do projeto).
PROJECTS_COLUMNS = [
    'project_history_id',
    'standard_acronym',
    'project_internal_id',
    'project_public_id',
    'project_url',
]

# Colunas da aba project_history (estado mensal do projeto).
PROJECT_HISTORY_COLUMNS = [
    'record_id',
    'project_history_id',
    'standard_name',
    'standard_acronym',
    'bronze_file_path',
    'source_file_name',
    'snapshot_date',
    'reference_month',
    'gold_selected_from_snapshot',
    'project_name',
    'project_description',
    'standard_program',
    'project_market',
    'standard_reported_project_status',
    'standard_pipeline_status_id',
    'project_type',
    'project_category',
    'project_subcategories',
    'project_developer',
    'project_owner',
    'project_operator',
    'validator_name',
    'verifier_name',
    'country_observed',
    'country_standard',
    'country_alpha_3',
    'state_or_region',
    'city_or_locality',
    'location_latitude',
    'location_longitude',
    'project_geometry',
    'registration_date',
    'status_date',
    'crediting_start_date',
    'crediting_end_date',
    'first_issuance_date',
    'last_issuance_date',
    'credits_issued_total',
    'credits_retired_total',
    'credits_cancelled_total',
    'credits_buffer_total',
    'estimated_annual_emission_reductions',
    'estimated_total_emission_reductions',
    'area_hectares',
]

# Colunas da aba project_history_methodologies.
METHODOLOGY_COLUMNS = [
    'record_id',
    'methodology_sequence',
    'standard_acronym',
    'methodology_name',
    'technical_area_id',
    'sectoral_scope_id',
]

# Colunas da aba project_history_sdgs.
SDG_COLUMNS = [
    'record_id',
    'sdg_goal_id',
]

# Colunas da aba project_history_sectors.
SECTOR_COLUMNS = [
    'record_id',
    'standard_reported_sector',
]


# Constroi o lookup de country_standard -> alpha_3 a partir da aba countries_standard.
def build_country_alpha3_lookup(workbook_path: Path) -> dict[str, str]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb['countries_standard']
        headers = [cell.value for cell in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        lookup: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name_en = row[idx['name_en']]
            alpha_3 = row[idx['alpha_3']]
            if name_en and alpha_3:
                lookup[str(name_en).strip()] = str(alpha_3).strip()
        return lookup
    finally:
        wb.close()


# Descobre datasets silver filtrados pelo reference_month solicitado.
# Preserva o campo country observado da silver no registro gold como _country_observed.
def collect_gold_candidates_for_month(
    reference_month: str,
    references: dict[str, Any],
    quality: Counter,
) -> tuple[list[dict[str, Any]], int, int]:
    candidates: list[dict[str, Any]] = []
    dataset_count = 0
    project_count = 0

    if not SILVER_BASE_DIR.exists():
        print(f'silver base directory not found: {SILVER_BASE_DIR}')
        return candidates, dataset_count, project_count

    for dataset_path in sorted(SILVER_BASE_DIR.rglob('allprojects.json')):
        payload = json.loads(dataset_path.read_text(encoding='utf-8'))
        dataset_ref_month = payload.get('reference_month')
        # Filtra somente datasets cujo reference_month corresponde ao solicitado.
        if dataset_ref_month != reference_month:
            continue

        dataset_generated_at = parse_datetime(payload.get('generated_at'))
        dataset_snapshot_date = parse_date(payload.get('snapshot_date'))
        projects = payload.get('projects') or []
        dataset_count += 1
        project_count += len(projects)
        file_snapshot_token = dataset_path.parent.name
        context = RecordContext(
            dataset_path=dataset_path,
            dataset_generated_at=dataset_generated_at,
            dataset_snapshot_date=dataset_snapshot_date,
            file_snapshot_token=file_snapshot_token,
        )
        print(f'  loading: {dataset_path.relative_to(ROOT_DIR)} ({len(projects)} projects)')
        for project in projects:
            gold_record = build_gold_project(project, context, references, quality)
            # Preserva a forma observada do pais da silver para rastreabilidade no Excel.
            gold_record['_country_observed'] = project.get('country')
            candidates.append(gold_record)

    return candidates, dataset_count, project_count


# Extrai a tabela projects (identidade fixa) a partir dos registros gold deduplicated.
def extract_projects_table(gold_projects: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: dict[str, dict[str, Any]] = {}
    for project in gold_projects:
        history_id = project['project_history_id']
        if history_id in seen:
            continue
        seen[history_id] = {
            'project_history_id': history_id,
            'standard_acronym': project.get('standard_acronym'),
            'project_internal_id': project.get('project_internal_id'),
            'project_public_id': project.get('project_public_id'),
            'project_url': project.get('project_url'),
        }
    return sorted(seen.values(), key=lambda r: r['project_history_id'])


# Extrai a tabela project_history (estado mensal) com country_alpha_3 resolvido.
def extract_project_history_table(
    gold_projects: list[dict[str, Any]],
    country_alpha3_lookup: dict[str, str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for project in gold_projects:
        country_std = project.get('country_standard')
        alpha_3 = country_alpha3_lookup.get(country_std) if country_std else None
        row: dict[str, Any] = {}
        for col in PROJECT_HISTORY_COLUMNS:
            if col == 'country_observed':
                # Preserva a forma observada original para rastreabilidade.
                row[col] = project.get('_country_observed')
            elif col == 'country_alpha_3':
                row[col] = alpha_3
            else:
                row[col] = project.get(col)
        rows.append(row)
    return rows


# Extrai a tabela project_history_methodologies (N por record).
def extract_methodologies_table(gold_projects: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for project in gold_projects:
        record_id = project['record_id']
        standard_acronym = project.get('standard_acronym')
        items = project.get('project_methodology') or []
        for seq, item in enumerate(items, start=1):
            rows.append({
                'record_id': record_id,
                'methodology_sequence': seq,
                'standard_acronym': standard_acronym,
                'methodology_name': item.get('project_methodology'),
                'technical_area_id': item.get('technical_area_id'),
                'sectoral_scope_id': item.get('sectoral_scope_id'),
            })
    return rows


# Extrai a tabela project_history_sdgs (N por record).
def extract_sdgs_table(gold_projects: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for project in gold_projects:
        record_id = project['record_id']
        for goal_id in project.get('sdg_goal_ids') or []:
            rows.append({
                'record_id': record_id,
                'sdg_goal_id': goal_id,
            })
    return rows


# Extrai a tabela project_history_sectors (N por record).
def extract_sectors_table(gold_projects: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for project in gold_projects:
        record_id = project['record_id']
        sectors = project.get('standard_reported_sector') or []
        if isinstance(sectors, str):
            sectors = [sectors]
        for sector in sectors:
            if sector in (None, ''):
                continue
            rows.append({
                'record_id': record_id,
                'standard_reported_sector': str(sector).strip(),
            })
    return rows


# Converte valores que nao sao nativos do Excel (listas, dicts) para texto.
# Remove caracteres de controle ilegais para o Excel (U+0000-U+001F exceto tab/newline/cr).
def to_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, list):
        items = [str(v) for v in value if v is not None]
        return _sanitize_excel_string(' | '.join(items)) if items else None
    if isinstance(value, dict):
        return _sanitize_excel_string(json.dumps(value, ensure_ascii=False))
    if isinstance(value, str):
        return _sanitize_excel_string(value)
    return value


_ILLEGAL_CHARS_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f\ufffe\uffff]'
)


# Remove caracteres que o openpyxl rejeita como illegais em celulas Excel.
def _sanitize_excel_string(text: str) -> str:
    return _ILLEGAL_CHARS_RE.sub('', text)


# Escreve uma lista de dicts como aba Excel com tabela formatada.
def write_sheet_from_rows(
    wb: Workbook,
    sheet_name: str,
    columns: list[str],
    rows: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(title=sheet_name)
    # Cabecalho.
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    # Dados.
    for row_idx, record in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=to_excel_value(record.get(col_name)))
    # Formata como tabela Excel quando houver dados.
    if rows:
        last_col = get_column_letter(len(columns))
        last_row = len(rows) + 1
        table_name = sheet_name.replace('-', '_')
        table = Table(
            displayName=table_name,
            ref=f'A1:{last_col}{last_row}',
        )
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium9',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


# Copia uma aba do reference_dataset.xlsx para o workbook gold.
def copy_reference_sheet(source_wb_path: Path, sheet_name: str, target_wb: Workbook) -> None:
    source_wb = load_workbook(source_wb_path, read_only=True, data_only=True)
    try:
        if sheet_name not in source_wb.sheetnames:
            print(f'  warning: reference sheet "{sheet_name}" not found, skipping')
            return
        source_ws = source_wb[sheet_name]
        target_ws = target_wb.create_sheet(title=f'ref_{sheet_name}')
        for row_idx, row in enumerate(source_ws.iter_rows(values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                target_ws.cell(row=row_idx, column=col_idx, value=value)
        # Formata como tabela se houver dados.
        max_row = target_ws.max_row
        max_col = target_ws.max_column
        if max_row > 1 and max_col > 0:
            last_col = get_column_letter(max_col)
            table_name = f'ref_{sheet_name}'.replace('-', '_')
            table = Table(
                displayName=table_name,
                ref=f'A1:{last_col}{max_row}',
            )
            table.tableStyleInfo = TableStyleInfo(
                name='TableStyleMedium2',
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            target_ws.add_table(table)
    finally:
        source_wb.close()


# Move artefatos Excel anteriores do mesmo reference_month para backup.
def backup_existing_excel(output_path: Path, backup_dir: Path) -> Path | None:
    if not output_path.exists():
        return None
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%dT%H%M%S')
    backup_snapshot_dir = backup_dir / timestamp
    backup_snapshot_dir.mkdir(parents=True, exist_ok=True)
    shutil.move(str(output_path), str(backup_snapshot_dir / output_path.name))
    return backup_snapshot_dir


# Monta o parser de argumentos.
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description='Gera a base gold em Excel (.xlsx) para um reference_month especifico.',
    )
    parser.add_argument(
        '--reference-month',
        required=True,
        help='Mes de referencia no formato YYYY-MM (ex: 2026-04). Sera convertido para YYYY-MM-01.',
    )
    parser.add_argument(
        '--output-dir',
        default=str(DEFAULT_OUTPUT_DIR),
        help='Diretorio de saida do arquivo Excel.',
    )
    parser.add_argument(
        '--reference-workbook',
        default=str(REFERENCE_WORKBOOK_PATH),
        help='Workbook consolidado de referencia.',
    )
    parser.add_argument(
        '--backup-dir',
        default=str(DEFAULT_BACKUP_DIR),
        help='Diretorio de backup das versoes substituidas.',
    )
    return parser


# Executa o fluxo completo de construcao da base gold em Excel.
def main() -> int:
    parser = build_parser()
    args = parser.parse_args()


    # Normaliza reference_month para formato YYYY-MM-DD (primeiro dia do mês) e define paths do JSON compactado
    raw_month = args.reference_month.strip()
    if len(raw_month) == 7:
        reference_month = f'{raw_month}-01'
        month_tag = raw_month.replace('-', '')
        month_folder = raw_month
    elif len(raw_month) == 10:
        reference_month = raw_month
        month_tag = raw_month[:7].replace('-', '')
        month_folder = raw_month[:7]
    else:
        print(f'error: formato invalido para --reference-month: {raw_month}. Use YYYY-MM ou YYYY-MM-DD.')
        return 1

    output_dir = Path(args.output_dir)
    output_path = output_dir / f'allprojects_{month_tag}.xlsx'
    reference_workbook = Path(args.reference_workbook)
    backup_dir = Path(args.backup_dir)

    # Caminho do JSON gold do mês
    gold_json_dir = output_dir / month_folder
    gold_json_name = f'allprojects_{month_tag}.json'
    gold_json_zip = gold_json_dir.parent / f'{gold_json_dir.name}.zip'
    gold_json_path = gold_json_dir / gold_json_name

    # 1. Descompacta o JSON do mês, se necessário
    if not gold_json_path.exists() and gold_json_zip.exists():
        print(f'Descompactando gold JSON: {gold_json_zip}')
        archive_data.unpack_archive(gold_json_zip, f'gold/{month_folder}', 1, 1)
    elif not gold_json_path.exists():
        print(f'Arquivo gold JSON não encontrado: {gold_json_path}')
        return 1

    print(f'=== Gold Excel Builder ===')
    print(f'reference_month: {reference_month}')
    print(f'output: {output_path}')
    print()

    # 1. Carrega referencias.
    print('loading reference workbook...')
    references = load_reference_data(reference_workbook)
    country_alpha3_lookup = build_country_alpha3_lookup(reference_workbook)
    print(f'  country alpha_3 entries: {len(country_alpha3_lookup)}')
    print()


    # 2. Carrega o JSON gold do mês
    print(f'Carregando gold JSON: {gold_json_path}')
    with open(gold_json_path, 'r', encoding='utf-8') as f:
        payload = json.load(f)
    candidates = payload.get('projects', [])
    dataset_count = payload.get('source_datasets_scanned', 0)
    source_project_count = payload.get('source_projects_scanned', 0)
    print(f'  candidate projects: {len(candidates)}')
    if not candidates:
        print('error: nenhum projeto encontrado no gold JSON para este reference_month.')
        return 1
    print()

    # 3. Deduplica por projeto dentro do mes.
    print('deduplicating by project within reference_month...')
    quality = Counter()
    gold_projects = deduplicate_gold_projects(candidates, quality)
    print(f'  final gold projects: {len(gold_projects)}')
    print()

    # 4. Extrai tabelas normalizadas.
    print('extracting normalized tables...')
    projects_table = extract_projects_table(gold_projects)
    history_table = extract_project_history_table(gold_projects, country_alpha3_lookup)
    methodologies_table = extract_methodologies_table(gold_projects)
    sdgs_table = extract_sdgs_table(gold_projects)
    sectors_table = extract_sectors_table(gold_projects)
    print(f'  projects:        {len(projects_table)} rows')
    print(f'  project_history: {len(history_table)} rows')
    print(f'  methodologies:   {len(methodologies_table)} rows')
    print(f'  sdgs:            {len(sdgs_table)} rows')
    print(f'  sectors:         {len(sectors_table)} rows')
    print()

    # 5. Backup de arquivo anterior, se existir.
    backup_path = backup_existing_excel(output_path, backup_dir)
    if backup_path:
        print(f'previous file backed up to: {backup_path}')

    # 6. Gera o workbook Excel.
    print('writing Excel workbook...')
    wb = Workbook()
    # Remove a aba default criada pelo openpyxl.
    wb.remove(wb.active)

    write_sheet_from_rows(wb, 'projects', PROJECTS_COLUMNS, projects_table)
    write_sheet_from_rows(wb, 'project_history', PROJECT_HISTORY_COLUMNS, history_table)
    write_sheet_from_rows(wb, 'project_history_method', METHODOLOGY_COLUMNS, methodologies_table)
    write_sheet_from_rows(wb, 'project_history_sdgs', SDG_COLUMNS, sdgs_table)
    write_sheet_from_rows(wb, 'project_history_sectors', SECTOR_COLUMNS, sectors_table)

    # 7. Copia abas de referencia do reference_dataset.xlsx.
    print('copying reference sheets...')
    for sheet_name in REFERENCE_SHEETS:
        copy_reference_sheet(reference_workbook, sheet_name, wb)

    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    total_sheets = len(wb.sheetnames)
    wb.close()
    print()
    print(f'gold Excel written: {output_path}')
    print(f'total sheets: {total_sheets}')


    # 8. Recompacta o JSON do mês após gerar o Excel
    if gold_json_path.exists():
        print(f'Compactando novamente o gold JSON: {gold_json_dir}')
        archive_data.pack_directory(gold_json_dir, f'gold/{month_folder}', 1, 1)

    print('done.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
