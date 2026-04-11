# Objetivo do script:
# Consolidar todos os datasets silver em uma base gold única com um registro por projeto por reference_month.
# Processo:
# 1. Ler argumentos CLI (--reference-month opcional).
# 2. Definir caminhos do workbook de referência e diretório gold (03_gold/projects).
# 3. Identificar e descompactar automaticamente todos os snapshots silver necessários (02_silver/<certificadora>/<YYYYMMDD>.zip) que ainda não estejam descompactados.
# 4. Escanear todos os datasets silver disponíveis em 02_silver/.
# 5. Deduplicar registros por projeto por reference_month.
# 6. Gerar chaves únicas: project_history_id e record_id.
# 7. Aplicar padronizações obrigatórias (SDGs, metodologias, país, status).
# 8. Fazer backup do gold anterior antes de sobrescrever.
# 9. Salvar allprojects.json, schema.json e quality_report.json.
# 10. Se --reference-month for usado, salvar o allprojects em uma subpasta 03_gold/projects/YYYY-MM/allprojects_YYYYMM.json e compactar ao final.
# 11. Recompactar novamente os snapshots silver que foram descompactados ao final da execução.

from __future__ import annotations

import argparse
import json
import math
import shutil
from importlib import import_module
import sys
import traceback
import types
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Importa utilitários de compactação/descompactação centralizados
from src.projects_standards.shared import archive_data


CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parents[3]
REFERENCE_WORKBOOK_PATH = ROOT_DIR / 'data' / 'project_standards' / '00_reference' / 'reference_dataset.xlsx'
DEFAULT_OUTPUT_DIR = ROOT_DIR / 'data' / 'project_standards' / '03_gold' / 'projects'
DEFAULT_OUTPUT_PATH = DEFAULT_OUTPUT_DIR / 'allprojects.json'
DEFAULT_QUALITY_PATH = DEFAULT_OUTPUT_DIR / 'quality_report.json'
DEFAULT_SCHEMA_PATH = DEFAULT_OUTPUT_DIR / 'schema.json'
DEFAULT_BACKUP_DIR = DEFAULT_OUTPUT_DIR / 'backup'
PLACEHOLDER_METHODOLOGIES = {
    'not provided',
    'other',
    'vcs v1 project specific',
}


# Agrupa metadados necessarios para comparar snapshots concorrentes do mesmo projeto no mes.
@dataclass(frozen=True)
class RecordContext:
    dataset_path: Path
    dataset_generated_at: datetime | None
    dataset_snapshot_date: datetime | None
    file_snapshot_token: str


# Normaliza texto para comparacoes de joins com as referencias.
def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text == '':
        return None
    return text.casefold()


# Converte texto ISO em datetime quando possivel.
def parse_datetime(value: Any) -> datetime | None:
    if value in (None, ''):
        return None
    text = str(value).strip()
    if text == '':
        return None
    try:
        if text.endswith('Z'):
            text = text[:-1] + '+00:00'
        return datetime.fromisoformat(text)
    except ValueError:
        return None


# Converte texto ISO de data em datetime simples para comparacao.
def parse_date(value: Any) -> datetime | None:
    if value in (None, ''):
        return None
    text = str(value).strip()
    if text == '':
        return None
    try:
        return datetime.strptime(text, '%Y-%m-%d')
    except ValueError:
        return None


# Remove duplicados preservando a ordem original dos itens validos.
def unique_preserve_order(values: list[Any]) -> list[Any]:
    seen: set[Any] = set()
    result: list[Any] = []
    for value in values:
        if value in (None, ''):
            continue
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def _is_finite_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def _coerce_position(value: Any) -> list[float] | None:
    if not isinstance(value, list) or len(value) < 2:
        return None
    lon = value[0]
    lat = value[1]
    if not _is_finite_number(lon) or not _is_finite_number(lat):
        return None
    return [float(lon), float(lat)]


def _is_ring(value: Any) -> bool:
    return isinstance(value, list) and len(value) >= 4 and all(_coerce_position(point) is not None for point in value)


def _close_ring(ring: list[list[float]]) -> list[list[float]]:
    if ring and ring[0] != ring[-1]:
        return [*ring, ring[0]]
    return ring


def _normalize_ring(value: Any) -> list[list[float]] | None:
    if not isinstance(value, list):
        return None
    normalized: list[list[float]] = []
    for point in value:
        coerced = _coerce_position(point)
        if coerced is None:
            return None
        normalized.append(coerced)
    normalized = _close_ring(normalized)
    if len(normalized) < 4:
        return None
    return normalized


def _normalize_polygon_coordinates(value: Any) -> list[list[list[float]]] | None:
    coords = value
    while isinstance(coords, list) and len(coords) == 1 and not _is_ring(coords):
        nested = coords[0]
        if not isinstance(nested, list):
            break
        coords = nested
    if _is_ring(coords):
        normalized_ring = _normalize_ring(coords)
        if normalized_ring is None:
            return None
        return [normalized_ring]
    if not isinstance(coords, list) or not coords:
        return None
    rings: list[list[list[float]]] = []
    for ring in coords:
        normalized_ring = _normalize_ring(ring)
        if normalized_ring is None:
            return None
        rings.append(normalized_ring)
    return rings


def normalize_project_geometry(value: Any) -> dict[str, Any] | None:
    if not isinstance(value, dict):
        return None
    geometry_type = value.get('type')
    coordinates = value.get('coordinates')

    if geometry_type == 'Point':
        point = _coerce_position(coordinates)
        if point is None:
            return None
        return {'type': 'Point', 'coordinates': point}

    if geometry_type == 'Polygon':
        polygon = _normalize_polygon_coordinates(coordinates)
        if polygon is None:
            return None
        return {'type': 'Polygon', 'coordinates': polygon}

    if geometry_type == 'MultiPolygon':
        if not isinstance(coordinates, list) or not coordinates:
            return None
        multipolygon: list[list[list[list[float]]]] = []
        for polygon in coordinates:
            normalized_polygon = _normalize_polygon_coordinates(polygon)
            if normalized_polygon is None:
                return None
            multipolygon.append(normalized_polygon)
        return {'type': 'MultiPolygon', 'coordinates': multipolygon}

    return None


def build_geojson_feature_collection(projects: list[dict[str, Any]]) -> tuple[dict[str, Any], int]:
    features: list[dict[str, Any]] = []
    skipped = 0
    for project in projects:
        normalized_geometry = normalize_project_geometry(project.get('project_geometry'))
        if normalized_geometry is None:
            skipped += 1
            continue
        features.append(
            {
                'type': 'Feature',
                'properties': {
                    'record_id': project.get('record_id'),
                    'project_history_id': project.get('project_history_id'),
                    'standard_acronym': project.get('standard_acronym'),
                    'standard_name': project.get('standard_name'),
                    'project_public_id': project.get('project_public_id'),
                    'project_name': project.get('project_name'),
                    'snapshot_date': project.get('snapshot_date'),
                    'reference_month': project.get('reference_month'),
                },
                'geometry': normalized_geometry,
            }
        )
    return {'type': 'FeatureCollection', 'features': features}, skipped


# Carrega as referencias consolidadas necessarias para a construcao da gold.
def load_reference_data(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        methodology_lookup: dict[tuple[str, str], str] = {}
        ws = workbook['methodologies']
        headers = [cell.value for cell in ws[1]]
        idx = {header: i for i, header in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            standard = normalize_text(row[idx['standard_acronym']])
            methodology = normalize_text(row[idx['project_methodology']])
            technical_area_id = row[idx['technical_area_id']]
            if standard and methodology and technical_area_id:
                methodology_lookup[(standard, methodology)] = str(technical_area_id).strip()

        technical_area_lookup: dict[str, str] = {}
        ws = workbook['technical_areas']
        headers = [cell.value for cell in ws[1]]
        idx = {header: i for i, header in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            technical_area_id = row[idx['technical_area_id']]
            sectoral_scope_id = row[idx['sectoral_scope_id']]
            if technical_area_id and sectoral_scope_id:
                technical_area_lookup[str(technical_area_id).strip()] = str(sectoral_scope_id).strip()

        country_lookup: dict[str, str] = {}
        ws = workbook['countries_observed_mapping']
        headers = [cell.value for cell in ws[1]]
        idx = {header: i for i, header in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_country = normalize_text(row[idx['country_raw']])
            standard_country = row[idx['country_standard']]
            if raw_country and standard_country:
                country_lookup[raw_country] = str(standard_country).strip()

        sdg_lookup: dict[str, int] = {}
        ws = workbook['sdg_observed_mapping']
        headers = [cell.value for cell in ws[1]]
        idx = {header: i for i, header in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_value = normalize_text(row[idx['sdg_raw']])
            goal_id = row[idx['sdg_goal_id']]
            if raw_value and goal_id is not None:
                sdg_lookup[raw_value] = int(goal_id)

        status_lookup: dict[tuple[str, str, str], str] = {}
        ws = workbook['standards_status']
        headers = [cell.value for cell in ws[1]]
        idx = {header: i for i, header in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            standard = normalize_text(row[idx['standard_acronym']])
            market = normalize_text(row[idx['market']])
            status = normalize_text(row[idx['status_standard']])
            pipeline_id = row[idx['common_pipeline_status_id']]
            if standard and market and status and pipeline_id:
                status_lookup[(standard, market, status)] = str(pipeline_id).strip()

        return {
            'methodology_lookup': methodology_lookup,
            'technical_area_lookup': technical_area_lookup,
            'country_lookup': country_lookup,
            'sdg_lookup': sdg_lookup,
            'status_lookup': status_lookup,
        }
    finally:
        workbook.close()


# Resolve a atividade tecnica do projeto a partir da lista de metodologias da silver.
def build_project_methodology_items(
    project: dict[str, Any],
    methodology_lookup: dict[tuple[str, str], str],
    technical_area_lookup: dict[str, str],
    quality: Counter,
) -> list[dict[str, Any]]:
    standard = normalize_text(project.get('standard_acronym'))
    methodologies = project.get('project_methodology') or []
    if not isinstance(methodologies, list):
        methodologies = [methodologies]
    normalized_methodologies = [normalize_text(methodology) for methodology in methodologies if normalize_text(methodology)]
    if not normalized_methodologies:
        quality['technical_area_missing_methodology'] += 1
        return []
    if all(methodology in PLACEHOLDER_METHODOLOGIES for methodology in normalized_methodologies):
        quality['technical_area_placeholder_methodology'] += 1
        return [
            {
                'project_methodology': str(methodology).strip(),
                'technical_area_id': None,
                'sectoral_scope_id': None,
            }
            for methodology in methodologies
            if normalize_text(methodology)
        ]

    items: list[dict[str, Any]] = []
    resolved_technical_areas: list[str] = []
    for methodology in methodologies:
        key = (standard, normalize_text(methodology))
        if not key[0] or not key[1]:
            continue
        technical_area_id = methodology_lookup.get(key)
        sectoral_scope_id = technical_area_lookup.get(technical_area_id) if technical_area_id else None
        if technical_area_id:
            resolved_technical_areas.append(technical_area_id)
        items.append(
            {
                'project_methodology': str(methodology).strip(),
                'technical_area_id': technical_area_id,
                'sectoral_scope_id': sectoral_scope_id,
            }
        )

    resolved_technical_areas = unique_preserve_order(resolved_technical_areas)
    if not resolved_technical_areas and items:
        quality['technical_area_unmapped'] += 1
    if len(resolved_technical_areas) > 1:
        quality['technical_area_multiple_candidates'] += 1
    return items


# Resolve o escopo setorial derivado a partir da atividade tecnica padronizada.
def resolve_sectoral_scope_id(technical_area_id: str | None, technical_area_lookup: dict[str, str], quality: Counter) -> str | None:
    if technical_area_id is None:
        return None
    sectoral_scope_id = technical_area_lookup.get(technical_area_id)
    if sectoral_scope_id is None:
        quality['sectoral_scope_unmapped'] += 1
    return sectoral_scope_id


# Resolve o pais padronizado conforme o mapeamento de referencia.
def resolve_country_standard(project: dict[str, Any], country_lookup: dict[str, str], quality: Counter) -> str | None:
    raw_country = project.get('country')
    normalized = normalize_text(raw_country)
    if normalized is None:
        return None
    country_standard = country_lookup.get(normalized)
    if country_standard is None:
        country_standard = str(raw_country).strip()
        quality['country_fallback_to_observed'] += 1
    return country_standard


# Resolve a lista padronizada de SDGs em goal_id a partir dos valores observados na silver.
def resolve_sdg_goal_ids(project: dict[str, Any], sdg_lookup: dict[str, int], quality: Counter) -> list[int]:
    observed_values = project.get('sdg_targets') or []
    if not isinstance(observed_values, list):
        observed_values = [observed_values]
    resolved: list[int] = []
    for value in observed_values:
        normalized = normalize_text(value)
        if normalized is None:
            continue
        goal_id = sdg_lookup.get(normalized)
        if goal_id is None:
            quality['sdg_unmapped_values'] += 1
            continue
        resolved.append(goal_id)
    return unique_preserve_order(resolved)


# Deriva o mercado do projeto a partir dos status observados na silver, com regra generica conservadora.
def derive_project_market(project: dict[str, Any], quality: Counter) -> str | None:
    voluntary_status = normalize_text(project.get('project_voluntary_status'))
    regulatory_status = normalize_text(project.get('project_regulatory_status'))
    if voluntary_status and not regulatory_status:
        return 'voluntary'
    if regulatory_status and not voluntary_status:
        return 'regulatory'
    if voluntary_status and regulatory_status:
        quality['project_market_both_statuses_present'] += 1
        return 'voluntary'
    quality['project_market_missing'] += 1
    return None


# Deriva o status efetivo do projeto conforme o mercado selecionado.
def derive_project_status(project: dict[str, Any], project_market: str | None) -> str | None:
    if project_market == 'regulatory':
        value = project.get('project_regulatory_status')
    else:
        value = project.get('project_voluntary_status')
    if value in (None, ''):
        return None
    return str(value).strip()


# Resolve o status padrao de pipeline usando o status efetivo e a referencia consolidada.
def resolve_pipeline_status_id(project: dict[str, Any], project_market: str | None, project_status: str | None, status_lookup: dict[tuple[str, str, str], str], quality: Counter) -> str | None:
    key = (
        normalize_text(project.get('standard_acronym')),
        normalize_text(project_market),
        normalize_text(project_status),
    )
    if None in key:
        return None
    pipeline_status_id = status_lookup.get(key)
    if pipeline_status_id is None:
        quality['pipeline_status_unmapped'] += 1
    return pipeline_status_id


# Constrói as chaves oficiais da camada gold.
def build_gold_keys(project: dict[str, Any]) -> tuple[str, str]:
    standard_acronym = str(project.get('standard_acronym') or '').strip()
    project_internal_id = str(project.get('project_internal_id') or '').strip()
    reference_month = str(project.get('reference_month') or '').strip()
    project_history_id = f'{standard_acronym}_{project_internal_id}'
    record_id = f'{project_history_id}_{reference_month}'
    return project_history_id, record_id


# Compara dois candidatos do mesmo projeto no mes e escolhe o mais atualizado.
def choose_preferred_record(current: dict[str, Any], challenger: dict[str, Any], quality: Counter) -> dict[str, Any]:
    current_key = current['_sort_key']
    challenger_key = challenger['_sort_key']
    if challenger_key > current_key:
        quality['deduplicated_records_replaced'] += 1
        return challenger
    quality['deduplicated_records_discarded'] += 1
    return current


# Monta o registro gold final a partir do projeto silver e das referencias carregadas.
def build_gold_project(project: dict[str, Any], context: RecordContext, references: dict[str, Any], quality: Counter) -> dict[str, Any]:
    project_history_id, record_id = build_gold_keys(project)
    project_methodology_items = build_project_methodology_items(
        project,
        references['methodology_lookup'],
        references['technical_area_lookup'],
        quality,
    )
    country_standard = resolve_country_standard(project, references['country_lookup'], quality)
    sdg_goal_ids = resolve_sdg_goal_ids(project, references['sdg_lookup'], quality)
    project_market = derive_project_market(project, quality)
    standard_reported_project_status = derive_project_status(project, project_market)
    pipeline_status_id = resolve_pipeline_status_id(
        project,
        project_market,
        standard_reported_project_status,
        references['status_lookup'],
        quality,
    )

    return {
        'record_id': record_id,
        'project_history_id': project_history_id,
        'standard_name': project.get('standard_name'),
        'standard_acronym': project.get('standard_acronym'),
        'project_public_id': project.get('project_public_id'),
        'project_internal_id': project.get('project_internal_id'),
        'project_url': project.get('project_url'),
        'bronze_file_path': project.get('bronze_file_path'),
        'source_file_name': project.get('source_file_name'),
        'snapshot_date': project.get('snapshot_date'),
        'reference_month': project.get('reference_month'),
        'gold_selected_from_snapshot': context.file_snapshot_token,
        'project_name': project.get('project_name'),
        'project_description': project.get('project_description'),
        'standard_program': project.get('standard_program'),
        'project_market': project_market,
        'standard_reported_project_status': standard_reported_project_status,
        'standard_pipeline_status_id': pipeline_status_id,
        'project_type': project.get('project_type'),
        'project_category': project.get('project_category'),
        'project_subcategories': project.get('project_subcategories'),
        'project_methodology': project_methodology_items,
        'standard_reported_sector': project.get('sector') or [],
        'sdg_goal_ids': sdg_goal_ids,
        'project_developer': project.get('project_developer'),
        'project_owner': project.get('project_owner'),
        'project_operator': project.get('project_operator'),
        'validator_name': project.get('validator_name'),
        'verifier_name': project.get('verifier_name'),
        'country_standard': country_standard,
        'state_or_region': project.get('state_or_region'),
        'city_or_locality': project.get('city_or_locality'),
        'location_latitude': project.get('location_latitude'),
        'location_longitude': project.get('location_longitude'),
        'project_geometry': project.get('project_geometry'),
        'registration_date': project.get('registration_date'),
        'status_date': project.get('status_date'),
        'crediting_start_date': project.get('crediting_start_date'),
        'crediting_end_date': project.get('crediting_end_date'),
        'first_issuance_date': project.get('first_issuance_date'),
        'last_issuance_date': project.get('last_issuance_date'),
        'credits_issued_total': project.get('credits_issued_total'),
        'credits_retired_total': project.get('credits_retired_total'),
        'credits_cancelled_total': project.get('credits_cancelled_total'),
        'credits_buffer_total': project.get('credits_buffer_total'),
        'estimated_annual_emission_reductions': project.get('estimated_annual_emission_reductions'),
        'estimated_total_emission_reductions': project.get('estimated_total_emission_reductions'),
        'area_hectares': project.get('area_hectares'),
        '_sort_key': (
            context.dataset_generated_at or datetime.min,
            parse_date(project.get('snapshot_date')) or context.dataset_snapshot_date or datetime.min,
            context.file_snapshot_token,
        ),
    }


# Descobre todos os datasets silver disponiveis e retorna seus projetos enriquecidos com contexto.
def collect_gold_candidates(references: dict[str, Any], quality: Counter) -> tuple[list[dict[str, Any]], int, int]:
    candidates: list[dict[str, Any]] = []
    dataset_count = 0
    project_count = 0
    for dataset_path in sorted((ROOT_DIR / 'data' / 'project_standards' / '02_silver').rglob('allprojects.json')):
        payload = json.loads(dataset_path.read_text(encoding='utf-8'))
        dataset_generated_at = parse_datetime(payload.get('generated_at'))
        dataset_snapshot_date = parse_date(payload.get('snapshot_date'))
        projects = payload.get('projects') or []
        dataset_count += 1
        project_count += len(projects)
        file_snapshot_token = dataset_path.parent.name
        context = RecordContext(
            dataset_path=dataset_path,
            dataset_generated_at=dataset_generated_at,
            dataset_snapshot_date=dataset_snapshot_date,
            file_snapshot_token=file_snapshot_token,
        )
        print(f'loading silver dataset: {dataset_path} ({len(projects)} projects)')
        for project in projects:
            candidates.append(build_gold_project(project, context, references, quality))
    return candidates, dataset_count, project_count


# Deduplica a gold por projeto dentro do mesmo reference_month.
def deduplicate_gold_projects(candidates: list[dict[str, Any]], quality: Counter) -> list[dict[str, Any]]:
    selected: dict[tuple[str, str, str], dict[str, Any]] = {}
    for candidate in candidates:
        dedupe_key = (
            str(candidate.get('standard_acronym') or ''),
            str(candidate.get('project_internal_id') or ''),
            str(candidate.get('reference_month') or ''),
        )
        if dedupe_key not in selected:
            selected[dedupe_key] = candidate
            continue
        selected[dedupe_key] = choose_preferred_record(selected[dedupe_key], candidate, quality)
    final_projects = []
    for record in selected.values():
        record.pop('_sort_key', None)
        final_projects.append(record)
    final_projects.sort(key=lambda item: (item['standard_acronym'], item['project_internal_id'], item['reference_month']))
    return final_projects


# Move os artefatos anteriores da gold para um diretorio de backup identificado por timestamp.
def backup_existing_gold_artifacts(output_path: Path, quality_path: Path, geojson_path: Path | None, backup_dir: Path) -> Path | None:
    candidate_paths: list[Path] = [output_path, quality_path]
    if geojson_path is not None:
        candidate_paths.append(geojson_path)
    files_to_backup = [path for path in candidate_paths if path.exists()]
    if not files_to_backup:
        return None

    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%dT%H%M%S')
    backup_snapshot_dir = backup_dir / timestamp
    backup_snapshot_dir.mkdir(parents=True, exist_ok=True)

    for file_path in files_to_backup:
        shutil.move(str(file_path), str(backup_snapshot_dir / file_path.name))

    return backup_snapshot_dir


# Gera o relatorio de qualidade e estatisticas da execucao gold.
def build_quality_report(projects: list[dict[str, Any]], quality: Counter, dataset_count: int, source_project_count: int, backup_path: Path | None) -> dict[str, Any]:
    final_quality = Counter(quality)
    final_quality['project_market_missing'] = sum(1 for project in projects if project.get('project_market') is None)
    final_quality['standard_pipeline_status_missing'] = sum(
        1 for project in projects if project.get('standard_pipeline_status_id') is None
    )
    final_quality['project_market_both_statuses_present'] = sum(
        1
        for project in projects
        if project.get('project_voluntary_status') not in (None, '')
        and project.get('project_regulatory_status') not in (None, '')
    )
    final_quality['technical_area_missing_methodology'] = sum(
        1
        for project in projects
        if not (project.get('project_methodology') or [])
    )
    final_quality['technical_area_placeholder_methodology'] = sum(
        1
        for project in projects
        if (project.get('project_methodology') or [])
        and all(item.get('technical_area_id') is None for item in project.get('project_methodology') or [])
        and all(
            normalize_text(item.get('project_methodology')) in PLACEHOLDER_METHODOLOGIES
            for item in project.get('project_methodology') or []
            if normalize_text(item.get('project_methodology'))
        )
    )
    final_quality['technical_area_unmapped'] = sum(
        1
        for project in projects
        if (project.get('project_methodology') or [])
        and all(item.get('technical_area_id') is None for item in project.get('project_methodology') or [])
        and not all(
            normalize_text(item.get('project_methodology')) in PLACEHOLDER_METHODOLOGIES
            for item in project.get('project_methodology') or []
            if normalize_text(item.get('project_methodology'))
        )
    )
    return {
        'generated_at': datetime.now().isoformat(timespec='seconds'),
        'source_datasets_scanned': dataset_count,
        'source_projects_scanned': source_project_count,
        'gold_projects_generated': len(projects),
        'backup_created': str(backup_path) if backup_path else None,
        'quality_summary': dict(final_quality),
    }


# Salva o dataset gold principal e o quality report correspondente.
def write_gold_outputs(
    projects: list[dict[str, Any]],
    quality_report: dict[str, Any],
    output_path: Path,
    quality_path: Path,
    geojson_path: Path | None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        'generated_at': quality_report['generated_at'],
        'source_datasets_scanned': quality_report['source_datasets_scanned'],
        'source_projects_scanned': quality_report['source_projects_scanned'],
        'total_projects': len(projects),
        'quality_summary': quality_report['quality_summary'],
        'projects': projects,
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding='utf-8')
    quality_path.write_text(json.dumps(quality_report, ensure_ascii=True, indent=2), encoding='utf-8')
    if geojson_path is not None:
        geojson_path.parent.mkdir(parents=True, exist_ok=True)
        feature_collection, skipped = build_geojson_feature_collection(projects)
        geojson_path.write_text(json.dumps(feature_collection, ensure_ascii=False, indent=2), encoding='utf-8')
        print(f'gold geojson written: {geojson_path} ({len(feature_collection["features"])} features, {skipped} skipped)')


# Monta o parser de argumentos do builder da gold.
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description='Consolida todos os datasets silver em uma base gold unica.')
    parser.add_argument('--reference-workbook', default=str(REFERENCE_WORKBOOK_PATH), help='Workbook consolidado de referencia.')
    parser.add_argument('--output', default=str(DEFAULT_OUTPUT_PATH), help='Arquivo final allprojects.json da gold.')
    parser.add_argument('--quality-report', default=str(DEFAULT_QUALITY_PATH), help='Arquivo quality_report.json da gold.')
    parser.add_argument(
        '--geojson-output',
        default=None,
        help='Arquivo GeoJSON de saida com 1 feicao por projeto. Padrao: mesmo nome do --output com extensao .geojson.',
    )
    parser.add_argument('--backup-dir', default=str(DEFAULT_BACKUP_DIR), help='Diretorio de backup das versoes substituidas da gold.')
    parser.add_argument('--reference-month', type=str, default=None, help='Filtrar para consolidar apenas projetos do mês AAAA-MM (ex: 2026-04).')
    return parser


# Executa o fluxo completo de construcao da base gold.
def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    reference_workbook = Path(args.reference_workbook)
    output_path = Path(args.output)
    quality_path = Path(args.quality_report)
    geojson_path = Path(args.geojson_output) if args.geojson_output else output_path.with_suffix('.geojson')
    backup_dir = Path(args.backup_dir)
    reference_month = args.reference_month

    # Se reference_month, definir subpasta e nome customizado para o allprojects.json
    month_tag = None
    if reference_month:
        # Aceita tanto AAAA-MM quanto AAAA-MM-DD
        if len(reference_month) == 7:
            month_tag = reference_month.replace('-', '')
            month_folder = reference_month
        elif len(reference_month) == 10:
            month_tag = reference_month[:7].replace('-', '')
            month_folder = reference_month[:7]
        else:
            print(f'Formato inválido para --reference-month: {reference_month}')
            return 1
        gold_month_dir = DEFAULT_OUTPUT_DIR / month_folder
        gold_month_dir.mkdir(parents=True, exist_ok=True)
        output_path = gold_month_dir / f'allprojects_{month_tag}.json'
        if args.geojson_output:
            geojson_path = Path(args.geojson_output)
        else:
            geojson_path = gold_month_dir / f'allprojects_{month_tag}.geojson'

    print(f'loading reference workbook: {reference_workbook}')
    references = load_reference_data(reference_workbook)
    quality = Counter()

    # 1. Descompactar todos os snapshots silver necessários antes de processar
    print("\n[gold] Verificando e descompactando snapshots silver necessários...")
    silver_dir = Path(archive_data.LAYER_DIRS['silver'])
    # Descompacta todos os .zip de silver que ainda não possuem pasta descompactada
    silver_archives = archive_data.find_archives(silver_dir, standard=None, date=None)
    if silver_archives:
        archive_data.print_plan('unpack', silver_archives, 'silver')
        for i, zip_path in enumerate(silver_archives, start=1):
            label = f"{zip_path.parent.name}/{zip_path.stem}"
            archive_data.unpack_archive(zip_path, label, i, len(silver_archives))
    else:
        print("Nenhum snapshot silver compactado pendente para descompactar.")

    candidates, dataset_count, source_project_count = collect_gold_candidates(references, quality)
    print(f'gold candidates collected: {len(candidates)}')

    # Filtra candidatos pelo reference_month, se fornecido
    if reference_month:
        # Aceita tanto AAAA-MM quanto AAAA-MM-DD
        def month_match(project):
            ref = str(project.get('reference_month') or '')
            return ref.startswith(reference_month)
        candidates = [p for p in candidates if month_match(p)]
        print(f'gold candidates after filtering by reference_month={reference_month}: {len(candidates)}')

    projects = deduplicate_gold_projects(candidates, quality)
    print(f'gold projects after monthly deduplication: {len(projects)}')

    backup_path = backup_existing_gold_artifacts(output_path, quality_path, geojson_path, backup_dir)
    if backup_path:
        print(f'previous gold backup created: {backup_path}')

    quality_report = build_quality_report(projects, quality, dataset_count, source_project_count, backup_path)
    write_gold_outputs(projects, quality_report, output_path, quality_path, geojson_path)
    print(f'gold dataset written: {output_path}')
    print(f'gold quality report written: {quality_path}')

    # Se reference_month, compactar o allprojects gerado
    if reference_month:
        from src.projects_standards.shared.archive_data import pack_directory
        print(f'Compactando arquivo gold JSON do mês: {output_path.parent}')
        pack_directory(output_path.parent, f'gold/{month_folder}', 1, 1)

    # 2. Recompactar novamente os snapshots silver que foram descompactados
    print("\n[gold] Recompactando snapshots silver após processamento...")
    silver_snapshots = archive_data.find_snapshots(silver_dir, standard=None, date=None)
    if silver_snapshots:
        archive_data.print_plan('pack', silver_snapshots, 'silver')
        for i, snapshot_dir in enumerate(silver_snapshots, start=1):
            label = f"{snapshot_dir.parent.name}/{snapshot_dir.name}"
            archive_data.pack_directory(snapshot_dir, label, i, len(silver_snapshots))
    else:
        print("Nenhum snapshot silver pendente para compactar.")

    return 0


if __name__ == '__main__':
    raise SystemExit(main())
