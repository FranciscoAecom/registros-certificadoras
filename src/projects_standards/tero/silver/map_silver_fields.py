# Objetivo do script:
# Analisar os arquivos bronze da TERO e gerar um mapeamento inicial entre o bruto e o schema canonico da camada silver.

# Processo:
# 1. Ler argumentos CLI (--date, --output, --sample-fraction, --limit).
# 2. Carregar amostra hibrida de arquivos bronze do snapshot (maiores + aleatorios).
# 3. Inspecionar campos presentes em list_data e detail_data de cada arquivo.
# 4. Mapear campos bronze para o schema canonico silver com regras de extracao.
# 5. Calcular cobertura percentual de cada campo candidato na amostra.
# 6. Gerar relatorio de mapeamento em JSON ou Markdown.

import html
import re
import sys
import unicodedata
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parents[3]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.projects_standards.shared.silver import (  # noqa: E402
    CandidateSource,
    path_candidate,
    run_mapping,
    scalar_or_list,
)


DISPLAY_NAME = "TERO"
BRONZE_SLUG = "tero"
MAPPING_OUTPUT_PATH = CURRENT_DIR / "docs" / "silver_field_mapping.md"
COUNTRY_REFERENCE_PATH = ROOT_DIR / "data" / "project_standards" / "00_reference" / "reference_dataset.xlsx"


# Ordena os arquivos do bronze pela chave derivada do nome do arquivo.
def sort_key(path: Path) -> str:
    return path.stem


# Limpa HTML simples para texto plano util aos parsers locais.
def clean_html_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = html.unescape(str(value))
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


# Retorna o payload real do detalhe da TERO preservado abaixo de detail_data.api_response.
def get_api_response(payload: dict[str, Any]) -> dict[str, Any]:
    detail_data = payload.get("detail_data", {})
    api_response = detail_data.get("api_response")
    if isinstance(api_response, dict):
        return api_response
    return detail_data if isinstance(detail_data, dict) else {}


# Normaliza chaves textuais para comparacao tolerante contra a referencia de paises.
def normalize_country_match_key(value: Any) -> str | None:
    text = clean_html_text(value)
    if not text:
        return None
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^A-Za-z0-9]+", " ", text).strip().casefold()
    return text or None


# Carrega aliases de pais a partir da tabela padrao para permitir identificacao por nomes canonicos.
@lru_cache(maxsize=1)
def load_country_aliases() -> dict[str, str]:
    workbook = load_workbook(COUNTRY_REFERENCE_PATH, read_only=True, data_only=True)
    try:
        worksheet = workbook["pais_padrao"]
        aliases: dict[str, str] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            nome_pt, nome_en, nome_es, alpha_2, alpha_3, numeric = row[:6]
            english_name = clean_html_text(nome_en)
            if not english_name:
                continue
            for candidate in (nome_pt, nome_en, nome_es):
                key = normalize_country_match_key(candidate)
                if key:
                    aliases[key] = english_name
        return aliases
    finally:
        workbook.close()


# Tenta identificar um pais padrao em ingles a partir de um texto livre do projeto.
def extract_country_from_text(value: Any) -> str | None:
    text = clean_html_text(value)
    if not text:
        return None

    aliases = load_country_aliases()
    normalized_text = normalize_country_match_key(text)
    if normalized_text:
        for alias, english_name in aliases.items():
            if len(alias) >= 4 and re.search(rf"(?<![A-Za-z0-9]){re.escape(alias)}(?![A-Za-z0-9])", normalized_text):
                return english_name
    return None


# Extrai o HTML renderizado salvo no bronze real da TERO.
def extract_content(payload: dict[str, Any]) -> str:
    api_response = get_api_response(payload)
    return str(api_response.get("content", {}).get("rendered", "") or payload.get("list_data", {}).get("content", {}).get("rendered", ""))


# Coleta os blocos principais do resumo exibido na pagina para reuso nos parsers.
def extract_header_spans(payload: dict[str, Any]) -> list[str]:
    content = extract_content(payload)
    matches = re.findall(r"<h4 class=\"et_pb_module_header\">(?:<a [^>]*>)?<span>(.*?)</span>", content, flags=re.IGNORECASE | re.DOTALL)
    cleaned = []
    for match in matches:
        text = clean_html_text(match)
        if text and text not in cleaned:
            cleaned.append(text)
    return cleaned


# Extrai o codigo de metodologia TERO.00X exibido na pagina.
def extract_methodology_code(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        match = re.search(r"\bTERO\.\d+\b", text)
        if match:
            return match.group(0)
    return None


# Extrai a linha textual completa da metodologia exibida no resumo do projeto.
def extract_methodology(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        if re.search(r"\bTERO\.\d+\b", text):
            return text
    return None


# Extrai o status voluntario textual exibido no resumo da pagina.
def extract_status(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        if re.fullmatch(r"Finalizado|Em desenvolvimento|Em valida[c??][a??]o|Listado|Registrado", text, flags=re.IGNORECASE):
            return text
    return None


# Extrai a localizacao textual do projeto usando o resumo ou a descricao SEO.
def extract_location(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        if "," in text and not re.search(r"\d{2}/\d{2}/\d{4}", text):
            return text
    description = clean_html_text(get_api_response(payload).get("yoast_head_json", {}).get("description"))
    if description:
        match = re.search(r"localizado em ([^.]+)", description, flags=re.IGNORECASE)
        if match:
            location = re.split(r",\s+certificad[oa]|\s+certificad[oa]", match.group(1).strip(), maxsplit=1, flags=re.IGNORECASE)[0]
            return location.strip()
    return None


# Identifica o pais do projeto a partir da referencia padrao de paises do projeto.
def extract_country(payload: dict[str, Any], _: Path) -> Any:
    api_response = get_api_response(payload)
    candidates = [
        extract_location(payload, _),
        api_response.get("yoast_head_json", {}).get("description"),
        api_response.get("excerpt", {}).get("rendered"),
        extract_content(payload),
    ]
    for candidate in candidates:
        country = extract_country_from_text(candidate)
        if country:
            return country
    return None


# Extrai o estado ou regiao a partir da localizacao textual exibida pela pagina.
def extract_state(payload: dict[str, Any], _: Path) -> Any:
    location = extract_location(payload, _)
    if not location:
        return None
    parts = [part.strip() for part in location.split(",") if part.strip()]
    if len(parts) >= 2:
        return parts[-2]
    return None


# Extrai a cidade ou localidade a partir da localizacao textual exibida pela pagina.
def extract_city(payload: dict[str, Any], _: Path) -> Any:
    location = extract_location(payload, _)
    if not location:
        return None
    parts = [part.strip() for part in location.split(",") if part.strip()]
    if parts:
        return parts[0]
    return None


# Extrai o nome do desenvolvedor a partir do resumo do projeto.
def extract_project_developer(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        if re.search(r"\bLtda\b|Consultoria|Carbon|Agropecu[a??]ria", text, flags=re.IGNORECASE) and not re.search(r"\bTERO\.\d+\b", text):
            return text
    return None


# Extrai o setor principal do projeto a partir do resumo textual.
def extract_sector(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        if text.upper() in {"AFOLU"}:
            return text
    return None


# Extrai a categoria macro do projeto mostrada no resumo da pagina.
def extract_project_category(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        if "Solu????o Baseada" in text:
            return text
    return None


# Deriva subcategorias taxonomicas a partir das classes CSS presentes no payload da API.
def extract_project_subcategories(payload: dict[str, Any], _: Path) -> Any:
    values = []
    for class_name in get_api_response(payload).get("class_list", []):
        if isinstance(class_name, str) and class_name.startswith("project_category-"):
            value = class_name.replace("project_category-", "").replace("-", " ").strip()
            if value:
                values.append(value)
    return scalar_or_list(values)


# Extrai o contador total de creditos emitidos do bloco de impacto positivo.
def extract_credits_issued(payload: dict[str, Any], _: Path) -> Any:
    content = extract_content(payload)
    for number, title in re.findall(r'data-number-value="([\d\.]+)".{0,250}?<h3 class="title">([^<]+)</h3>', content, flags=re.IGNORECASE | re.DOTALL):
        clean_title = clean_html_text(title) or ""
        if "cr??ditos de carbono" in clean_title.lower() or "creditos de carbono" in clean_title.lower():
            return int(number.replace(".", ""))
    return None


# Monta as fontes candidatas do mapeamento canonicamente documentado da TERO.
def build_candidate_sources() -> dict[str, list[CandidateSource]]:
    return {
        "standard_name": [path_candidate("source", "carbon_standard", rule_type="rename")],
        "standard_acronym": [
            CandidateSource(
                source_section="reference",
                source_path="data/project_standards/00_reference/reference_dataset.xlsx (standards_catalog)",
                rule_type="lookup",
                notes="Deve ser obtido na referencia Certificadoras, a partir da certificadora do registro.",
                extractor=lambda payload, _file_path: "TER",
            )
        ],
        "project_public_id": [path_candidate("source", "project_public_id"), path_candidate("list_data", "slug")],
        "project_internal_id": [path_candidate("source", "project_internal_id"), path_candidate("list_data", "id")],
        "project_url": [path_candidate("source", "project_url"), path_candidate("list_data", "link")],
        "bronze_file_path": [path_candidate("file_system", "bronze_file_path", rule_type="derived", notes="Derivado do caminho do arquivo de detalhe no filesystem.")],
        "source_file_name": [path_candidate("file_system", "source_file_name", rule_type="derived", notes="Derivado do nome do arquivo de detalhe no filesystem.")],
        "project_name": [path_candidate("detail_data", "api_response.title.rendered"), path_candidate("list_data", "title.rendered")],
        "project_voluntary_status": [CandidateSource(source_section="detail_data", source_path="api_response.content.rendered", rule_type="parsed_html", notes="Extrai o status do bloco resumo no HTML renderizado da pagina.", extractor=extract_status)],
        "project_regulatory_status": [],
        "standard_program": [path_candidate("source", "carbon_standard", rule_type="rename")],
        "project_description": [path_candidate("detail_data", "api_response.yoast_head_json.description"), path_candidate("detail_data", "api_response.excerpt.rendered"), path_candidate("list_data", "excerpt.rendered")],
        "project_methodology": [CandidateSource(source_section="detail_data", source_path="api_response.content.rendered", rule_type="parsed_html", notes="Extrai a linha completa da metodologia do bloco resumo do HTML renderizado.", extractor=extract_methodology)],
        "project_type": [CandidateSource(source_section="detail_data", source_path="api_response.content.rendered", rule_type="parsed_html", notes="Usa a categoria de alto nivel Solucao Baseada na Natureza quando presente no resumo.", extractor=extract_project_category)],
        "sector": [CandidateSource(source_section="detail_data", source_path="api_response.content.rendered", rule_type="parsed_html", notes="Extrai o setor AFOLU do bloco resumo do HTML renderizado.", extractor=extract_sector)],
        "project_category": [CandidateSource(source_section="detail_data", source_path="api_response.content.rendered", rule_type="parsed_html", notes="Extrai a categoria macro do resumo do projeto no HTML renderizado.", extractor=extract_project_category)],
        "project_subcategories": [CandidateSource(source_section="detail_data", source_path="api_response.class_list", rule_type="normalized", notes="Deriva subcategorias taxonomicas a partir das classes CSS project_category-* do payload da API.", extractor=extract_project_subcategories)],
        "sdg_targets": [],
        "project_developer": [CandidateSource(source_section="detail_data", source_path="api_response.content.rendered", rule_type="parsed_html", notes="Extrai o nome da organizacao proponente do bloco resumo do HTML renderizado.", extractor=extract_project_developer)],
        "project_owner": [],
        "project_operator": [],
        "validator_name": [],
        "verifier_name": [],
        "country": [CandidateSource(source_section="detail_data", source_path="api_response.yoast_head_json.description", rule_type="reference_lookup", notes="Identifica o pais usando a localizacao textual e a tabela tb_paisPadrao, retornando o nome_en padrao.", extractor=extract_country)],
        "state_or_region": [CandidateSource(source_section="detail_data", source_path="api_response.content.rendered", rule_type="parsed_html", notes="Infere a unidade federativa a partir da localizacao textual exibida na pagina.", extractor=extract_state)],
        "city_or_locality": [CandidateSource(source_section="detail_data", source_path="api_response.content.rendered", rule_type="parsed_html", notes="Infere a cidade a partir da localizacao textual exibida na pagina.", extractor=extract_city)],
        "location_latitude": [],
        "location_longitude": [],
        "snapshot_date": [path_candidate("source", "snapshot_date")],
        "reference_month": [path_candidate("source", "reference_month")],
        "registration_date": [path_candidate("detail_data", "api_response.date_gmt"), path_candidate("list_data", "date_gmt")],
        "status_date": [path_candidate("detail_data", "api_response.modified_gmt"), path_candidate("list_data", "modified_gmt")],
        "crediting_start_date": [],
        "crediting_end_date": [],
        "first_issuance_date": [],
        "last_issuance_date": [],
        "credits_issued_total": [CandidateSource(source_section="detail_data", source_path="api_response.content.rendered", rule_type="parsed_html", notes="Extrai o contador de creditos emitidos do bloco impacto positivo no HTML renderizado.", extractor=extract_credits_issued)],
        "credits_retired_total": [],
        "credits_cancelled_total": [],
        "credits_buffer_total": [],
        "estimated_annual_emission_reductions": [],
        "estimated_total_emission_reductions": [],
        "area_hectares": [],
    }


CONFIG = {
    "display_name": DISPLAY_NAME,
    "bronze_slug": BRONZE_SLUG,
    "mapping_output_path": MAPPING_OUTPUT_PATH,
    "mapping_candidates": build_candidate_sources,
    "sort_key": sort_key,
}


if __name__ == "__main__":
    raise SystemExit(run_mapping(CONFIG))
