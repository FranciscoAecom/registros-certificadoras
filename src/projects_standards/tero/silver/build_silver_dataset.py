# Objetivo do script:
# Consolidar os arquivos bronze da TERO em um unico dataset JSON na camada silver seguindo o schema canonico do projeto.

# Processo:
# 1. Definir configuracao da certificadora (nome, slug bronze, sigla).
# 2. Construir dicionario de transformadores campo a campo (build_transformers).
# 3. Registrar hooks de pos-build para sincronizar referencias (status, pais, metodologia).
# 4. Delegar ao framework compartilhado run_dataset(CONFIG) que:
#    a. Descompacta automaticamente o bronze e o silver se estiverem zipados.
#    b. Carrega os arquivos bronze do snapshot informado.
#    c. Aplica cada transformador para extrair e normalizar campos.
#    d. Gera o dataset silver (allprojects.json).
#    e. Gera o relatorio de qualidade (quality_report.json).
#    f. Gera o relatorio de mapeamento (mapping_report.json).
#    g. Compacta novamente o bronze e o silver ao final da execução.
# 5. Executar hooks de sincronizacao de referencias.
# 6. Retornar codigo de saida.

import html
import re
import sys
import unicodedata
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook


CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parents[3]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.projects_standards.shared.silver import (  # noqa: E402
    first_non_empty,
    get_path,
    parse_date,
    path_candidate,
    run_dataset,
    scalar_or_list,
)
from sync_country_reference import sync_country_reference_for_projects  # noqa: E402
from sync_methodology_reference import sync_methodology_reference_for_projects  # noqa: E402
from sync_status_reference import sync_status_reference_for_projects  # noqa: E402


DISPLAY_NAME = "TERO"
BRONZE_SLUG = "tero"
DATASET_OUTPUT_TEMPLATE = ROOT_DIR / "data" / "project_standards" / "02_silver" / BRONZE_SLUG / "{date}" / "allprojects.json"
FAILURE_OUTPUT_TEMPLATE = CURRENT_DIR / "logs" / "build_silver_dataset_failures_{date}.json"
COUNTRY_REFERENCE_PATH = ROOT_DIR / "data" / "reference" / "reference_dataset.xlsx"


# Ordena os arquivos do bronze pela chave derivada do nome do arquivo.
def sort_key(path: Path) -> str:
    return path.stem


# Limpa HTML simples para texto plano util aos parsers locais.
def clean_html_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = html.unescape(str(value))
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


# Retorna o payload real do detalhe da TERO preservado abaixo de detail_data.api_response.
def get_api_response(payload: dict[str, Any]) -> dict[str, Any]:
    detail_data = payload.get("detail_data", {})
    api_response = detail_data.get("api_response")
    if isinstance(api_response, dict):
        return api_response
    return detail_data if isinstance(detail_data, dict) else {}


# Normaliza chaves textuais para comparacao tolerante contra a referencia de paises.
def normalize_country_match_key(value: Any) -> str | None:
    text = clean_html_text(value)
    if not text:
        return None
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^A-Za-z0-9]+", " ", text).strip().casefold()
    return text or None


# Carrega aliases de pais a partir da tabela padrao para permitir identificacao por nomes canonicos.
@lru_cache(maxsize=1)
def load_country_aliases() -> dict[str, str]:
    workbook = load_workbook(COUNTRY_REFERENCE_PATH, read_only=True, data_only=True)
    try:
        worksheet = workbook["countries_standard"]
        aliases: dict[str, str] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            name_pt, name_en, name_es, alpha_2, alpha_3, numeric = row[:6]
            english_name = clean_html_text(name_en)
            if not english_name:
                continue
            for candidate in (name_pt, name_en, name_es):
                key = normalize_country_match_key(candidate)
                if key:
                    aliases[key] = english_name
        return aliases
    finally:
        workbook.close()


# Tenta identificar um pais padrao em ingles a partir de um texto livre do projeto.
def extract_country_from_text(value: Any) -> str | None:
    text = clean_html_text(value)
    if not text:
        return None

    aliases = load_country_aliases()
    normalized_text = normalize_country_match_key(text)
    if normalized_text:
        for alias, english_name in aliases.items():
            if len(alias) >= 4 and re.search(rf"(?<![A-Za-z0-9]){re.escape(alias)}(?![A-Za-z0-9])", normalized_text):
                return english_name
    return None


# Extrai o HTML renderizado salvo no bronze real da TERO.
def extract_content(payload: dict[str, Any]) -> str:
    api_response = get_api_response(payload)
    return str(api_response.get("content", {}).get("rendered", "") or payload.get("list_data", {}).get("content", {}).get("rendered", ""))


# Coleta os blocos principais do resumo exibido na pagina para reuso nos parsers.
def extract_header_spans(payload: dict[str, Any]) -> list[str]:
    content = extract_content(payload)
    matches = re.findall(r"<h4 class=\"et_pb_module_header\">(?:<a [^>]*>)?<span>(.*?)</span>", content, flags=re.IGNORECASE | re.DOTALL)
    cleaned = []
    for match in matches:
        text = clean_html_text(match)
        if text and text not in cleaned:
            cleaned.append(text)
    return cleaned


# Extrai o codigo de metodologia TERO.00X exibido na pagina.
def extract_methodology_code(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        match = re.search(r"\bTERO\.\d+\b", text)
        if match:
            return match.group(0)
    return None


# Extrai a linha textual completa da metodologia exibida no resumo do projeto.
def extract_methodology(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        if re.search(r"\bTERO\.\d+\b", text):
            return text
    return None


# Extrai o status voluntario textual exibido no resumo da pagina.
def extract_status(payload: dict[str, Any], _: Path) -> Any:
    canonical_status = {
        "finalizado": "Finalizado",
        "em desenvolvimento": "Em desenvolvimento",
        "em validação": "Em validação",
        "em validacao": "Em validação",
        "listado": "Listado",
        "registrado": "Registrado",
    }
    candidates = extract_header_spans(payload)
    candidates.append(clean_html_text(extract_content(payload)))
    for text in candidates:
        if not text:
            continue
        match = re.search(r"\b(Finalizado|Em desenvolvimento|Em validação|Em validacao|Listado|Registrado)\b", text, flags=re.IGNORECASE)
        if match:
            return canonical_status.get(match.group(1).casefold(), match.group(1))
    return None


# Extrai a localizacao textual do projeto usando o resumo ou a descricao SEO.
def extract_location(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        if "," in text and not re.search(r"\d{2}/\d{2}/\d{4}", text):
            return text
    description = clean_html_text(get_api_response(payload).get("yoast_head_json", {}).get("description"))
    if description:
        match = re.search(r"localizado em ([^.]+)", description, flags=re.IGNORECASE)
        if match:
            location = re.split(r",\s+certificad[oa]|\s+certificad[oa]", match.group(1).strip(), maxsplit=1, flags=re.IGNORECASE)[0]
            return location.strip()
    return None


# Identifica o pais do projeto a partir da referencia padrao de paises do projeto.
def extract_country(payload: dict[str, Any], file_path: Path) -> Any:
    api_response = get_api_response(payload)
    candidates = [
        extract_location(payload, file_path),
        api_response.get("yoast_head_json", {}).get("description"),
        api_response.get("excerpt", {}).get("rendered"),
        extract_content(payload),
    ]
    for candidate in candidates:
        country = extract_country_from_text(candidate)
        if country:
            return country
    return None


# Extrai o estado ou regiao a partir da localizacao textual exibida pela pagina.
def extract_state(payload: dict[str, Any], file_path: Path) -> Any:
    location = extract_location(payload, file_path)
    if not location:
        return None
    parts = [part.strip() for part in location.split(",") if part.strip()]
    if len(parts) >= 2:
        return parts[-2]
    return None


# Extrai a cidade ou localidade a partir da localizacao textual exibida pela pagina.
def extract_city(payload: dict[str, Any], file_path: Path) -> Any:
    location = extract_location(payload, file_path)
    if not location:
        return None
    parts = [part.strip() for part in location.split(",") if part.strip()]
    return parts[0] if parts else None


# Extrai o nome do desenvolvedor a partir do resumo do projeto.
def extract_project_developer(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        if re.search(r"\bLtda\b|Consultoria|Carbon|Agropecu[a??]ria", text, flags=re.IGNORECASE) and not re.search(r"\bTERO\.\d+\b", text):
            return text
    return None


# Extrai o setor principal do projeto a partir do resumo textual.
def extract_sector(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        if text.upper() == "AFOLU":
            return text
    return None


# Extrai a categoria macro do projeto mostrada no resumo da pagina.
def extract_project_category(payload: dict[str, Any], _: Path) -> Any:
    for text in extract_header_spans(payload):
        if "Solu????o Baseada" in text:
            return text
    return None


# Deriva subcategorias taxonomicas a partir das classes CSS presentes no payload da API.
def extract_project_subcategories(payload: dict[str, Any], _: Path) -> Any:
    values = []
    for class_name in get_api_response(payload).get("class_list", []):
        if isinstance(class_name, str) and class_name.startswith("project_category-"):
            value = class_name.replace("project_category-", "").replace("-", " ").strip()
            if value:
                values.append(value)
    return scalar_or_list(values)


# Extrai o contador total de creditos emitidos do bloco de impacto positivo.
def extract_credits_issued(payload: dict[str, Any], _: Path) -> Any:
    content = extract_content(payload)
    for number, title in re.findall(r'data-number-value="([\d\.]+)".{0,250}?<h3 class="title">([^<]+)</h3>', content, flags=re.IGNORECASE | re.DOTALL):
        clean_title = clean_html_text(title) or ""
        if "cr??ditos de carbono" in clean_title.lower() or "creditos de carbono" in clean_title.lower():
            return int(number.replace(".", ""))
    return None


# Monta os transformadores canonicos do dataset silver da TERO.
def build_transformers(standard_acronym: str | None) -> dict[str, Callable[[dict[str, Any], Path], Any]]:
    return {
        "standard_name": lambda payload, _file_path: get_path(payload, "source.carbon_standard"),
        "standard_acronym": lambda payload, _file_path: standard_acronym,
        "project_public_id": lambda payload, _file_path: first_non_empty(payload, "source.project_public_id", "list_data.slug"),
        "project_internal_id": lambda payload, _file_path: first_non_empty(payload, "source.project_internal_id", "list_data.id"),
        "project_url": lambda payload, _file_path: first_non_empty(payload, "source.project_url", "list_data.link"),
        "bronze_file_path": lambda payload, file_path: path_candidate("file_system", "bronze_file_path").extractor(payload, file_path),
        "source_file_name": lambda payload, file_path: path_candidate("file_system", "source_file_name").extractor(payload, file_path),
        "project_name": lambda payload, _file_path: first_non_empty(payload, "detail_data.api_response.title.rendered", "list_data.title.rendered"),
        "project_voluntary_status": extract_status,
        "project_regulatory_status": lambda payload, _file_path: None,
        "standard_program": lambda payload, _file_path: get_path(payload, "source.carbon_standard"),
        "project_description": lambda payload, _file_path: clean_html_text(first_non_empty(payload, "detail_data.api_response.yoast_head_json.description", "detail_data.api_response.excerpt.rendered", "list_data.excerpt.rendered")),
        "project_methodology": extract_methodology,
        "project_type": extract_project_category,
        "sector": extract_sector,
        "project_category": extract_project_category,
        "project_subcategories": extract_project_subcategories,
        "sdg_targets": lambda payload, _file_path: None,
        "project_developer": extract_project_developer,
        "project_owner": lambda payload, _file_path: None,
        "project_operator": lambda payload, _file_path: None,
        "validator_name": lambda payload, _file_path: None,
        "verifier_name": lambda payload, _file_path: None,
        "country": extract_country,
        "state_or_region": extract_state,
        "city_or_locality": extract_city,
        "location_latitude": lambda payload, _file_path: None,
        "location_longitude": lambda payload, _file_path: None,
        "snapshot_date": lambda payload, _file_path: get_path(payload, "source.snapshot_date"),
        "reference_month": lambda payload, _file_path: get_path(payload, "source.reference_month"),
        "registration_date": lambda payload, _file_path: parse_date(first_non_empty(payload, "detail_data.api_response.date_gmt", "list_data.date_gmt")),
        "status_date": lambda payload, _file_path: parse_date(first_non_empty(payload, "detail_data.api_response.modified_gmt", "list_data.modified_gmt")),
        "crediting_start_date": lambda payload, _file_path: None,
        "crediting_end_date": lambda payload, _file_path: None,
        "first_issuance_date": lambda payload, _file_path: None,
        "last_issuance_date": lambda payload, _file_path: None,
        "credits_issued_total": extract_credits_issued,
        "credits_retired_total": lambda payload, _file_path: None,
        "credits_cancelled_total": lambda payload, _file_path: None,
        "credits_buffer_total": lambda payload, _file_path: None,
        "estimated_annual_emission_reductions": lambda payload, _file_path: None,
        "estimated_total_emission_reductions": lambda payload, _file_path: None,
        "area_hectares": lambda payload, _file_path: None,
    }


CONFIG = {
    "display_name": DISPLAY_NAME,
    "bronze_slug": BRONZE_SLUG,
    "reference_name": DISPLAY_NAME,
    "dataset_output_template": DATASET_OUTPUT_TEMPLATE,
    "failure_output_template": FAILURE_OUTPUT_TEMPLATE,
    "transformers": build_transformers,
    "sort_key": sort_key,
    "post_build_hooks": [sync_status_reference_for_projects, sync_country_reference_for_projects, sync_methodology_reference_for_projects],
}


if __name__ == "__main__":
    raise SystemExit(run_dataset(CONFIG))
